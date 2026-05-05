import os
import io
import csv
import random
import hashlib
import json
import tempfile
from datetime import date, datetime, timedelta
from typing import cast
from urllib import request
import subprocess

# Django / stdlib
from django.conf import settings
from django.core.exceptions import PermissionDenied
from django.core.cache import cache
from django.db import transaction
from django.db.models import Count, Min, Q
from django.shortcuts import render, redirect, get_object_or_404
from django.http import FileResponse, Http404, HttpResponse, JsonResponse
from django.urls import reverse
from django.template.loader import render_to_string
from django.utils import timezone
from django.utils.timezone import now
from django.contrib import messages
from django.contrib.auth import login as auth_login, logout, get_user_model
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.auth.views import LoginView
from django.contrib.admin.views.decorators import staff_member_required
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_http_methods
from django.views import View

# Third-party
from weasyprint import HTML
from pypdf import PdfWriter, PdfReader
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from gtts import gTTS
from captcha.models import CaptchaStore, logger
from deep_translator import GoogleTranslator
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status

from website import urls

# Local App Imports
from .utils import (
    load_employee_data, send_system_email, get_allowed_quarters,
    ensure_current_financial_year
)
from .employeeform import EmployeeForm
from .forms import (
    CertificateDataForm, CustomLoginForm,
    CustomUserCreationForm, TypingUsageReportForm
)
from .models import (
    ArchivedUser, CertificateData, CustomUser, DataAccessLog,
    EditRequest, Employee, FinancialYear, HindiPost, ManagerRequest, Office,
    QPRPartTwo, QPRRecord, Role, Section1FilesData, Section2MeetingsData,
    Section3OfficialLanguagesData, Section4HindiLettersData,
    Section5EnglishRepliedHindiData, Section6IssuedLettersData,
    Section7NotingsData, Section8WorkshopsData,
    Section9ImplementationCommitteeData, Section10HindiAdvisoryData,
    Section11SpecificAchievementsData, StaffHindiKnowledge, TranslationKnowledge, TypingStenographyKnowledge,
    TypingUsageReport, UserProfile, cipher_suite, ProfileChangeRequest,
    ManagerRequest, EditRequest
)
from .serializers import EmployeeSerializer
from .signals import User
from .static_event_service import (
    delete_event, get_all_events, update_event_meta,
    upload_event, upload_images_to_existing_event
)

# Font Registration
FONT_PATH = os.path.join(settings.BASE_DIR, 'static', 'fonts', 'NIRMALA.TTF')
if os.path.exists(FONT_PATH):
    pdfmetrics.registerFont(TTFont('HindiFont', FONT_PATH))
from .templatetags.translate_tags import translate_text
from .utils import (
    ensure_current_financial_year, get_allowed_quarters, 
    load_employee_data, send_system_email
)

# Font Registration
FONT_PATH = os.path.join(settings.BASE_DIR, 'static', 'fonts', 'NIRMALA.TTF')
pdfmetrics.registerFont(TTFont('HindiFont', FONT_PATH))
from django.contrib.admin.views.decorators import staff_member_required
from django.shortcuts import render, redirect
from django.http import JsonResponse
from django.contrib import messages
from .utils import load_employee_data
from .utils import ensure_current_financial_year
from .models import FinancialYear
from django.http import JsonResponse
import csv
import hashlib
import io
import json
import os
import random
import tempfile
from datetime import date, datetime, timedelta
from typing import cast
from urllib import request

# Third-party / Django Imports
from captcha.models import CaptchaStore, logger
from deep_translator import GoogleTranslator
from django.conf import settings
from django.contrib import messages
from django.contrib.admin.views.decorators import staff_member_required
from django.contrib.auth import get_user_model, logout
from django.contrib.auth import login as auth_login
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.auth.views import LoginView
from django.core.cache import cache
from django.core.exceptions import PermissionDenied
from django.db.models import Count, Min, Q
from django.http import FileResponse, Http404, HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.template.loader import render_to_string
from django.urls import reverse
from django.utils import timezone
from django.utils.timezone import now
from django.views import View
from django.views.decorators.csrf import csrf_exempt
from gtts import gTTS
from pypdf import PdfReader, PdfWriter
from reportlab.lib.pagesizes import letter
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.pdfgen import canvas
from rest_framework import status
from rest_framework.response import Response
from rest_framework.views import APIView
from weasyprint import HTML

# Local App Imports
from .employeeform import EmployeeForm
from .forms import (
    CertificateDataForm, CustomLoginForm, 
    CustomUserCreationForm, TypingUsageReportForm
)
from .models import (
    ArchivedUser, CertificateData, CustomUser, DataAccessLog, 
    EditRequest, Employee, HindiPost, ManagerRequest, Office, 
    QPRPartTwo, QPRRecord, Role, Section1FilesData, Section2MeetingsData, 
    Section3OfficialLanguagesData, Section4HindiLettersData, 
    Section5EnglishRepliedHindiData, Section6IssuedLettersData, 
    Section7NotingsData, Section8WorkshopsData, 
    Section9ImplementationCommitteeData, Section10HindiAdvisoryData, 
    Section11SpecificAchievementsData, StaffHindiKnowledge, 
    TypingUsageReport, UserProfile, cipher_suite
)
from website.models import CodeManualStandardForms, HindiPost, WebsiteDetail, OfficersWorkInHindi   
from .serializers import EmployeeSerializer
from .templatetags.translate_tags import translate_text
from .utils import get_allowed_quarters, load_employee_data, send_system_email
from .signals import User
if os.path.exists(FONT_PATH):
    pdfmetrics.registerFont(TTFont('HindiFont', FONT_PATH))

def get_employee_details_form(request):
    if request.method == "POST":
        empcode = request.POST.get('empcode', '').strip()
        
        if not empcode:
            return JsonResponse({'status': 'error', 'message': 'Employee code required'})
        
        try:
            import openpyxl
            import os
            from django.conf import settings
            from django.http import JsonResponse
            
            # Load Excel file
            excel_file = os.path.join(settings.MEDIA_ROOT, 'data', 'tg_hod_officers_employee_report.xlsx')
            
            if not os.path.exists(excel_file):
                return JsonResponse({'status': 'error', 'message': 'Employee database file not found'})
            
            wb = openpyxl.load_workbook(excel_file)
            ws = wb.active
            
            headers = []
            for cell in ws[1]:
                headers.append(cell.value)
            
            found = False
            row_data = {}
            
            for row in range(2, ws.max_row + 1):
                row_data = {}
                for col_idx, header in enumerate(headers, 1):
                    row_data[header] = ws.cell(row=row, column=col_idx).value
                
                if str(row_data.get('Empcode', '')).strip() == str(empcode).strip():
                    found = True
                    break
            
            if not found:
                return JsonResponse({'status': 'error', 'message': 'Invalid Employee Code'})
            
            # Return employee data
            return JsonResponse({
                'status': 'success',
                'name': row_data.get('Name', '') or '',
                'mobile': row_data.get('Mobile', '') or '',
                'ip_number': row_data.get('IP Number', '') or '',
                'state': row_data.get('State', '') or '',
                'hindi_name': row_data.get('Name in Hindi', '') or '',
                'designation': row_data.get('Designation', '') or '',
                'email': '',
            })
        
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': f'Error: {str(e)}'})
    
    return JsonResponse({'status': 'error', 'message': 'Invalid request'})

@login_required
@require_http_methods(["POST"])
def submit_profile_change_request(request):
    try:
        data = json.loads(request.body)
        reason = data.get('change_reason', '').strip()
        allowed_fields = {'alternate_email', 'designation', 'highest_exam'}
        requested_fields = data.get('requested_fields') or []
        requested_fields = [
            field for field in requested_fields
            if isinstance(field, str) and field in allowed_fields
        ]

        if not reason:
            return JsonResponse({'success': False, 'message': 'Reason is required'})

        if not requested_fields:
            return JsonResponse({'success': False, 'message': 'Please select at least one field to edit.'})

        profile = getattr(request.user, 'profile', None)
        if (
            profile is None
            or not profile.profile_updated
            or profile.approval_status != 'approved'
            or request.user.is_edit_allowed
        ):
            return JsonResponse({
                'success': False,
                'message': 'Profile change requests are only allowed for locked, approved profiles.'
            }, status=403)

        hod_identifier = (profile.hod_name or "").strip()

        if not hod_identifier:
            return JsonResponse({
                'success': False,
                'message': 'HOD is not assigned to your profile.'
            })

        # PREVENT DUPLICATE REQUESTS
        existing_request = ProfileChangeRequest.objects.filter(
            profile=profile,
            status='pending'
        ).first()

        if existing_request:
            return JsonResponse({
                'success': False,
                'message': 'You already have a pending request. Please wait for approval.'
            })

        # 🔍 Find HOD (same logic, untouched)
        hod_profile = UserProfile.objects.filter(
            Q(roles__name='hod') | Q(user__roles__name='hod'),
            Q(employee_code=hod_identifier) |
            Q(name__iexact=hod_identifier) |
            Q(hod_name__iexact=hod_identifier) |
            Q(user__username__iexact=hod_identifier)
        ).distinct().first()

        if not hod_profile:
            return JsonResponse({
                'success': False,
                'message': f'HOD "{hod_identifier}" not found in system. Please ensure your HOD has registered and is approved.'
            })

        # CREATE REQUEST (unchanged logic)
        ProfileChangeRequest.objects.create(
            profile=profile,
            change_reason=reason,
            requested_fields=requested_fields,
            hod=hod_profile.user,
            status='pending'
        )

        # UPDATE PROFILE STATUS (unchanged)
        # profile.approval_status = 'change_pending'
        # profile.save(update_fields=['approval_status'])

        return JsonResponse({
            'success': True,
            'message': 'Change request submitted successfully. Awaiting HOD approval.'
        })

    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'message': 'Invalid JSON data'})

    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)})
from django.http import JsonResponse
from website.static_event_service import get_all_events

def can_manage_events(user):
    if not user or not user.is_authenticated:
        return False

    if user.is_staff or user.is_superuser:
        return True

    profile = getattr(user, 'profile', None)
    return (
        user.roles.filter(name__in=['manager', 'admin']).exists()
        or (profile and profile.roles.filter(name__in=['manager', 'admin']).exists())
    )

def require_event_manager(user):
    if not can_manage_events(user):
        raise PermissionDenied

def get_event_images(request, folder):
    """Get event images (non-API version)"""
    require_event_manager(request.user)

    if request.method != 'POST':
        return JsonResponse({'images': []}, status=400)
    
    try:
        from website.static_event_service import get_all_events
        events = get_all_events()
        event = next((e for e in events if e['folder'] == folder), None)
        
        if not event:
            return JsonResponse({'images': []})
        
        return JsonResponse({'images': event['images']})
    except Exception as e:
        return JsonResponse({'images': [], 'error': str(e)})

def update_event_titles(request):
    """Update event titles"""
    require_event_manager(request.user)

    if request.method != 'POST':
        return redirect('admin_events_dashboard')
    
    folder = request.POST.get('folder')
    title_en = request.POST.get('title_en')
    title_hi = request.POST.get('title_hi')
    
    from website.static_event_service import update_event_meta
    try:
        update_event_meta(folder, title_en, title_hi)
        messages.success(request, "Titles updated")
    except Exception as e:
        messages.error(request, str(e))
    
    return redirect('admin_events_dashboard')



@login_required
def admin_events_dashboard(request):
    require_event_manager(request.user)

    events = get_all_events()
    return render(request, "admin_events_dashboard.html", {"events": events})
 
 
@login_required
def admin_upload_event(request):
    require_event_manager(request.user)

    folder = request.GET.get("folder")
 
    if request.method == "POST":
        event_date   = request.POST.get("event_date")
        event_name   = request.POST.get("event_name")
        event_name_hi = request.POST.get("event_name_hi", "")
        images       = request.FILES.getlist("images")
 
        try:
            if folder:
                upload_images_to_existing_event(folder, images)
            else:
                upload_event(event_date, event_name, event_name_hi, images)
 
            return JsonResponse({"status": "success"})
 
        except Exception as e:
            return JsonResponse({"status": "error", "message": str(e)})
 
    return render(request, "admin_upload_event.html", {"folder": folder})
 
 
@login_required
def admin_delete_event(request, folder):
    require_event_manager(request.user)

    try:
        delete_event(folder)
        messages.success(request, "Event deleted successfully")
    except Exception as e:
        messages.error(request, f"Failed to delete event: {e}")
    return redirect("admin_events_dashboard")
 
 
@login_required
def admin_edit_event_titles(request):
    require_event_manager(request.user)

    """AJAX endpoint — update title_en and title_hi in an event's meta.json"""
    if request.method == "POST":
        try:
            data     = json.loads(request.body)
            folder   = data.get("folder", "").strip()
            title_en = data.get("title_en", "").strip()
            title_hi = data.get("title_hi", "").strip()
 
            if not folder or not title_en:
                return JsonResponse({"status": "error", "message": "folder and title_en are required"})
 
            update_event_meta(folder, title_en, title_hi)
            return JsonResponse({"status": "success"})
 
        except Exception as e:
            return JsonResponse({"status": "error", "message": str(e)})
 
    return JsonResponse({"status": "error", "message": "POST only"})
from website.static_event_service import update_event_meta, _read_meta

@login_required
def set_thumbnail(request, folder):
    require_event_manager(request.user)

    if request.method != 'POST':
        return JsonResponse({'status': 'error', 'message': 'POST required'}, status=400)
    
    try:
        data = json.loads(request.body)
        thumbnail = data.get('thumbnail')
        
        if not thumbnail:
            return JsonResponse({'status': 'error', 'message': 'Thumbnail filename required'})
        
        from website.static_event_service import update_event_meta
        update_event_meta(folder, None, None, thumbnail=thumbnail)
        
        return JsonResponse({'status': 'success'})
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)})
    
# Helper functions to safely access a user's roles for type-checkers
def user_has_role(user, role_name):
    """Return True if user (or their profile) has the given role or any in the list."""
    profile = getattr(user, 'profile', None)
    if isinstance(role_name, (list, tuple)):
        user_has = user.roles.filter(name__in=role_name).exists()
        profile_has = profile.roles.filter(name__in=role_name).exists() if profile else False
        return user_has or profile_has
    else:
        user_has = user.roles.filter(name=role_name).exists()
        profile_has = profile.roles.filter(name=role_name).exists() if profile else False
        return user_has or profile_has

def user_role(user):
    """Return user's primary role (for backward compatibility)
    Returns the first role from: admin > manager > hod > user > None"""
    if user is None or not user.is_authenticated:
        return None
    
    priority_roles = ['admin', 'manager', 'hod', 'user', 'backup_user']
    profile = getattr(user, 'profile', None)
    for role in priority_roles:
        if user.roles.filter(name=role).exists():
            return role
        if profile and profile.roles.filter(name=role).exists():
            return role
    return None

def user_get_all_roles(user):
    """Get all role names as a list"""
    if user is None or not user.is_authenticated:
        return []
    profile = getattr(user, 'profile', None)
    user_roles = set(user.roles.values_list('name', flat=True))
    profile_roles = set(profile.roles.values_list('name', flat=True)) if profile else set()
    return list(sorted(user_roles.union(profile_roles)))

def is_admin(user):
    """Check if user is an admin"""
    return user.is_authenticated and user_has_role(user, 'admin')

def can_access_user_site(user):
    """User site accessible to all authenticated users (everyone starts as 'user')"""
    return user.is_authenticated and user_has_role(user, 'user')

def can_access_hod_site(user):
    """HOD site accessible to users with 'hod' role"""
    return user.is_authenticated and user_has_role(user, 'hod')

def can_access_manager_site(user):
    """Manager site accessible to users with 'manager' role"""
    return user.is_authenticated and user_has_role(user, 'manager')


def get_active_hods(office_code=None):
    """
    Returns a list of HOD names/usernames. 
    Matches office_code if provided, otherwise returns all HODs.
    """
    hod_query = UserProfile.objects.filter(Q(roles__name='hod') | Q(user__roles__name='hod'))
    hod_names = lambda qs: list(
        qs.exclude(user__username__isnull=True)
          .exclude(user__username='')
          .values_list('user__username', flat=True)
          .order_by('user__username')
          .distinct()
    )
    
    if office_code:
        # Try to find HODs in same office, but fallback to all HODs if none found in that office
        specific_hods = hod_names(hod_query.filter(office_code=office_code))
        if specific_hods:
            return specific_hods
            
    return hod_names(hod_query)

def _convert_to_int(value):
    if value == '' or value is None: return None
    try: return int(value)
    except (ValueError, TypeError): return None

def _convert_to_date(value):
    if value == '' or value is None: return None
    try:
        if isinstance(value, str): return datetime.fromisoformat(value).date()
        return value
    except (ValueError, TypeError, AttributeError): return None

def get_current_quarter():
    m = date.today().month

    if m <= 3:
        return "31 मार्च / Mar 31"
    elif m <= 6:
        return "30 जून / Jun 30"
    elif m <= 9:
        return "30 सितंबर / Sep 30"
    else:
        return "31 दिसंबर / Dec 31"


def get_current_year_label():
    today = date.today()
    # Financial year runs from Apr 1 -> Mar 31. If current month is April or later,
    # the fiscal year starts this calendar year; otherwise it started last calendar year.
    if today.month >= 4:
        start = today.year
    else:
        start = today.year - 1
    return f"{start}-{start+1}"


def get_quarter_end_dates():
    """
    Returns a dictionary with quarter end dates for current and upcoming quarters.
    {
        'current': date of quarterend for current quarter,
        'next': date of quarter end for next quarter
    }
    """
    today = date.today()
    month = today.month
    year = today.year
    
    if month <= 3:
        current_end = date(year, 3, 31)
        next_end = date(year, 6, 30)
    elif month <= 6:
        current_end = date(year, 6, 30)
        next_end = date(year, 9, 30)
    elif month <= 9:
        current_end = date(year, 9, 30)
        next_end = date(year, 12, 31)
    else:
        current_end = date(year, 12, 31)
        next_end = date(year + 1, 3, 31)
    
    return {'current': current_end, 'next': next_end}


def get_base_year(year_label):
    return int(year_label.split("-")[0])

def _save_section_data(record, details):
    # Section 1
    s1, _ = Section1FilesData.objects.get_or_create(qpr_record=record)
    s1.total_files = _convert_to_int(details.get('s1_total'))
    s1.hindi_files = _convert_to_int(details.get('s1_hindi'))
    s1.save()
    # Section 2
    s2, _ = Section2MeetingsData.objects.get_or_create(qpr_record=record)
    s2.meetings_count = _convert_to_int(details.get('s2_meetings'))
    s2.hindi_minutes = _convert_to_int(details.get('s2_minutes'))
    s2.total_papers = _convert_to_int(details.get('s2_papers_total'))
    s2.hindi_papers = _convert_to_int(details.get('s2_papers_hindi'))
    s2.save()
    # Section 3
    s3, _ = Section3OfficialLanguagesData.objects.get_or_create(qpr_record=record)
    s3.total_documents = _convert_to_int(details.get('s3_total'))
    s3.bilingual_documents = _convert_to_int(details.get('s3_bilingual'))
    s3.english_only_documents = _convert_to_int(details.get('s3_english'))
    s3.hindi_only_documents = _convert_to_int(details.get('s3_hindi_only'))
    s3.save()
    # Section 4
    s4, _ = Section4HindiLettersData.objects.get_or_create(qpr_record=record)
    s4.total_letters = _convert_to_int(details.get('s4_total'))
    s4.no_reply_letters = _convert_to_int(details.get('s4_no_reply'))
    s4.replied_hindi_letters = _convert_to_int(details.get('s4_replied_hindi'))
    s4.replied_english_letters = _convert_to_int(details.get('s4_replied_eng'))
    s4.save()
    # Section 5
    s5, _ = Section5EnglishRepliedHindiData.objects.get_or_create(qpr_record=record)
    s5.region_a_english_letters = _convert_to_int(details.get('s5_total'))
    s5.region_a_replied_hindi = _convert_to_int(details.get('s5_hindi'))
    s5.region_a_replied_english = _convert_to_int(details.get('s5_english'))
    s5.region_a_no_reply = _convert_to_int(details.get('s5_noreply'))
    s5.save()
    # Section 6
    s6, _ = Section6IssuedLettersData.objects.get_or_create(qpr_record=record)
    s6.region_a_hindi_bilingual = _convert_to_int(details.get('s6_a_hindi'))
    s6.region_a_english_only = _convert_to_int(details.get('s6_a_eng'))
    s6.region_a_total = _convert_to_int(details.get('s6_a_total'))
    s6.region_b_hindi_bilingual = _convert_to_int(details.get('s6_b_hindi'))
    s6.region_b_english_only = _convert_to_int(details.get('s6_b_eng'))
    s6.region_b_total = _convert_to_int(details.get('s6_b_total'))
    s6.region_c_hindi_bilingual = _convert_to_int(details.get('s6_c_hindi'))
    s6.region_c_english_only = _convert_to_int(details.get('s6_c_eng'))
    s6.region_c_total = _convert_to_int(details.get('s6_c_total'))
    s6.save()
    # Section 7
    s7, _ = Section7NotingsData.objects.get_or_create(qpr_record=record)
    s7.hindi_pages = _convert_to_int(details.get('s7_hindi'))
    s7.english_pages = _convert_to_int(details.get('s7_eng'))
    s7.total_pages = _convert_to_int(details.get('s7_total'))
    s7.eoffice_notings = _convert_to_int(details.get('s7_eoffice'))
    s7.save()
    # Section 8
    s8, _ = Section8WorkshopsData.objects.get_or_create(qpr_record=record)
    s8.full_day_workshops = _convert_to_int(details.get('s8_workshops'))
    s8.officers_trained = _convert_to_int(details.get('s8_officers'))
    s8.employees_trained = _convert_to_int(details.get('s8_employees'))
    s8.save()
    # Section 9
    s9, _ = Section9ImplementationCommitteeData.objects.get_or_create(qpr_record=record)
    s9.meeting_date = _convert_to_date(details.get('s9_date'))
    s9.sub_committees_count = _convert_to_int(details.get('s9_sub_committees'))
    s9.meetings_organized = _convert_to_int(details.get('s9_meetings_count'))
    s9.agenda_hindi = details.get('s9_agenda_hindi', '')
    s9.save()
    # Section 10
    s10, _ = Section10HindiAdvisoryData.objects.get_or_create(qpr_record=record)
    s10.meeting_date = _convert_to_date(details.get('s10_date'))
    s10.save()
    # Section 11
    s11, _ = Section11SpecificAchievementsData.objects.get_or_create(qpr_record=record)
    s11.innovative_work = details.get('s12_1', '')
    s11.special_events = details.get('s12_2', '')
    s11.hindi_medium_works = details.get('s12_3', '')
    s11.save()

def _quarter_label_to_daterange(quarter_label, year_label):
    """Return (start_date, end_date) for given quarter label and fiscal year label like '2025-2026'"""
    try:
        base = get_base_year(year_label)
    except Exception:
        base = date.today().year
    q = (quarter_label or '').strip()
    # Apr-Jun
    if 'Jun' in q or 'जून' in q:
        start = date(base, 4, 1)
        end = date(base, 6, 30)
    # Jul-Sep
    elif 'Sep' in q or 'सितंबर' in q or 'सित' in q:
        start = date(base, 7, 1)
        end = date(base, 9, 30)
    # Oct-Dec
    elif 'Dec' in q or 'दिसंबर' in q or 'दिस' in q:
        start = date(base, 10, 1)
        end = date(base, 12, 31)
    # Jan-Mar
    else:
        # This quarter belongs to next calendar year
        start = date(base+1, 1, 1)
        end = date(base+1, 3, 31)
    return (start, end)

NUMERIC_KEYS = [
    's1_total','s1_hindi','s2_meetings','s2_minutes','s2_papers_total','s2_papers_hindi',
    's3_total','s3_bilingual','s3_english','s3_hindi_only',
    's4_total','s4_no_reply','s4_replied_hindi','s4_replied_eng',
    's5_total','s5_hindi','s5_english','s5_noreply',
    's6_a_hindi','s6_a_eng','s6_a_total','s6_b_hindi','s6_b_eng','s6_b_total','s6_c_hindi','s6_c_eng','s6_c_total',
    's7_hindi','s7_eng','s7_total','s7_eoffice',
    's8_workshops','s8_officers','s8_employees'
]


def _serialize_managerqpr(m):
    """Map a ManagerQPR instance to NUMERIC_KEYS-shaped dict."""
    out = {k: 0 for k in NUMERIC_KEYS}
    if not m:
        return out
    try:
        out['s2_meetings'] = int(getattr(m, 's2_meetings_count', 0) or 0)
        out['s2_minutes'] = int(getattr(m, 's2_hindi_minutes', 0) or 0)
        out['s2_papers_total'] = int(getattr(m, 's2_total_papers', 0) or 0)
        out['s2_papers_hindi'] = int(getattr(m, 's2_hindi_papers', 0) or 0)

        out['s4_total'] = int(getattr(m, 's4_total_letters', 0) or 0)
        out['s4_no_reply'] = int(getattr(m, 's4_no_reply_letters', 0) or 0)
        out['s4_replied_hindi'] = int(getattr(m, 's4_replied_hindi_letters', 0) or 0)
        out['s4_replied_eng'] = int(getattr(m, 's4_replied_english_letters', 0) or 0)

        out['s5_total'] = int(getattr(m, 's5_region_a_english_letters', 0) or 0)
        out['s5_hindi'] = int(getattr(m, 's5_region_a_replied_hindi', 0) or 0)
        out['s5_english'] = int(getattr(m, 's5_region_a_replied_english', 0) or 0)
        out['s5_noreply'] = int(getattr(m, 's5_region_a_no_reply', 0) or 0)

        out['s6_a_hindi'] = int(getattr(m, 's6_region_a_hindi_bilingual', 0) or 0)
        out['s6_a_eng'] = int(getattr(m, 's6_region_a_english_only', 0) or 0)
        out['s6_a_total'] = int(getattr(m, 's6_region_a_total', 0) or 0)
        out['s6_b_hindi'] = int(getattr(m, 's6_region_b_hindi_bilingual', 0) or 0)
        out['s6_b_eng'] = int(getattr(m, 's6_region_b_english_only', 0) or 0)
        out['s6_b_total'] = int(getattr(m, 's6_region_b_total', 0) or 0)
        out['s6_c_hindi'] = int(getattr(m, 's6_region_c_hindi_bilingual', 0) or 0)
        out['s6_c_eng'] = int(getattr(m, 's6_region_c_english_only', 0) or 0)
        out['s6_c_total'] = int(getattr(m, 's6_region_c_total', 0) or 0)

        out['s7_hindi'] = int(getattr(m, 's7_hindi_pages', 0) or 0)
        out['s7_eng'] = int(getattr(m, 's7_english_pages', 0) or 0)
        out['s7_total'] = int(getattr(m, 's7_total_pages', 0) or 0)
        out['s7_eoffice'] = int(getattr(m, 's7_eoffice_notings', 0) or 0)

        out['s8_workshops'] = int(getattr(m, 's8_full_day_workshops', 0) or 0)
        out['s8_officers'] = int(getattr(m, 's8_officers_trained', 0) or 0)
        out['s8_employees'] = int(getattr(m, 's8_employees_trained', 0) or 0)
    except Exception:
        pass
    return out


def _serialize_adminqpr(a):
    """Map an AdminQPR instance to NUMERIC_KEYS-shaped dict."""
    out = {k: 0 for k in NUMERIC_KEYS}
    if not a:
        return out
    try:
        out['s2_meetings'] = int(getattr(a, 'a_s2_meetings_count', 0) or 0)
        out['s2_minutes'] = int(getattr(a, 'a_s2_hindi_minutes', 0) or 0)
        out['s2_papers_total'] = int(getattr(a, 'a_s2_total_papers', 0) or 0)
        out['s2_papers_hindi'] = int(getattr(a, 'a_s2_hindi_papers', 0) or 0)

        out['s3_total'] = int(getattr(a, 'a_s3_total_documents', 0) or 0)
        out['s3_bilingual'] = int(getattr(a, 'a_s3_bilingual_documents', 0) or 0)
        out['s3_english'] = int(getattr(a, 'a_s3_english_only_documents', 0) or 0)
        out['s3_hindi_only'] = int(getattr(a, 'a_s3_hindi_only_documents', 0) or 0)

        out['s4_total'] = int(getattr(a, 'a_s4_total_letters', 0) or 0)
        out['s4_no_reply'] = int(getattr(a, 'a_s4_no_reply_letters', 0) or 0)
        out['s4_replied_hindi'] = int(getattr(a, 'a_s4_replied_hindi_letters', 0) or 0)
        out['s4_replied_eng'] = int(getattr(a, 'a_s4_replied_english_letters', 0) or 0)

        # map other admin fields if present, best-effort
        out['s7_hindi'] = int(getattr(a, 'a_s7_hindi_pages', 0) or 0)
        out['s7_eng'] = int(getattr(a, 'a_s7_english_pages', 0) or 0)
        out['s7_total'] = int(getattr(a, 'a_s7_total_pages', 0) or 0)
        out['s7_eoffice'] = int(getattr(a, 'a_s7_eoffice_notings', 0) or 0)
    except Exception:
        pass
    return out

def _aggregate_records_for_range(user, start_dt, end_dt, source_frequency='daily'):
    """Sum numeric fields of submitted records for a user whose period overlaps [start_dt,end_dt].

    Only records matching `source_frequency` are considered. Records without explicit
    period_start/period_end are ignored (no fallback to quarter) to avoid accidental
    full-quarter overlaps.
    """
    total = {k: 0 for k in NUMERIC_KEYS}
    if not start_dt or not end_dt:
        return total

    # Base queryset: filter by user, submission state and frequency only.
    # We intentionally avoid requiring explicit period_start/period_end here so
    # older records that may miss one of those fields are still considered.
    qs = QPRRecord.objects.filter(
        user=user,
        is_submitted=True,
        frequency__iexact=(source_frequency or '')
    )

    for r in qs:
        # Determine effective start/end for the record with safe fallbacks.
        try:
            r_start = getattr(r, 'period_start', None)
            r_end = getattr(r, 'period_end', None)

            # If one of the explicit bounds is missing, try to infer sensibly
            # without modifying the DB. These heuristics keep behaviour non-destructive.
            if r_start and not r_end:
                freq = (getattr(r, 'frequency', '') or '').lower()
                if freq == 'daily':
                    r_end = r_start
                elif freq == 'weekly':
                    r_end = r_start + timedelta(days=5)
                elif freq == 'monthly':
                    # last day of month for r_start
                    y, m = r_start.year, r_start.month
                    if m == 12:
                        r_end = date(y, 12, 31)
                    else:
                        r_end = date(y, m + 1, 1) - timedelta(days=1)
                elif freq == 'quarterly':
                    try:
                        q_s, q_e = _quarter_label_to_daterange(getattr(r, 'quarter', None), getattr(r, 'year', None))
                        r_start = r_start or q_s
                        r_end = r_end or q_e
                    except Exception:
                        r_end = r_start
                else:
                    r_end = r_start

            if not r_start and r_end:
                r_start = r_end

            # As a last resort use created_at date if neither bound exists
            if not r_start and not r_end:
                created = getattr(r, 'created_at', None)
                if created:
                    r_start = created.date()
                    r_end = r_start
                else:
                    # Skip records with no usable date info
                    continue

            # Now check overlap with requested range
            if r_start <= end_dt and r_end >= start_dt:
                try:
                    data = serialize_qpr_record(r)
                except Exception:
                    continue

                for k in NUMERIC_KEYS:
                    v = data.get(k)
                    if v is None or v == '':
                        continue
                    try:
                        total[k] += int(v)
                    except Exception:
                        continue
        except Exception:
            continue

    return total


def _aggregate_records_with_fallback(user, start_dt, end_dt, preferred='daily'):
    """Try preferred frequency then fall back to more granular sources.

    Returns a totals dict keyed by NUMERIC_KEYS. This mirrors the previous
    inner helper used elsewhere but exposes it at module level for reuse.
    """
    pref = (preferred or '').lower()

    # Helper: check if totals dict has any non-zero numeric keys
    def _has_nonzero(tot):
        return any((tot.get(k, 0) or 0) != 0 for k in NUMERIC_KEYS)

    # 1) Daily: trivial
    if pref == 'daily':
        try:
            return _aggregate_records_for_range(user, start_dt, end_dt, source_frequency='daily')
        except Exception:
            return {k: 0 for k in NUMERIC_KEYS}

    # 2) Weekly: prefer weekly for the entire range, else fall back to daily
    if pref == 'weekly':
        try:
            totals = _aggregate_records_for_range(user, start_dt, end_dt, source_frequency='weekly')
        except Exception:
            totals = {k: 0 for k in NUMERIC_KEYS}
        if _has_nonzero(totals):
            return totals
        try:
            return _aggregate_records_for_range(user, start_dt, end_dt, source_frequency='daily')
        except Exception:
            return {k: 0 for k in NUMERIC_KEYS}

    # 3) Monthly: Prefer a monthly record for the whole month; otherwise iterate weeks
    if pref == 'monthly':
        try:
            monthly_tot = _aggregate_records_for_range(user, start_dt, end_dt, source_frequency='monthly')
        except Exception:
            monthly_tot = {k: 0 for k in NUMERIC_KEYS}
        if _has_nonzero(monthly_tot):
            return monthly_tot

        # iterate Mon-Sat weeks within [start_dt,end_dt]
        acc = {k: 0 for k in NUMERIC_KEYS}
        w_start = start_dt - timedelta(days=start_dt.weekday())
        # ensure we start from the Monday on/ before start_dt and iterate weeks
        cur = w_start
        while cur <= end_dt:
            week_start = cur
            week_end = cur + timedelta(days=5)
            # clip to requested month range
            actual_start = max(week_start, start_dt)
            actual_end = min(week_end, end_dt)
            if actual_start <= actual_end:
                # try weekly for that week, else daily
                try:
                    wtot = _aggregate_records_for_range(user, actual_start, actual_end, source_frequency='weekly')
                except Exception:
                    wtot = {k: 0 for k in NUMERIC_KEYS}
                if not _has_nonzero(wtot):
                    try:
                        wtot = _aggregate_records_for_range(user, actual_start, actual_end, source_frequency='daily')
                    except Exception:
                        wtot = {k: 0 for k in NUMERIC_KEYS}
                for k in NUMERIC_KEYS:
                    acc[k] += int(wtot.get(k, 0) or 0)
            cur = cur + timedelta(days=7)
        return acc

    # 4) Quarterly or other: prefer quarterly then per-month breakdown
    # Attempt quarterly first
    try:
        qtot = _aggregate_records_for_range(user, start_dt, end_dt, source_frequency='quarterly')
    except Exception:
        qtot = {k: 0 for k in NUMERIC_KEYS}
    if _has_nonzero(qtot):
        return qtot

    # otherwise iterate months inside [start_dt,end_dt]
    acc = {k: 0 for k in NUMERIC_KEYS}
    m = start_dt
    while m <= end_dt:
        month_start = date(m.year, m.month, 1)
        if m.month == 12:
            month_end = date(m.year, 12, 31)
        else:
            month_end = date(m.year, m.month + 1, 1) - timedelta(days=1)
        if month_end > end_dt:
            month_end = end_dt
        if month_start < start_dt:
            month_start = start_dt

        # prefer monthly record for this month
        try:
            mtot = _aggregate_records_for_range(user, month_start, month_end, source_frequency='monthly')
        except Exception:
            mtot = {k: 0 for k in NUMERIC_KEYS}
        if _has_nonzero(mtot):
            for k in NUMERIC_KEYS:
                acc[k] += int(mtot.get(k, 0) or 0)
        else:
            # iterate weeks in this month
            w_start = month_start - timedelta(days=month_start.weekday())
            cur = w_start
            while cur <= month_end:
                week_start = cur
                week_end = cur + timedelta(days=5)
                actual_start = max(week_start, month_start)
                actual_end = min(week_end, month_end)
                if actual_start <= actual_end:
                    try:
                        wtot = _aggregate_records_for_range(user, actual_start, actual_end, source_frequency='weekly')
                    except Exception:
                        wtot = {k: 0 for k in NUMERIC_KEYS}
                    if not _has_nonzero(wtot):
                        try:
                            wtot = _aggregate_records_for_range(user, actual_start, actual_end, source_frequency='daily')
                        except Exception:
                            wtot = {k: 0 for k in NUMERIC_KEYS}
                    for k in NUMERIC_KEYS:
                        acc[k] += int(wtot.get(k, 0) or 0)
                cur = cur + timedelta(days=7)

        # move to next month
        if m.month == 12:
            m = date(m.year + 1, 1, 1)
        else:
            m = date(m.year, m.month + 1, 1)

    return acc


def _aggregate_section11_text_for_range(user, start_dt, end_dt, text_field_name, source_frequency='daily'):
    """Concatenate text from Section 11 fields of submitted records whose period overlaps [start_dt, end_dt]."""
    text_parts = []
    if not start_dt or not end_dt:
        return ''

    qs = QPRRecord.objects.filter(
        user=user,
        is_submitted=True,
        frequency__iexact=(source_frequency or '')
    )

    for r in qs:
        try:
            r_start = getattr(r, 'period_start', None)
            r_end = getattr(r, 'period_end', None)

            if r_start and not r_end:
                freq = (getattr(r, 'frequency', '') or '').lower()
                if freq == 'daily':
                    r_end = r_start
                elif freq == 'weekly':
                    r_end = r_start + timedelta(days=5)
                elif freq == 'monthly':
                    y, m = r_start.year, r_start.month
                    if m == 12:
                        r_end = date(y, 12, 31)
                    else:
                        r_end = date(y, m + 1, 1) - timedelta(days=1)
                elif freq == 'quarterly':
                    try:
                        q_s, q_e = _quarter_label_to_daterange(getattr(r, 'quarter', None), getattr(r, 'year', None))
                        r_start = r_start or q_s
                        r_end = r_end or q_e
                    except Exception:
                        r_end = r_start
                else:
                    r_end = r_start

            if not r_start and r_end:
                r_start = r_end

            if not r_start and not r_end:
                created = getattr(r, 'created_at', None)
                if created:
                    r_start = created.date()
                    r_end = r_start
                else:
                    continue

            if r_start <= end_dt and r_end >= start_dt:
                s11 = getattr(r, 'section11', None)
                if s11:
                    text_value = getattr(s11, text_field_name, '')
                    if text_value and text_value.strip():
                        text_parts.append(text_value.strip())
        except Exception:
            continue

    return '\n---\n'.join(text_parts)


def _get_quarter_range_for_date(dt):
    m = dt.month
    y = dt.year
    if m in (4,5,6):
        return (date(y,4,1), date(y,6,30))
    if m in (7,8,9):
        return (date(y,7,1), date(y,9,30))
    if m in (10,11,12):
        return (date(y,10,1), date(y,12,31))
    # Jan-Mar
    return (date(y,1,1), date(y,3,31))


def determine_submission_frequency(user, submission_date=None, is_submitted=True):
    """Enforce submission rules and return (frequency, period_start, period_end).

    Rules implemented:
    - Normal operation: `daily` submissions allowed on each working day (Mon-Sat).
    - If a user has at least one `daily` in the current week but has missed earlier working day(s) up to today,
      the server will require a single `weekly` submission for that week (Mon-Sat). After a weekly is created,
      further `daily` submissions for that week are blocked.
    - If a user has zero `daily` submissions for the entire current week, they are blocked from daily/weekly
      submissions until month end; on month end they may submit a single `monthly` for that month.
    - If a user has zero `daily` submissions for the entire month, they are blocked from daily/weekly/monthly
      until quarter end; on quarter end they may submit a single `quarterly` for that quarter.

    If `is_submitted` is False (saving as Draft), this function will return `daily` for drafts and not enforce blocks.
    Raises ValueError with a descriptive message when submission is not allowed at this time.
    """
    if submission_date is None:
        submission_date = date.today()

    # Week (Mon-Sat) starting Monday
    week_start = submission_date - timedelta(days=submission_date.weekday())
    week_end = week_start + timedelta(days=5)  # Mon-Sat (Saturday is week end)

    # Month range
    month_start = date(submission_date.year, submission_date.month, 1)
    if submission_date.month == 12:
        month_end = date(submission_date.year, 12, 31)
    else:
        month_end = date(submission_date.year, submission_date.month + 1, 1) - timedelta(days=1)

    # Quarter range
    q_start, q_end = _get_quarter_range_for_date(submission_date)

    def _last_working_before(d):
        while d.weekday() > 5:  # Sunday (6)
            d = d - timedelta(days=1)
        return d

    month_last_working = _last_working_before(month_end)
    quarter_last_working = _last_working_before(q_end)

    # Drafts are given a neutral daily default without enforcement
    if not is_submitted:
        return ('daily', submission_date, submission_date)


def compute_period(frequency, selected_date=None, quarter=None, year=None):
    """Compute (period_start, period_end) for given frequency.

    - frequency: 'daily'|'weekly'|'monthly'|'quarterly'
    - selected_date: datetime.date used for daily/weekly/monthly
    - quarter, year: used for quarterly
    """
    if selected_date is None:
        selected_date = timezone.localdate()

    if frequency == 'daily':
        return (selected_date, selected_date)

    if frequency == 'weekly':
        # Week is Mon-Sat (server convention)
        start = selected_date - timedelta(days=selected_date.weekday())
        end = start + timedelta(days=5)
        # Clamp within quarter boundaries to avoid crossing into adjacent quarters
        try:
            q_start, q_end = _get_quarter_range_for_date(selected_date)
            if start < q_start: start = q_start
            if end > q_end: end = q_end
        except Exception:
            pass
        return (start, end)

    if frequency == 'monthly':
        start = date(selected_date.year, selected_date.month, 1)
        if selected_date.month == 12:
            end = date(selected_date.year, 12, 31)
        else:
            end = date(selected_date.year, selected_date.month + 1, 1) - timedelta(days=1)
        # Clamp within quarter boundaries to avoid spanning adjacent quarters
        try:
            q_start, q_end = _get_quarter_range_for_date(selected_date)
            if start < q_start: start = q_start
            if end > q_end: end = q_end
        except Exception:
            pass
        return (start, end)

    if frequency == 'quarterly':
        # Use existing helper to map quarter label + fiscal year to range
        if quarter and year:
            try:
                return _quarter_label_to_daterange(quarter, year)
            except Exception:
                pass
        # fallback: compute quarter containing selected_date
        return _get_quarter_range_for_date(selected_date)

    # default fallback
    return (selected_date, selected_date)


def is_period_overlapping(user, start, end, exclude_id=None, new_frequency=None):
    """Return True if a submitted QPRRecord for user conflicts with [start,end].

    Behaviour:
    - By default (new_frequency is None) returns True if any submitted record
      overlaps the range (preserves original strict behaviour).
    - If `new_frequency=='weekly'` then:
        * overlapping records with frequency != 'daily' (weekly/monthly/quarterly)
          are considered conflicts (e.g. another weekly already exists).
        * daily records are allowed to partially overlap unless they fully
          cover every working day in [start,end], in which case the week is
          considered already covered and this is a conflict.
    - If `new_frequency` in ['monthly', 'quarterly'] then:
        * overlapping records with same frequency and same period are conflicts.
        * daily/weekly records are allowed (monthly/quarterly aggregate).
    """
    if not start or not end:
        return False

    base_qs = QPRRecord.objects.filter(user=user, is_submitted=True)
    if exclude_id:
        base_qs = base_qs.exclude(pk=exclude_id)

    # Default strict behaviour: any overlap is a conflict
    if not new_frequency:
        return base_qs.filter(period_start__lte=end, period_end__gte=start).exists()

    # Special-case: creating a weekly record — allow partial overlap with daily
    if str(new_frequency).lower() == 'weekly':
        # 1) If any non-daily submitted record overlaps, treat as conflict
        non_daily_conflict = base_qs.exclude(frequency__iexact='daily').filter(period_start__lte=end, period_end__gte=start).exists()
        if non_daily_conflict:
            return True

        # 2) Count submitted daily records within [start,end] and compare to expected working days
        daily_count = base_qs.filter(frequency__iexact='daily', period_start__range=(start, end)).count()
        expected_days = 0
        d = start
        while d <= end:
            if d.weekday() <= 5:  # Mon-Sat are working days
                expected_days += 1
            d = d + timedelta(days=1)

        # If all working days already have daily submissions, it's a conflict
        if expected_days > 0 and daily_count >= expected_days:
            return True

        # Otherwise allow weekly creation (no conflict)
        return False

    # Special-case: creating a daily record — only conflict if same day already exists
    if str(new_frequency).lower() == 'daily':
        # Daily records should only conflict if another daily record for the same day exists
        # (they represent the same day's report)
        same_day_overlap = base_qs.filter(
            frequency__iexact='daily',
            period_start=start,
            period_end=end
        ).exists()
        return same_day_overlap

    # Special-case: creating a monthly/quarterly record — only conflict with same frequency/period
    if str(new_frequency).lower() in ['monthly', 'quarterly']:
        # Monthly/quarterly are cumulative aggregations; only conflict if exact same period exists
        # Don't conflict with daily/weekly records (those are incorporated into the aggregate)
        same_freq_overlap = base_qs.filter(
            frequency__iexact=new_frequency,
            period_start=start,
            period_end=end
        ).exists()
        return same_freq_overlap

    # Fallback to strict behaviour for other frequencies
    return base_qs.filter(period_start__lte=end, period_end__gte=start).exists()


def _allowed_frequencies_for_date(user, selected_date, allow_future_days=True):
    """Return a dict with allowed frequencies and missing days for the selected_date.

    Result example:
    {
      'allowed': ['daily','weekly'],
      'missing_days_week': ['2026-03-23','2026-03-24'],
      'missing_days_month': [...],
      'min_date': '2025-04-01', 'max_date': '2026-04-25'
    }
    """
    today = timezone.localdate()
    # min_date: earliest submitted period_start for user or start of current financial year
    earliest = QPRRecord.objects.filter(user=user).order_by('period_start').first()
    if earliest and earliest.period_start:
        min_date = earliest.period_start
    else:
        # fiscal year start: Apr 1 of current fiscal year
        fy_start = today.year if today.month >= 4 else today.year - 1
        min_date = date(fy_start, 4, 1)

    # Max date: by default allow one month ahead from today (user can plan one month in advance)
    # If allow_future_days is False (used by the user/HOD QPR form), restrict max_date to today.
    if not allow_future_days:
        max_date = today
    else:
        try:
            # Handle month overflow (e.g., Jan 31 -> Feb doesn't have 31 days)
            if today.month == 12:
                next_month_year, next_month_month = today.year + 1, 1
            else:
                next_month_year, next_month_month = today.year, today.month + 1
            
            # Try to create the same day in next month; fallback to last day of month if it doesn't exist
            try:
                max_date = date(next_month_year, next_month_month, today.day)
            except ValueError:
                # Day doesn't exist in target month (e.g., Jan 31 -> Feb 31 doesn't exist)
                # Use last day of the month
                if next_month_month == 2:
                    max_date = date(next_month_year, 2, 29 if next_month_year % 4 == 0 else 28)
                elif next_month_month in [4, 6, 9, 11]:
                    max_date = date(next_month_year, next_month_month, 30)
                else:
                    max_date = date(next_month_year, next_month_month, 31)
        except Exception:
            # Fallback to today + 30 days if anything fails
            max_date = today + timedelta(days=30)

    # Normalize selected_date within bounds
    if selected_date < min_date:
        selected_date = min_date
    if selected_date > max_date:
        selected_date = max_date

    # Week (Mon-Sat)
    q_start, q_end = _get_quarter_range_for_date(selected_date)

    week_start = selected_date - timedelta(days=selected_date.weekday())
    week_end = week_start + timedelta(days=5)

    # CLIP TO QUARTER
    week_start = max(week_start, q_start)
    week_end = min(week_end, q_end)
    week_days = [
    d for d in (week_start + timedelta(days=i) for i in range((week_end - week_start).days + 1))
    if d.weekday() <= 5 and q_start <= d <= q_end ]
    # Submitted weekly dates in week
    submitted_week = set(QPRRecord.objects.filter(user=user, is_submitted=True, frequency__iexact='weekly', period_start__range=(week_start, week_end)).values_list('period_start', flat=True))
    missing_week = [d for d in week_days if d not in submitted_week and d >= min_date and d <= max_date]

    # Month
    month_start = date(selected_date.year, selected_date.month, 1)
    if selected_date.month == 12:
        month_end = date(selected_date.year, 12, 31)
    else:
        month_end = date(selected_date.year, selected_date.month + 1, 1) - timedelta(days=1)
    month_days = [month_start + timedelta(days=i) for i in range((month_end - month_start).days + 1) if (month_start + timedelta(days=i)).weekday() <= 5]
    submitted_month = set(QPRRecord.objects.filter(user=user, is_submitted=True, frequency__iexact='monthly', period_start__range=(month_start, month_end)).values_list('period_start', flat=True))
    missing_month = [d for d in month_days if d not in submitted_month and d >= min_date and d <= max_date]

    # Quarter
    q_start, q_end = _get_quarter_range_for_date(selected_date)
    quarter_days = [q_start + timedelta(days=i) for i in range((q_end - q_start).days + 1) if (q_start + timedelta(days=i)).weekday() <= 5]
    submitted_quarter = set(QPRRecord.objects.filter(user=user, is_submitted=True, frequency__iexact='quarterly', period_start__range=(q_start, q_end)).values_list('period_start', flat=True))
    missing_quarter = [d for d in quarter_days if d not in submitted_quarter and d >= min_date and d <= max_date]

    allowed = ['daily']
    
    # Helper to find last working day before a date
    def _last_working_before(d):
        while d.weekday() > 5:  # Sunday is 6
            d = d - timedelta(days=1)
        return d
    
    month_last = _last_working_before(month_end)
    quarter_last = _last_working_before(q_end)
    
    # weekly allowed only for fully completed past weeks (Mon-Sat) that have missing working days.
    # Do not allow weekly for current ongoing week or any future weeks.
    if len(missing_week) > 0 and week_end <= today:
        allowed.append('weekly')
    # monthly allowed only at month end if there are missing working days in the month
    if len(missing_month) > 0 and selected_date >= month_last:
        allowed.append('monthly')
    # quarterly allowed only at quarter end if there are missing working days in the quarter
    if len(missing_quarter) > 0 and selected_date >= quarter_last:
        allowed.append('quarterly')

    return {
        'allowed': allowed,
        'missing_week': [d.isoformat() for d in missing_week],
        'missing_month': [d.isoformat() for d in missing_month],
        'missing_quarter': [d.isoformat() for d in missing_quarter],
        'min_date': min_date.isoformat(),
        'max_date': max_date.isoformat(),
        'default_date': timezone.localdate().isoformat()
    }


def compute_cumulative_for_record(record):
    """Return dict with daily/weekly/monthly/quarterly aggregates for the given record."""
    user = record.user
    # determine quarter range first (always parse from record's quarter/year)
    q_start, q_end = None, None
    try:
        q_start, q_end = _quarter_label_to_daterange(record.quarter, record.year or '')
    except Exception:
        # fallback to today's month
        today = date.today()
        q_start = date(today.year, today.month, 1)
        if today.month == 12:
            q_end = date(today.year, 12, 31)
        else:
            q_end = date(today.year, today.month + 1, 1) - timedelta(days=1)
    
    # determine base date for daily aggregation
    base = None
    if getattr(record, 'period_start', None):
        base = record.period_start
    elif getattr(record, 'period_end', None):
        base = record.period_end
    else:
        # If period is missing, try to infer from record.frequency and created_at
        freq = (getattr(record, 'frequency', '') or '').lower()
        created = getattr(record, 'created_at', None)
        if freq == 'weekly' and created:
            base = created.date()
        elif freq == 'monthly' and created:
            base = created.date()
        elif freq == 'daily' and created:
            base = created.date()
        else:
            # use quarter end as fallback base for daily/weekly/monthly when no specific period
            base = q_end
    
    # daily
    day_start = base
    day_end = base
    # weekly (Mon-Sat week used elsewhere; keep to Mon-Sat)
    week_start = day_start - timedelta(days=day_start.weekday())
    week_end = week_start + timedelta(days=5)
    # monthly
    month_start = date(day_start.year, day_start.month, 1)
    # compute month end
    if day_start.month == 12:
        month_end = date(day_start.year, 12, 31)
    else:
        month_end = date(day_start.year, day_start.month + 1, 1) - timedelta(days=1)

    # Helper to try preferred source then fall back to lower-frequency sources
    def _aggregate_with_fallback(user, start_dt, end_dt, preferred):
        order = []
        pref = (preferred or '').lower()
        if pref == 'daily':
            order = ['daily']
        elif pref == 'weekly':
            order = ['weekly', 'daily']
        elif pref == 'monthly':
            order = ['monthly', 'weekly', 'daily']
        else:
            # quarterly or unknown: try monthly -> weekly -> daily
            order = ['monthly', 'weekly', 'daily']

        last_totals = None
        for src in order:
            try:
                totals = _aggregate_records_for_range(user, start_dt, end_dt, source_frequency=src)
            except Exception:
                totals = {k: 0 for k in NUMERIC_KEYS}
            last_totals = totals
            # if any numeric key is non-zero, accept these totals
            if any((totals.get(k, 0) or 0) != 0 for k in NUMERIC_KEYS):
                return totals
        # if all zero, return the last computed (likely zeros)
        return last_totals or {k: 0 for k in NUMERIC_KEYS}

    try:
        daily_tot = _aggregate_with_fallback(user, day_start, day_end, 'daily')
        weekly_tot = _aggregate_with_fallback(user, week_start, week_end, 'weekly')
        monthly_tot = _aggregate_with_fallback(user, month_start, month_end, 'monthly')
        quarterly_tot = _aggregate_with_fallback(user, q_start, q_end, 'quarterly')
        return {
            'daily': daily_tot,
            'weekly': weekly_tot,
            'monthly': monthly_tot,
            'quarterly': quarterly_tot,
        }
    except Exception:
        zeros = {k: 0 for k in NUMERIC_KEYS}
        return {'daily': zeros.copy(), 'weekly': zeros.copy(), 'monthly': zeros.copy(), 'quarterly': zeros.copy()}

def _aggregate_text_section_11(user, start_dt, end_dt):
    """Gathers Section 11 text from submitted Daily records within a range."""
    from .models import Section11SpecificAchievementsData
    
    daily_s11 = Section11SpecificAchievementsData.objects.filter(
        qpr_record__user=user,
        qpr_record__frequency__iexact='daily',
        qpr_record__is_submitted=True,
        qpr_record__period_start__range=[start_dt, end_dt]
    ).select_related('qpr_record').order_by('qpr_record__period_start')

    innovative, events, medium = [], [], []

    for item in daily_s11:
        date_label = item.qpr_record.period_start.strftime('%d-%m-%Y')
        if item.innovative_work and item.innovative_work.strip():
            innovative.append(f"[{date_label}]: {item.innovative_work.strip()}")
        if item.special_events and item.special_events.strip():
            events.append(f"[{date_label}]: {item.special_events.strip()}")
        if item.hindi_medium_works and item.hindi_medium_works.strip():
            medium.append(f"[{date_label}]: {item.hindi_medium_works.strip()}")

    return {
        's12_1': "\n\n".join(innovative),
        's12_2': "\n\n".join(events),
        's12_3': "\n\n".join(medium),
    }

def serialize_qpr_record(record):
    """Serialize a QPRRecord with all related sections."""
    data = {
        'id': record.id,
        'officeName': record.officeName,
        'officeCode': record.officeCode,
        'region': record.region,
        'quarter': record.quarter,
        'year': record.year or '2025-2026',
        'status': record.status,
        'is_submitted': record.is_submitted,
        'phone': record.phone or '',
        'email': record.email or '',
        # Section 1
        's1_total': getattr(record.section1, 'total_files', '') if hasattr(record, 'section1') else '',
        's1_hindi': getattr(record.section1, 'hindi_files', '') if hasattr(record, 'section1') else '',
        # Section 2
        's2_meetings': getattr(record.section2, 'meetings_count', '') if hasattr(record, 'section2') else '',
        's2_minutes': getattr(record.section2, 'hindi_minutes', '') if hasattr(record, 'section2') else '',
        's2_papers_total': getattr(record.section2, 'total_papers', '') if hasattr(record, 'section2') else '',
        's2_papers_hindi': getattr(record.section2, 'hindi_papers', '') if hasattr(record, 'section2') else '',
        # Section 3
        's3_total': getattr(record.section3, 'total_documents', '') if hasattr(record, 'section3') else '',
        's3_bilingual': getattr(record.section3, 'bilingual_documents', '') if hasattr(record, 'section3') else '',
        's3_english': getattr(record.section3, 'english_only_documents', '') if hasattr(record, 'section3') else '',
        's3_hindi_only': getattr(record.section3, 'hindi_only_documents', '') if hasattr(record, 'section3') else '',
        # Section 4
        's4_total': getattr(record.section4, 'total_letters', '') if hasattr(record, 'section4') else '',
        's4_no_reply': getattr(record.section4, 'no_reply_letters', '') if hasattr(record, 'section4') else '',
        's4_replied_hindi': getattr(record.section4, 'replied_hindi_letters', '') if hasattr(record, 'section4') else '',
        's4_replied_eng': getattr(record.section4, 'replied_english_letters', '') if hasattr(record, 'section4') else '',
        # Section 5
        's5_total': getattr(record.section5, 'region_a_english_letters', '') if hasattr(record, 'section5') else '',
        's5_hindi': getattr(record.section5, 'region_a_replied_hindi', '') if hasattr(record, 'section5') else '',
        's5_english': getattr(record.section5, 'region_a_replied_english', '') if hasattr(record, 'section5') else '',
        's5_noreply': getattr(record.section5, 'region_a_no_reply', '') if hasattr(record, 'section5') else '',
        # Section 6
        's6_a_hindi': getattr(record.section6, 'region_a_hindi_bilingual', '') if hasattr(record, 'section6') else '',
        's6_a_eng': getattr(record.section6, 'region_a_english_only', '') if hasattr(record, 'section6') else '',
        's6_a_total': getattr(record.section6, 'region_a_total', '') if hasattr(record, 'section6') else '',
        's6_b_hindi': getattr(record.section6, 'region_b_hindi_bilingual', '') if hasattr(record, 'section6') else '',
        's6_b_eng': getattr(record.section6, 'region_b_english_only', '') if hasattr(record, 'section6') else '',
        's6_b_total': getattr(record.section6, 'region_b_total', '') if hasattr(record, 'section6') else '',
        's6_c_hindi': getattr(record.section6, 'region_c_hindi_bilingual', '') if hasattr(record, 'section6') else '',
        's6_c_eng': getattr(record.section6, 'region_c_english_only', '') if hasattr(record, 'section6') else '',
        's6_c_total': getattr(record.section6, 'region_c_total', '') if hasattr(record, 'section6') else '',
        # Section 7
        's7_hindi': getattr(record.section7, 'hindi_pages', '') if hasattr(record, 'section7') else '',
        's7_eng': getattr(record.section7, 'english_pages', '') if hasattr(record, 'section7') else '',
        's7_total': getattr(record.section7, 'total_pages', '') if hasattr(record, 'section7') else '',
        's7_eoffice': getattr(record.section7, 'eoffice_notings', '') if hasattr(record, 'section7') else '',
        # Section 8
        's8_workshops': getattr(record.section8, 'full_day_workshops', '') if hasattr(record, 'section8') else '',
        's8_officers': getattr(record.section8, 'officers_trained', '') if hasattr(record, 'section8') else '',
        's8_employees': getattr(record.section8, 'employees_trained', '') if hasattr(record, 'section8') else '',
        # Section 9
        's9_date': getattr(record.section9, 'meeting_date', '') if hasattr(record, 'section9') else '',
        's9_sub_committees': getattr(record.section9, 'sub_committees_count', '') if hasattr(record, 'section9') else '',
        's9_meetings_count': getattr(record.section9, 'meetings_organized', '') if hasattr(record, 'section9') else '',
        's9_agenda_hindi': getattr(record.section9, 'agenda_hindi', '') if hasattr(record, 'section9') else '',
        # Section 10
        's10_date': getattr(record.section10, 'meeting_date', '') if hasattr(record, 'section10') else '',
        # Section 11
        's12_1': getattr(record.section11, 'innovative_work', '') if hasattr(record, 'section11') else '',
        's12_2': getattr(record.section11, 'special_events', '') if hasattr(record, 'section11') else '',
        's12_3': getattr(record.section11, 'hindi_medium_works', '') if hasattr(record, 'section11') else '',
        'details': {}
    }
    # Include submission frequency and explicit period when available
    data['frequency'] = getattr(record, 'frequency', 'quarterly') if record else 'quarterly'
    data['period_start'] = getattr(record, 'period_start', None)
    data['period_end'] = getattr(record, 'period_end', None)
    data['is_quarterly_frozen'] = getattr(record, 'is_quarterly_frozen', False)
    # Normalize numeric keys: convert None/empty to 0 so cumulative sums include them
    try:
        for k in NUMERIC_KEYS:
            if data.get(k) is None or data.get(k) == '':
                data[k] = 0
    except Exception:
        pass
    # Also provide a `details` dictionary that mirrors form input ids so client
    # `editRecord()` can populate fields from `record.details` when editing.
    try:
        details = {}
        # numeric and simple keys used by the form
        for k in NUMERIC_KEYS:
            details[k] = data.get(k, 0)
        # date / text fields
        details['s9_date'] = data.get('s9_date', '')
        details['s10_date'] = data.get('s10_date', '')
        details['s12_1'] = data.get('s12_1', '')
        details['s12_2'] = data.get('s12_2', '')
        details['s12_3'] = data.get('s12_3', '')
        data['details'] = details
    except Exception:
        data['details'] = {}
    return data

def send_otp_email(user, lang, target_email=None, email_type='otp'):
    user.otp = str(random.randint(100000, 999999))
    user.otp_created_at = timezone.now()
    user.save(update_fields=['otp', 'otp_created_at'])
    send_system_email(user, None, email_type, extra_context={'otp': user.otp, 'lang': lang}, target_email=target_email)
    return user.otp


def custom_logout(request):
    logout(request)
    messages.success(request, "You have been logged out successfully.")
    return redirect('home')

def home(request):
    events = get_all_events()
    return render(request, "home.html", {"events": events})

def event_detail(request, folder):
    events = get_all_events()
    selected_event = next((e for e in events if e["folder"] == folder), None)

    if not selected_event:
        return redirect("home")

    return render(request, "event_detail.html", {"event": selected_event})

def universal_error_view(request, exception=None, status_code=500):
    lang = request.session.get('lang', 'en')
    error_map = {
        400: {'title': "Bad Request",'msg': "The server could not understand the request due to invalid syntax."},
        403: {'title': "Security Verification Failed",'msg': "You do not have permission to access this resource or your session has expired."},
        404: {'title': "Page Not Found",'msg': "The page you are looking for might have been removed or does not exist."},
        500: {'title': "Internal Server Error",'msg': "Something went wrong on our end. We're working on fixing it."}
    }
    config = error_map.get(status_code, error_map[500])
    context = {'current_lang': lang, 'status_code': status_code, 'error_title': config['title'], 'error_message': config['msg']}
    return render(request, 'error.html', context, status=status_code)

def error_400(request, exception=None): return universal_error_view(request, exception, 400)
def error_403(request, exception=None): return universal_error_view(request, exception, 403)
def csrf_failure(request, reason=""): return universal_error_view(request, None, 403)
def error_404(request, exception=None): return universal_error_view(request, exception, 404)
def error_500(request): return universal_error_view(request, None, 500)

@login_required
def dashboard(request):
    """Central Dashboard Router - Routes each role to their dedicated dashboard"""
    user = request.user
    role = request.session.get('active_role', user_role(user))
    profile = getattr(user, 'profile', None)
    
    # Dashboard routing uses session active_role (set at login) or falls back to user's primary role
    
    context = {
        'current_lang': request.session.get('lang', 'en'),
        'role': role
    }
    if role == 'user' and profile:
        if profile.approval_status == 'pending':
            if profile.hod_name =="ADMIN":
                messages.warning(request, "Your registration is pending Admin approval.")
            else:    
                messages.warning(request, "Your registration is pending HOD approval. You may edit your details while you wait.")
            return redirect('qpr_user_profile') # Locks them into the profile edit page
        elif profile.approval_status == 'rejected':
            messages.error(request, "Your registration was rejected. Please verify your details and update them, or contact admin.")
            return redirect('qpr_user_profile')
    # 1. ADMIN - System administration, HOD management, archive/unarchive, typing reports
    if role == 'admin':
        return redirect('qpr_admin_dashboard')

    # 2. MANAGER - Edit request approvals, employee records management, designations
    elif role == 'manager':
        return redirect('manager_dashboard')
    
    # 3. HOD - Department oversight, employee statistics, detail list
    elif role == 'hod':
        return redirect('qpr_hod_dashboard')
        
    # 4. BACKUP USER - Database download and backup management
    elif role == 'backup_user':
        return render(request, 'dashboard.html', context)

    # 5. USER (Default) - Profile management, QPR forms, employee forms
    else:
        return redirect('qpr_user_dashboard')

def privacy_policy(request):
    return render(request, 'privacy_policy.html')

def toggle_language(request):
    current = request.session.get('lang', 'en')
    request.session['lang'] = 'hi' if current == 'en' else 'en'
    return redirect(request.META.get('HTTP_REFERER', 'home'))

class CustomLoginView(LoginView):
    authentication_form = CustomLoginForm
    template_name = 'registration/login.html'

    def get_success_url(self):
        return reverse('dashboard')

    def form_valid(self, form):
        user = cast(CustomUser, form.get_user())
        current_lang = self.request.session.get('lang', 'en')
        
        selected_role = form.cleaned_data.get('role')
        if selected_role and user_has_role(user, selected_role):
            active_role = selected_role
        else:
            active_role = user_role(user)

        # --- NEW OTP LOGIC ---
        email_choice = form.cleaned_data.get('email_choice', 'primary')
        target_email = user.get_email()
        profile = getattr(user, 'profile', None)
        alternate_email = getattr(profile, 'alternate_email', None)
        
        if email_choice == 'alternate':
            if alternate_email:
                target_email = alternate_email
            else:
                messages.warning(self.request, translate_text("No alternate email found in your profile. Sending to official email.", current_lang))

        # Send OTP
        send_otp_email(user, current_lang, target_email=target_email, email_type='login_otp')
        
        # Save pre-login state
        self.request.session['pre_login_user_id'] = user.id
        self.request.session['login_target_email'] = target_email
        self.request.session['is_login_otp'] = True
        self.request.session['lang'] = current_lang
        self.request.session['active_role'] = active_role
        self.request.session.modified = True
        
        messages.success(self.request, translate_text("OTP sent successfully.", current_lang))
        return redirect('verify_otp')

    def form_invalid(self, form):
        username = form.data.get('username')
        user = CustomUser.objects.filter(username=username).first()
        raw_password = form.data.get('password')
        if not isinstance(raw_password, str):
            return super().form_invalid(form)

        if user and not user.is_active and user.check_password(raw_password):
            lang = self.request.session.get('lang', 'en')
            messages.error(self.request, translate_text("Your account has been archived. Please contact the admin.", lang))
            return self.render_to_response(self.get_context_data(form=form))
        return super().form_invalid(form)

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs.update({'request': self.request})
        return kwargs

def signup(request):
    if request.user.is_authenticated:
        return redirect('dashboard')

    lang = request.session.get('lang', 'en')
    form = CustomUserCreationForm(request.POST or None, request=request)

    if request.method == "POST":
        if form.is_valid():
            user = form.save(commit=False)

            # hod_name removed from user-editable form (admin-managed)
            employee_code = request.POST.get('employee_code', '').strip()
            phone = request.POST.get('phone', '').strip()

            # Generate OTP and keep signup data in session until verification
            otp = str(random.randint(100000, 999999))
            signup_data = {
                'username': user.username,
                'email': form.cleaned_data['email'],
                'password': user.password,
                'first_name': user.first_name,
                'otp': otp,
                'otp_time': timezone.now().timestamp()
            }
            request.session['signup_data'] = signup_data
            request.session['is_signup'] = True

            send_system_email(user, request, 'otp', extra_context={'otp': otp, 'lang': lang})

            messages.success(request, "Account verification initiated! Please verify your email with the OTP sent.")
            return redirect('verify_otp')
        else:
            messages.error(request, "Please correct the errors below.")

    return render(request, 'registration/signup.html', {'form': form})

# ==================== PASSWORD & OTP ====================
class LoginOTPView(View):
    def get(self, request):
        user_id = request.session.get('pre_otp_user_id')
        if not user_id:
            return redirect('login')
            
        user = CustomUser.objects.get(id=user_id)
        profile = getattr(user, 'profile', None)
        lang = request.session.get('lang', 'en')
        
        # Helper to mask emails for security (e.g., p***@domain.com)
        def mask_email(email):
            if not email or '@' not in email: return ""
            parts = email.split('@')
            return f"{parts[0][0]}***@{parts[1]}"
            
        context = {
            'primary_email': mask_email(user.get_email()),
            'has_alternate': bool(profile and profile.alternate_email),
            'alternate_email': mask_email(profile.alternate_email) if profile and profile.alternate_email else "",
            'current_lang': lang
        }
        return render(request, 'registration/verify_otp.html', context)
        
    def post(self, request):
        user_id = request.session.get('pre_otp_user_id')
        if not user_id:
            return redirect('login')
            
        user = CustomUser.objects.get(id=user_id)
        profile = getattr(user, 'profile', None)
        action = request.POST.get('action')
        lang = request.session.get('lang', 'en')
        
        if action == 'send_otp':
            email_choice = request.POST.get('email_choice', 'primary')
            target_email = user.get_email()
            if email_choice == 'alternate' and profile and profile.alternate_email:
                target_email = profile.alternate_email
                
            send_otp_email(user, lang, target_email=target_email)
            messages.success(request, translate_text("OTP sent to your selected email.", lang))
            return redirect('login_otp_step')
            
        elif action == 'verify_otp':
            otp_input = request.POST.get('otp', '').strip()
            is_magic_code = settings.DEBUG and otp_input == "123456"
            
            # Check if actual OTP is valid (standard logic)
            is_real_otp_valid = (
                user.otp and
                user.otp == otp_input and 
                user.otp_created_at and 
                (timezone.now() - user.otp_created_at).total_seconds() < 300
            )
            if is_real_otp_valid or is_magic_code:
                # OTP is valid! Log them in properly
                user.otp = None
                user.save(update_fields=['otp'])
                
                auth_login(request, user)
                
                send_system_email(user, request, 'login')
                if user_role(user) == 'user' and profile and not profile.profile_updated:
                    return redirect('qpr_user_profile')
                    
                request.session.pop('pre_otp_user_id', None)
                return redirect('dashboard')
            else:
                messages.error(request, translate_text("Invalid or expired OTP.", lang))
                return redirect('login_otp_step')
        else:
            messages.error(request, translate_text("Invalid request.", lang))
            return redirect('login_otp_step')
class ForgotPasswordView(View):
    def get(self, request):
        return render(request, 'registration/forgot_password.html')
    def post(self, request):
        request.session.pop('is_signup', None)
        request.session.pop('signup_data', None)
        lang = request.session.get('lang', 'en')
        username = request.POST.get('username', '').strip()
        user = CustomUser.objects.filter(username=username).first()
        if user:
            send_otp_email(user, lang, email_type='reset_otp')
            email = user.get_email()
            if email:
                request.session['reset_email_hash'] = hashlib.sha256(email.encode()).hexdigest()
                messages.success(request, translate_text("OTP sent successfully.", lang))
                return redirect('verify_otp')
        messages.error(request, translate_text("User does not exist.", lang))
        return redirect('forgot_password')

class VerifyOTPView(View):
    def get(self, request):
        if not request.session.get('reset_email_hash') and not request.session.get('is_signup') and not request.session.get('is_login_otp'): 
            return redirect('forgot_password')
        lang = request.session.get('lang', 'en')
        context = {'title_text': translate_text("Verify OTP", lang), 'button_text': translate_text("Verify Code", lang), 'current_lang': lang}
        return render(request, 'registration/verify_otp.html', context)

    def post(self, request):
        if request.user.is_authenticated:
            return redirect('dashboard')

        otp_input = request.POST.get('otp', '').strip()
        lang = request.session.get('lang', 'en')
        if request.session.get('is_login_otp'):
            user_id = request.session.get('pre_login_user_id')
            if not user_id: return redirect('login')
            user = CustomUser.objects.get(id=user_id)
            
            att_key, blk_key = f"otp_att_login_{user_id}", f"otp_blk_login_{user_id}"
            if cache.get(blk_key):
                return render(request, 'registration/verify_otp.html', {'is_blocked': True, 'current_lang': lang})
            
            is_magic_code = settings.DEBUG and otp_input == "123456"
            is_real_otp_valid = (
                user.otp and
                user.otp == otp_input and 
                user.otp_created_at and 
                (timezone.now() - user.otp_created_at).total_seconds() < 300
            )
            
            if is_real_otp_valid or is_magic_code:
                user.otp = None
                user.save(update_fields=['otp'])
                auth_login(request, user, backend='django.contrib.auth.backends.ModelBackend')
                send_system_email(user, request, 'login')
                
                request.session.pop('pre_login_user_id', None)
                request.session.pop('is_login_otp', None)
                request.session.pop('login_target_email', None)
                
                profile = getattr(user, 'profile', None)
                if user_role(user) == 'user' and profile and not profile.profile_updated:
                    return redirect('qpr_user_profile')
                return redirect('dashboard')
            else:
                attempts = cache.get(att_key, 0) + 1
                cache.set(att_key, attempts, 600)
                if attempts >= 5: cache.set(blk_key, True, 600)
                messages.error(request, translate_text("Invalid or expired OTP.", lang))
                return render(request, 'registration/verify_otp.html', {'current_lang': lang})
        elif request.session.get('is_signup'):
            signup_data = request.session.get('signup_data')
            if not signup_data:
                messages.error(request, "Session expired. Please sign up again.")
                return redirect('signup')
            email_hash = hashlib.sha256(signup_data['email'].encode()).hexdigest()
            att_key, blk_key = f"otp_att_{email_hash}", f"otp_blk_{email_hash}"
            if cache.get(blk_key):
                return render(request, 'registration/verify_otp.html', {'is_blocked': True, 'current_lang': lang})  
            if otp_input == signup_data['otp']:
                if (timezone.now().timestamp() - signup_data['otp_time']) < 300: 
                    try:
                        with transaction.atomic():
                            user, created = CustomUser.objects.get_or_create(
                                username=signup_data['username'],
                                defaults={
                                    'first_name': signup_data.get('first_name', ''),
                                    'is_active': True,
                                    'consent_given_at': timezone.now()
                                }
                            )
                            user.password = signup_data['password']
                            user.set_email(signup_data['email'])
                            user.save()
                            profile, _ = UserProfile.objects.get_or_create(
                                user=user,
                                defaults={"employee_code": user.username}
                            )
                            profile.approval_status = 'pending'
                            profile.profile_updated = False
                            profile.save()
                        auth_login(request, user, backend='django.contrib.auth.backends.ModelBackend')
                        request.session['lang'] = lang
                        request.session['active_role'] = 'user'
                        send_system_email(user, request, 'welcome')
                        request.session.pop('signup_data', None)
                        request.session.pop('is_signup', None)
                        messages.success(request, "Email verified! Account created successfully.")
                        return redirect('dashboard')
                    except Exception as e:
                        messages.error(request, f"Registration error: {e}")
                        return redirect('signup')
            attempts = cache.get(att_key, 0) + 1
            cache.set(att_key, attempts, 600)
            if attempts >= 5: cache.set(blk_key, True, 600)
            messages.error(request, translate_text("Invalid or expired OTP.", lang))
            return render(request, 'registration/verify_otp.html', {'current_lang': lang})
        elif request.session.get('reset_email_hash'):
            email_hash = request.session.get('reset_email_hash')
            att_key, blk_key = f"otp_att_{email_hash}", f"otp_blk_{email_hash}"
            if cache.get(blk_key):
                return render(request, 'registration/verify_otp.html', {'is_blocked': True, 'current_lang': lang})
            user = CustomUser.objects.filter(email_hash=email_hash).first()
            if user and user.otp == otp_input:
                if user.otp_created_at and (timezone.now() - user.otp_created_at).total_seconds() < 300:
                    request.session['otp_verified'] = True
                    return redirect('reset_password')
            attempts = cache.get(att_key, 0) + 1
            cache.set(att_key, attempts, 600)
            if attempts >= 5: cache.set(blk_key, True, 600)
            messages.error(request, translate_text("Invalid or expired OTP.", lang))
            return render(request, 'registration/verify_otp.html', {'current_lang': lang})
            
        else:
            return redirect('login')

class ResendOTPView(View):
    def get(self, request):
        lang = request.session.get('lang', 'en')
        if request.session.get('is_signup'):
            signup_data = request.session.get('signup_data')
            if not signup_data: return redirect('signup')
            new_otp = str(random.randint(100000, 999999))
            signup_data['otp'] = new_otp
            signup_data['otp_time'] = timezone.now().timestamp()
            request.session['signup_data'] = signup_data
            dummy_user = CustomUser(username=signup_data['username'])
            dummy_user.set_email(signup_data['email'])
            send_system_email(dummy_user, request, 'otp', extra_context={'otp': new_otp, 'lang': lang})
            messages.success(request, translate_text("New OTP sent.", lang))
            return redirect('verify_otp')
        if request.session.get('is_login_otp'):
            user_id = request.session.get('pre_login_user_id')
            target_email = request.session.get('login_target_email')
            if not user_id: return redirect('login')
            user = CustomUser.objects.get(id=user_id)
            send_otp_email(user, lang, target_email=target_email, email_type='login_otp')
            messages.success(request, translate_text("New OTP sent.", lang))
            return redirect('verify_otp')
        email_hash = request.session.get('reset_email_hash')
        if not email_hash: return redirect('forgot_password')
        user = CustomUser.objects.filter(email_hash=email_hash).first()
        if not user: return redirect('forgot_password')
        send_otp_email(user, lang, email_type='reset_otp')
        messages.success(request, translate_text("New OTP sent.", lang))
        return redirect('verify_otp')

class ResetPasswordView(View):
    def get(self, request):
        if not request.session.get('reset_email_hash'):
            return redirect('forgot_password')
        return render(request, 'registration/reset_password.html')

    def post(self, request):
        email_hash = request.session.get('reset_email_hash')
        pwd = request.POST.get('password')
        cfm = request.POST.get('confirm_password')
        if not email_hash:
            return redirect('forgot_password')
        if pwd == cfm:
            user = CustomUser.objects.filter(email_hash=email_hash).first()
            if user:
                user.set_password(pwd)
                user.otp = None
                user.save()
                send_system_email(user, request, 'reset')
                request.session.pop('reset_email_hash', None)
                messages.success(request, "Password reset successfully.")
            return redirect('login')
        messages.error(request, "Passwords do not match.")
        return render(request, 'registration/reset_password.html')

@login_required
def change_password(request):
    if request.method == 'POST':
        old_password = request.POST.get('old_password', '')
        new_password1 = request.POST.get('new_password1', '')
        new_password2 = request.POST.get('new_password2', '')
        if not request.user.check_password(old_password):
            messages.error(request, 'Current password is incorrect')
        elif new_password1 != new_password2:
            messages.error(request, 'New passwords do not match')
        elif len(new_password1) < 6:
            messages.error(request, 'New password must be at least 6 characters')
        else:
            request.user.set_password(new_password1)
            request.user.save()
            messages.success(request, 'Password changed successfully!')
            return redirect('dashboard')
    return render(request, 'qpr/change_password.html')

# ==================== DATA & PRIVACY & ARCHIVING (RESTORED) ====================

@login_required
def user_detail_view(request, user_id):
    """Restored User Detail View with Access Logging"""
    target_user = get_object_or_404(CustomUser, id=user_id)
    lang = request.session.get('lang', 'en')
    active_role = request.session.get('active_role', 'user')
    if request.user != target_user and active_role in ['admin', 'hod']:
        DataAccessLog.objects.create(
            accessed_by=request.user,
            target_user=target_user,
            reason="Manager/Admin Dashboard Review"
        )
    return render(request, 'user_detail.html', {
        'target_user': target_user,
        'current_lang': lang,
        'role': active_role
    })

@login_required
def export_user_data(request):
    user = request.user
    send_system_email(user, request, 'export')
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="{user.username}_data.csv"'
    writer = csv.writer(response)
    writer.writerow(['Category', 'Value'])
    writer.writerow(['Username', user.username])
    writer.writerow(['Email', user.get_email()])
    return response

@login_required
def delete_account(request):
    if request.method == "POST":
        request.user.delete()
        logout(request)
        messages.success(request, "Your personal data has been erased successfully.")
        return redirect('login')
    return render(request, 'registration/confirm_erasure.html')

@user_passes_test(lambda u: u.is_superuser)
def download_privacy_audit(request):
    buffer = io.BytesIO()
    p = canvas.Canvas(buffer, pagesize=letter)
    width, height = letter
    
    # Use your registered Hindi font for the title
    p.setFont("HindiFont", 16)
    p.drawString(50, height - 50, "DPDP Privacy Audit Report")
    
    y = height - 100
    logs = DataAccessLog.objects.all().order_by('-access_time')
    
    for log in logs:
        # Switch to HindiFont here so Hindi names are visible
        p.setFont("HindiFont", 10)
        
        log_text = f"{log.access_time.strftime('%Y-%m-%d')}: {log.accessed_by.username} accessed {log.target_user.username}"
        p.drawString(50, y, log_text)
        
        y -= 20
        if y < 50:
            p.showPage()
            p.setFont("HindiFont", 10) # Reset font on new page
            y = height - 50
    p.setFont("HindiFont", 20)
    p.drawString(100, 100, "रिकी टेस्ट")        
    p.save()
    buffer.seek(0)
    return FileResponse(buffer, as_attachment=True, filename=f'privacy_audit_{timezone.now().date()}.pdf')

@user_passes_test(lambda u: u.is_superuser)
def privacy_audit_report(request):
    logs = DataAccessLog.objects.all().order_by('-access_time')
    lang = request.session.get('lang', 'en')
    return render(request, 'privacy_audit.html', {'logs': logs, 'current_lang': lang})

@login_required
def download_db_backup(request):
    if request.session.get('active_role') != 'backup_user':
        return JsonResponse({"status": "error", "message": "Unauthorized access."}, status=403)

    try:
        host = os.getenv("POSTGRES_HOST")
        db = os.getenv("DB_NAME")
        db_user = os.getenv("DB_USER")

        if not all([host, db, db_user]):
            return JsonResponse({"status": "error", "message": "Database environment variables are missing."}, status=500)
        timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M")
        filename = f"~/backup_{timestamp}.sql"
        cmd = [
            "ssh",
            f"shannu-1@{host}",
            f"pg_dump -U {db_user} {db} -f {filename}"
        ]
        subprocess.run(cmd, check=True)
        return JsonResponse({
            "status": "success", 
            "message": f"Database backup created successfully at {filename}"
        })

    except subprocess.CalledProcessError as e:
        return JsonResponse({"status": "error", "message": "Backup command failed on the remote server."}, status=500)
    except Exception as e:
        return JsonResponse({"status": "error", "message": str(e)}, status=500)

@login_required
@user_passes_test(is_admin) 
def archive_user(request, user_id):  
    user_to_archive = get_object_or_404(CustomUser, id=user_id)
    
    if getattr(user_to_archive, 'id', None) == getattr(request.user, 'id', None):
        messages.error(request, "You cannot archive yourself.")
        return redirect('dashboard')

    empcode_val = None
    profile = getattr(user_to_archive, 'profile', None)
    if profile and getattr(profile, 'employee_code', None):
        try:
            empcode_val = int(profile.employee_code)
        except (TypeError, ValueError):
            empcode_val = None

    # Fallback: try to use username if it's numeric
    if empcode_val is None:
        try:
            empcode_val = int(user_to_archive.username)
        except (TypeError, ValueError):
            empcode_val = None

    employee = None
    if empcode_val is not None:
        employee = Employee.objects.filter(empcode=empcode_val).first()

    snapshot = {}
    if employee:
        snapshot = {
            "name": employee.ename,
            "designation": employee.designation,
            "status": employee.status,
            "last_updated": str(employee.lastupdate)
        }

    # 4. Create Archive Record
    ArchivedUser.objects.create(
        username=user_to_archive.username,
        email_hash=user_to_archive.email_hash,
        encrypted_email_data=user_to_archive.encrypted_email_data,
        original_user_id=user_to_archive.pk,
        employee_snapshot=json.dumps(snapshot) 
    )
    
    user_to_archive.is_active = False    
    user_to_archive.is_archived = True
    user_to_archive.save()

    messages.success(request, f"User {user_to_archive.username} has been archived successfully.")
    return redirect('dashboard')

@login_required
@user_passes_test(is_admin)
def unarchive_user(request, archive_id):
    """ Restores a user from Archive """
    archived_record = get_object_or_404(ArchivedUser, id=archive_id)
    
    try:
        user_to_restore = CustomUser.objects.get(id=archived_record.original_user_id)
        user_to_restore.is_active = True
        user_to_restore.is_archived = False
        user_to_restore.save()
        
        # Cleanup Archive Record
        archived_record.delete()
        
        messages.success(request, f"User {user_to_restore.username} has been unarchived/restored.")
        return redirect('dashboard')
        
    except CustomUser.DoesNotExist:
        # Fallback if the original user was actually deleted
        messages.error(request, "Original user record not found. Cannot restore.")
        return redirect('dashboard')
    
def _can_edit_profile(user, profile, pending_change_request=None):
    """Server-side authority for whether profile data may be changed."""
    if user_has_role(user, ['manager', 'admin']):
        return True

    if profile is None:
        return True

    if not profile.profile_updated:
        return True

    if pending_change_request is not None:
        return False
    # Allow editing if the profile was explicitly rejected so the user can
    # correct and resubmit their details. For pending approvals, editing
    # remains disallowed until HOD action or an approved change request.
    if getattr(profile, 'approval_status', None) == 'rejected':
        return True

    return profile.approval_status == 'approved' and user.is_edit_allowed


@login_required
def profile_view(request):
    """
    Profile view with correct lock/unlock workflow.
    
    Flow:
    1. NEW USER → form UNLOCKED, can fill and save
    2. After save -> form LOCKED (status=pending, no request box yet)
    3. HOD approves -> form LOCKED (status=approved, request box appears)
    4. User requests change → form LOCKED (pending_change_request exists)
    5. HOD approves change -> form UNLOCKED (is_edit_allowed=True)
    6. User saves approved changes -> form LOCKED again
    """
    from .models import Employee, Office, ProfileChangeRequest, QPRRecord
    from .employeeform import EmployeeForm

    lang = request.session.get('lang', 'en')
    user = request.user
    profile = getattr(user, 'profile', None)
    scoped_profile_fields = {'alternate_email', 'designation', 'highest_exam'}
    profile_approval_required = not user_has_role(user, ['manager', 'admin'])
    
    # Get change requests
    pending_change_request = ProfileChangeRequest.objects.filter(
        profile=profile,
        status='pending'
    ).first() if profile and profile_approval_required else None

    approved_change_request = ProfileChangeRequest.objects.filter(
        profile=profile,
        status='approved'
    ).order_by('-approved_at').first() if profile and profile_approval_required else None

    can_edit = _can_edit_profile(user, profile, pending_change_request)
    approved_fields = []
    if approved_change_request:
        approved_fields = [
            field for field in (approved_change_request.requested_fields or [])
            if field in scoped_profile_fields
        ]

    # ===============================
    # STATE FLAGS & LOCK LOGIC
    # ===============================

    is_approved = profile and profile.approval_status == "approved"


    # ===============================
    # POST LOGIC (SAVE CHANGES)
    # ===============================
    if request.method == 'POST':
        if not can_edit:
            messages.error(request, "Your profile is locked. Please request edit permission.", extra_tags='danger')
            return redirect('profile')

        if approved_change_request:
            if not approved_fields:
                messages.error(request, "No approved profile fields are available to edit. Please submit a new change request.", extra_tags='danger')
                return redirect('profile')

            if 'alternate_email' in approved_fields:
                profile.alternate_email = request.POST.get('alternate_email', '').strip()
                profile.save(update_fields=['alternate_email'])

            if 'designation' in approved_fields or 'highest_exam' in approved_fields:
                employee = Employee.objects.filter(empcode=profile.employee_code).first()
                if not employee:
                    messages.error(request, "Employee record not found. Please contact admin.")
                    return redirect('profile')

                update_fields = []
                if 'designation' in approved_fields:
                    employee.designation = request.POST.get('designation') or employee.designation
                    update_fields.append('designation')

                if 'highest_exam' in approved_fields:
                    employee.highest_exam = ",".join(request.POST.getlist("hindi_exam"))
                    update_fields.append('highest_exam')

                if update_fields:
                    employee.save(update_fields=update_fields)

            user.is_edit_allowed = False
            user.save(update_fields=['is_edit_allowed'])
            approved_change_request.status = 'completed'
            approved_change_request.save(update_fields=['status'])
            messages.success(request, "Approved profile changes saved successfully. Your profile is locked again.")
            return redirect('profile')

        # Get form data
        empcode = request.POST.get('empcode', '').strip()
        username = request.POST.get('username', '').strip()
        phone = request.POST.get('phone', '').strip()

        # SIMPLE VALIDATION: Just check required fields are filled
        # Employee code was already verified via API (fetchEmployeeData)
        if not empcode:
            messages.error(request, "Employee Code is required.")
            return redirect('profile')
        if not username:
            messages.error(request, "Employee Name is required.")
            return redirect('profile')
        if not phone:
            messages.error(request, "Phone Number is required.")
            return redirect('profile')

        # Email & Security
        new_email = request.POST.get('email', '').lower().strip()
        if not new_email:
            messages.error(request, "Email is required.", extra_tags='danger')
            return redirect('profile')

        email_hash = hashlib.sha256(new_email.encode()).hexdigest()

        # HOD Selection
        hod_name_post = request.POST.get('hod_name', '').strip()
        if not profile_approval_required and not hod_name_post:
            hod_name_post = "ADMIN"
        if profile_approval_required and not hod_name_post:
            messages.error(request, "HOD/Approver selection is required.")
            return redirect('profile')

        # Validate employee data before changing user/profile state.
        employee = Employee.objects.filter(empcode=empcode).first()
        form = EmployeeForm(request.POST, instance=employee)
        if not form.is_valid():
            error_messages = []
            for field, errors in form.errors.items():
                label = form.fields[field].label if field in form.fields else field
                error_messages.append(f"{label}: {', '.join(errors)}")
            details = " ".join(error_messages)
            messages.error(request, f"Form validation failed. {details}", extra_tags='danger')
            return redirect('profile')

        with transaction.atomic():
            # 1. Update User
            user.set_email(new_email)
            user.save()

            # 2. Update or Create Profile
            if not profile:
                from .models import UserProfile
                profile = UserProfile(user=user)

            profile.employee_code = empcode
            profile.phone = phone
            profile.office_code = request.POST.get('office_code', '').strip()
            profile.office_name = request.POST.get('office_name', '').strip()
            profile.office_state = request.POST.get('office_state', '').strip()
            profile.email = new_email
            profile.language_region = request.POST.get('language_region', '')
            profile.hod_name = hod_name_post
            profile.ip_number = request.POST.get('ip_number', '').strip()
            profile.alternate_email = request.POST.get('alternate_email', '').strip()

            if not profile_approval_required:
                profile.approval_status = "approved"
            elif profile.approval_status != "approved":
                profile.approval_status = "pending_admin" if hod_name_post == "ADMIN" else "pending"

            profile.profile_updated = True
            profile.save()

            # 3. Update Employee Model
            emp_instance = form.save(commit=False)
            emp_instance.highest_exam = ",".join(request.POST.getlist("hindi_exam"))
            emp_instance.super_annuation_date = form.cleaned_data.get('super_annuation_date')
            emp_instance.empcode = empcode
            emp_instance.save()
            if profile:
                profile.employee = emp_instance
                profile.save(update_fields=['employee'])

        # 4. Cleanup: Mark approved change request as completed
        if approved_change_request:
            approved_change_request.status = 'completed'
            approved_change_request.save()

        send_system_email(user, request, 'update')
        if profile_approval_required:
            messages.success(request, "Profile submitted successfully! It is now awaiting HOD approval.")
        else:
            messages.success(request, "Profile saved successfully.")
        return redirect('profile')

    # ===============================
    # gET LOGIC (PAGE LOAD)
    # ===============================
    empcode = profile.employee_code if profile else None
    employee = Employee.objects.filter(empcode=empcode).first() if empcode else None
    form = EmployeeForm(instance=employee)
    current_office_code = profile.office_code if profile else "0012"
    super_annuation_date_value = ''
    if employee:
        decrypted_super_annuation_date = employee.get_super_annuation_date()
        if decrypted_super_annuation_date:
            super_annuation_date_value = decrypted_super_annuation_date.strftime('%Y-%m-%d')

    # Context Generation
    offices = Office.objects.all()
    context = {
        'form': form,
        'employee': employee,
        'profile': profile,
        'offices': offices,
        'region_choices': QPRRecord.region_choices,
        
        # HOD/Approver info
        'available_hods': get_active_hods(current_office_code),
        'current_hod': profile.hod_name if profile else None,
        'ip_number': profile.ip_number if profile else '',
        'alternate_email': profile.alternate_email if profile else '',
        'super_annuation_date_value': super_annuation_date_value,

        # Flags for Template
        'can_edit': can_edit,
        'profile_approval_required': profile_approval_required,
        'profile_locked': not can_edit,
        'profile_approved': is_approved,
        'pending_change_request': pending_change_request,
        'has_pending_change_request': bool(pending_change_request),
        'has_approved_change_request': bool(approved_change_request),
        'approved_profile_fields': approved_fields,
        'approved_profile_fields_json': json.dumps(approved_fields),
        'profile_updated': profile.profile_updated if profile else False,
    }

    return render(request, 'profile.html', context)
@login_required
def approve_profile_change_hod(request, request_id):
    """HOD approves profile change request → unlock form"""

    from .models import ProfileChangeRequest
    from django.utils import timezone

    # ROLE CHECK
    if not user_has_role(request.user, ['hod', 'admin']):
        messages.error(request, "Unauthorized", extra_tags='danger')
        return redirect('qpr_hod_detail_list')

    change_request = get_object_or_404(ProfileChangeRequest, id=request_id)

    # Prevent re-processing
    if change_request.status != 'pending':
        messages.warning(request, "This request is already processed.")
        return redirect('qpr_hod_detail_list')

    #  HOD VALIDATION
    if change_request.hod != request.user and not request.user.is_staff:
        messages.error(request, "Not authorized for this request", extra_tags='danger')
        return redirect('qpr_hod_detail_list')

    # APPROVE REQUEST
    change_request.status = 'approved'
    change_request.approved_at = timezone.now()
    change_request.save()

    #  UNLOCK USER FORM
    user = change_request.profile.user
    user.is_edit_allowed = True
    user.save(update_fields=['is_edit_allowed'])

    messages.success(
        request,
        f"Edit request approved for {change_request.profile.name}. Form unlocked.",
        extra_tags='success'
    )

    return redirect('qpr_hod_detail_list')
@login_required
def reject_profile_change_hod(request, request_id):
    """HOD rejects profile change request → keep form locked"""

    from .models import ProfileChangeRequest
    from django.utils import timezone

    if request.method != 'POST':
        return redirect('qpr_hod_detail_list')

    #  ROLE CHECK
    if not user_has_role(request.user, ['hod', 'admin']):
        messages.error(request, "Unauthorized", extra_tags='danger')
        return redirect('qpr_hod_detail_list')

    change_request = get_object_or_404(ProfileChangeRequest, id=request_id)

    #  Prevent re-processing
    if change_request.status != 'pending':
        messages.warning(request, "This request is already processed.")
        return redirect('qpr_hod_detail_list')

    #  HOD VALIDATION
    if change_request.hod != request.user and not request.user.is_staff:
        messages.error(request, "Not authorized for this request", extra_tags='danger')
        return redirect('qpr_hod_detail_list')

    rejection_reason = request.POST.get('rejection_reason', '').strip()

    if not rejection_reason:
        messages.error(request, "Rejection reason is required", extra_tags='danger')
        return redirect('qpr_hod_detail_list')

    #  REJECT REQUEST
    change_request.status = 'rejected'
    change_request.approved_at = timezone.now()  # you can rename to reviewed_at later
    change_request.approval_comments = rejection_reason
    change_request.save()

    # KEEP USER LOCKED (explicit for clarity)
    user = change_request.profile.user
    user.is_edit_allowed = False
    user.save(update_fields=['is_edit_allowed'])

    messages.success(
        request,
        f"Edit request rejected for {change_request.profile.name}",
        extra_tags='success'
    )

    return redirect('qpr_hod_detail_list')

@login_required
def freeze_profile(request):
    lang = request.session.get('lang', 'en')
    user = request.user
    user.is_frozen = True
    user.save()
    send_system_email(user, request, 'freeze')
    messages.success(request, translate_text("Your profile has been frozen.", lang))
    return redirect('dashboard')

@login_required
def request_edit(request):
    lang = request.session.get('lang', 'en')
    user = request.user
    if not user.is_frozen: return redirect('dashboard')
    
    # Check if already pending
    pending_request = EditRequest.objects.filter(
        user=request.user,
        request_type='profile',
        status='pending'
    ).exists()
    
    if pending_request:
        messages.warning(request, translate_text("You already have a pending profile edit request.", lang))
    else:
        # Try to find the user's manager/HOD and send the request there so manager can unlock
        profile = getattr(request.user, 'profile', None)
        hod_name = profile.hod_name if profile else None
        manager_user = None
        if hod_name:
            # Try to locate a CustomUser whose profile.hod_name or profile.name matches
            manager_user = CustomUser.objects.filter(profile__hod_name__iexact=hod_name).first()
            if not manager_user:
                manager_user = CustomUser.objects.filter(profile__name__iexact=hod_name).first()

        if manager_user:
            ManagerRequest.objects.create(
                hod=manager_user,
                user=request.user,
                request_type='profile',
                reason='User requested permission to edit frozen profile',
                status='pending'
            )
            messages.success(request, translate_text("Profile edit request sent to your manager for approval.", lang))
            # Notify manager
            msg = f"User {user.username} has requested permission to edit their profile."
            send_system_email(manager_user, request, 'manager_alert', extra_context={'body_text': msg})
        else:
            # Fallback to admin approval if no manager found
            EditRequest.objects.create(
                user=request.user,
                request_type='profile',
                requested_data={'reason': 'User requested permission to edit profile'},
                reason='User requested permission to edit frozen profile',
                status='pending'
            )
            messages.success(request, translate_text("Profile edit request sent to admin for approval.", lang))
            admin = CustomUser.objects.filter(roles__name='admin').first()
            if admin:
                msg = f"User {user.username} has requested permission to edit their profile."
                send_system_email(admin, request, 'manager_alert', extra_context={'body_text': msg})
    
    return redirect('dashboard')

@login_required
def user_office_form(request):
    profile = request.user.profile
    if request.method == 'POST':
        office_name = request.POST.get('office_name', '')
        office_code = request.POST.get('office_code', '')
        if not office_name or not office_code:
            messages.error(request, 'Office name and code are required')
        else:
            profile.office_name = office_name
            profile.office_code = office_code
            profile.save()
            messages.success(request, 'Office details updated successfully!')
            return redirect('qpr_user_dashboard')
    context = {'profile': profile}
    return render(request, 'user_office_form.html', context)

# ==================== UNIFIED DASHBOARD VIEWS ====================

@login_required
def user_dashboard(request):
    """User Dashboard View - Unified"""
    profile, created = UserProfile.objects.get_or_create(
        user=request.user,
        defaults={"employee_code": f"EMP{getattr(request.user, 'id', '')}"}
    )
    profile.refresh_from_db()
    # If profile not completed, redirect user to fill profile first (only first time)
    if not profile.profile_updated:
        return redirect('qpr_user_profile')
    qpr_records = QPRRecord.objects.filter(user=request.user)
    # Consider 'submitted' on dashboard only when user has submitted a DAILY QPR for today
    today = timezone.localdate()
    submitted_qprs = QPRRecord.objects.filter(user=request.user, is_submitted=True, frequency__iexact='daily', period_start=today).count()
    
    # Get list of available HODs for dropdown
    available_hods = get_active_hods(profile.office_code)
    
    # Check if user has HOD or Manager roles (disable HOD selection if they do)
    is_hod_or_manager = user_has_role(request.user, ['hod', 'manager'])
    # Compute role-based UI controls without changing any role models/permissions.
    roles = set(user_get_all_roles(request.user))
    roles_up = {r.upper() for r in roles}
    has_user = 'USER' in roles_up
    has_manager = 'MANAGER' in roles_up
    has_admin = 'ADMIN' in roles_up
    has_hod = 'HOD' in roles_up

    disable_user_dashboard_actions = (
        has_user and (has_manager or has_admin) and (not has_hod)
    )

    context = {
        'role': 'user',  # Explicitly set role for template to avoid showing other roles' content
        'profile': profile,
        'profile_status': 'Updated' if profile.profile_updated else 'Needs Update',
        'qpr_submitted': submitted_qprs > 0,
        'qpr_count': qpr_records.count(),
        'user': request.user,
        'available_hods': available_hods,
        'current_hod': profile.hod_name or '',
        'is_hod_or_manager': is_hod_or_manager,
        'disable_user_dashboard_actions': disable_user_dashboard_actions,
        'has_manager': has_manager,
        'has_admin': has_admin,
    }
    response = render(request, 'dashboard.html', context)
    response['Cache-Control'] = 'no-cache, no-store, must-revalidate'
    response['Pragma'] = 'no-cache'
    response['Expires'] = '0'
    return response 


@login_required
def manager_qpr_view(request, id=None):
    # Role check
    if not user_has_role(request.user, 'manager'):
        return HttpResponseForbidden("Manager role required")

    from .forms import ManagerQPRForm
    from .models import ManagerQPR

    instance = None
    if id:
        instance = get_object_or_404(ManagerQPR, pk=id)

    if request.method == 'POST':
        if instance:
            form = ManagerQPRForm(request.POST, instance=instance)
        else:
            form = ManagerQPRForm(request.POST)

        if form.is_valid():
            quarter = form.cleaned_data.get('quarter')
            # If creating new and one already exists for this quarter, show error
            if not instance and ManagerQPR.objects.filter(user=request.user, quarter=quarter).exists():
                messages.error(request, "Manager QPR for this quarter has already been filled.")
            else:
                # Only save if it's an edit OR if no duplicate exists
                obj = form.save(commit=False)
                obj.user = request.user
                # Ensure submitted flag/time on save
                obj.is_submitted = True
                obj.submitted_at = timezone.now()
                obj.save()
                messages.success(request, "Manager QPR saved successfully.")
                return redirect('manager_qpr_detail', id=obj.id)
    else:
        if instance:
            form = ManagerQPRForm(instance=instance)
        else:
            form = ManagerQPRForm()

    return render(request, 'qpr/manager_qpr_form.html', {'form': form, 'instance': instance})


@login_required
def manager_qpr_detail(request, id):
    from .forms import ManagerQPRForm
    from .models import ManagerQPR
    obj = get_object_or_404(ManagerQPR, id=id)
    # Only owner or staff can view
    if obj.user != request.user and not (request.user.is_staff or user_has_role(request.user, 'admin')):
        return HttpResponseForbidden()

    # Build a form populated with the instance and disable all inputs for readonly view
    form = ManagerQPRForm(instance=obj)
    for name in form.fields:
        try:
            form.fields[name].widget.attrs['disabled'] = 'disabled'
        except Exception:
            pass

    return render(request, 'qpr/manager_qpr_form.html', {'form': form, 'instance': obj, 'readonly': True})


@login_required
def manager_section11_select_texts(request, manager_qpr_id=None):
    """Manager selects which users' Section 11 texts to include in their aggregated report."""
    if not user_has_role(request.user, 'manager'):
        return HttpResponseForbidden("Manager role required")

    from .models import ManagerQPR, QPRRecord, Section11SpecificAchievementsData, CustomUser
    import json

    # Get manager's office code
    manager_office = getattr(request.user.profile, 'office_code', None)
    if not manager_office:
        messages.error(request, "Your profile doesn't have an office code configured.")
        return redirect('manager_qpr_list')

    # Get or create manager QPR record
    manager_qpr = None
    if manager_qpr_id:
        manager_qpr = get_object_or_404(ManagerQPR, pk=manager_qpr_id, user=request.user)

    if request.method == 'POST':
        # Process selected user texts
        selected_user_ids = request.POST.getlist('selected_users')
        
        # Get all users in manager's office
        office_users = CustomUser.objects.filter(
            profile__office_code=manager_office,
            is_active=True
        ).exclude(id=request.user.id)  # Exclude the manager themselves

        # Get Section11 texts from selected users' latest QPRs
        texts_by_field = {
            'innovative_work': [],
            'special_events': [],
            'hindi_medium_works': []
        }

        for user_id in selected_user_ids:
            try:
                user_id = int(user_id)
                user = office_users.get(id=user_id)
                
                # Get user's most recent submitted QPRRecord for this quarter/year
                latest_qpr = QPRRecord.objects.filter(
                    user=user,
                    quarter=manager_qpr.quarter if manager_qpr else None,
                    year=manager_qpr.financial_year if manager_qpr else None,
                    is_submitted=True
                ).order_by('-updated_at').first()

                if latest_qpr and hasattr(latest_qpr, 'section11'):
                    s11 = latest_qpr.section11
                    user_display = f"{user.first_name} {user.last_name}" if user.first_name else user.username
                    
                    if s11.innovative_work:
                        texts_by_field['innovative_work'].append(f"[{user_display}]: {s11.innovative_work}")
                    if s11.special_events:
                        texts_by_field['special_events'].append(f"[{user_display}]: {s11.special_events}")
                    if s11.hindi_medium_works:
                        texts_by_field['hindi_medium_works'].append(f"[{user_display}]: {s11.hindi_medium_works}")
            except (ValueError, CustomUser.DoesNotExist):
                continue

        # Join texts with line breaks
        aggregated_data = {
            'innovative_work': '\n\n'.join(texts_by_field['innovative_work']),
            'special_events': '\n\n'.join(texts_by_field['special_events']),
            'hindi_medium_works': '\n\n'.join(texts_by_field['hindi_medium_works'])
        }

        # Save to manager QPR if editing, otherwise show in flash
        if manager_qpr:
            manager_qpr.s11_innovative_work = aggregated_data['innovative_work']
            manager_qpr.s11_special_events = aggregated_data['special_events']
            manager_qpr.s11_hindi_medium_works = aggregated_data['hindi_medium_works']
            manager_qpr.save()
            messages.success(request, "Section 11 texts aggregated and saved successfully.")
            return redirect('manager_qpr_detail', id=manager_qpr.id)
        else:
            # Store in session for display
            request.session['section11_preview'] = aggregated_data
            messages.success(request, "Section 11 texts aggregated. Review below:")

    # Get all users in this office
    office_users = CustomUser.objects.filter(
        profile__office_code=manager_office,
        is_active=True
    ).exclude(id=request.user.id).select_related('profile')

    # Get their Section11 data with latest QPRs
    users_section11 = []
    quarter = manager_qpr.quarter if manager_qpr else None
    financial_year = manager_qpr.financial_year if manager_qpr else None

    for user in office_users:
        # Get user's most recent submitted QPRRecord
        latest_qpr = QPRRecord.objects.filter(
            user=user,
            quarter=quarter,
            year=financial_year,
            is_submitted=True
        ).order_by('-updated_at').first() if (quarter and financial_year) else None

        if latest_qpr and hasattr(latest_qpr, 'section11'):
            s11 = latest_qpr.section11
            if s11.innovative_work or s11.special_events or s11.hindi_medium_works:
                users_section11.append({
                    'user': user,
                    'section11': s11,
                    'qpr': latest_qpr
                })

    context = {
        'manager_qpr': manager_qpr,
        'users_section11': users_section11,
        'manager_office': manager_office,
        'section11_preview': request.session.pop('section11_preview', None)
    }

    return render(request, 'qpr/manager_section11_select.html', context)



@login_required
def admin_qpr_view(request, id=None):
    # Role check
    if not user_has_role(request.user, 'admin'):
        return HttpResponseForbidden("Admin role required")

    from .forms import AdminQPRForm
    from .models import AdminQPR

    instance = None
    if id:
        instance = get_object_or_404(AdminQPR, pk=id)

    if request.method == 'POST':
        if instance:
            form = AdminQPRForm(request.POST, instance=instance)
        else:
            form = AdminQPRForm(request.POST)

        if form.is_valid():
            quarter = form.cleaned_data.get('quarter')
            # If creating new and one already exists for this quarter, show error
            if not instance and AdminQPR.objects.filter(user=request.user, quarter=quarter).exists():
                messages.error(request, "Admin QPR for this quarter has already been filled.")
            else:
                # Only save if it's an edit OR if no duplicate exists
                obj = form.save(commit=False)
                obj.user = request.user
                # Ensure submitted flag/time on save
                obj.is_submitted = True
                obj.submitted_at = timezone.now()
                obj.save()
                messages.success(request, "Admin QPR saved successfully.")
                return redirect('admin_qpr_detail', id=obj.id)
    else:
        if instance:
            form = AdminQPRForm(instance=instance)
        else:
            form = AdminQPRForm()

    return render(request, 'qpr/admin_qpr_form.html', {'form': form, 'instance': instance})


@login_required
def admin_qpr_detail(request, id):
    from .forms import AdminQPRForm
    from .models import AdminQPR
    obj = get_object_or_404(AdminQPR, id=id)
    # Only owner or staff can view
    if obj.user != request.user and not (request.user.is_staff or user_has_role(request.user, 'admin')):
        return HttpResponseForbidden()

    # Build a form populated with the instance and disable all inputs for readonly view
    form = AdminQPRForm(instance=obj)
    for name in form.fields:
        try:
            form.fields[name].widget.attrs['disabled'] = 'disabled'
        except Exception:
            pass

    return render(request, 'qpr/admin_qpr_form.html', {'form': form, 'instance': obj, 'readonly': True})

@login_required
def qpr_hod_dashboard(request):
    """HOD Dashboard - Department overview and employee statistics"""

    if not user_has_role(request.user, 'hod'):
        return redirect('/')

    lang = request.session.get('lang', 'en')

    from .models import ProfileChangeRequest, UserProfile
    from django.db.models import Q

    # ===============================
    # CHANGE REQUESTS (FOR DASHBOARD)
    # ===============================
    profile_change_requests = ProfileChangeRequest.objects.filter(
        hod=request.user,
        status='pending'
    ).select_related('profile', 'profile__user').order_by('-requested_at')

    # ===============================
    # BASIC INFO
    # ===============================
    current_quarter = get_current_quarter()
    current_year = get_current_year_label()

    hod_profile = UserProfile.objects.select_related('user').get(user=request.user)
    hod_name = (hod_profile.hod_name or hod_profile.name or "").strip()


    if hod_name:
        user_role_q = Q(roles__name='user') | Q(user__roles__name='user')

        # Only include approved employees under this HOD (exclude pending/rejected)
        users_under_hod = UserProfile.objects.filter(
            ((user_role_q & Q(hod_name__iexact=hod_name)) |
            Q(user=request.user)) & Q(approval_status__iexact='approved')
        ).distinct()
    else:
        users_under_hod = UserProfile.objects.filter(user=request.user, approval_status__iexact='approved').distinct()

    total_users = users_under_hod.count()
    qpr_submitted_count = 0

    today = timezone.localdate()
    for up in users_under_hod:
        try:
            submitted_today = up.user.qpr_records.filter(
                frequency__iexact='daily',
                period_start=today,
                is_submitted=True
            ).exists()
            if submitted_today:
                qpr_submitted_count += 1
        except Exception:
            continue

    qpr_pending = total_users - qpr_submitted_count
    profile_updated_count = users_under_hod.filter(profile_updated=True).count()

    pending_approvals = UserProfile.objects.filter(
            approval_status='pending'
        ).filter(
            Q(hod_name__iexact=hod_name) |
            Q(hod_name__iexact=hod_profile.employee_code) |
            Q(hod_name=str(hod_profile.employee_code))
        ).select_related('user', 'employee')
    context = {
        'role': 'hod',
        'total_users': total_users,
        'qpr_submitted': qpr_submitted_count,
        'qpr_pending': qpr_pending,
        'profile_updated': profile_updated_count,
        'hod_name': hod_name,

        # IMPORTANT (for dashboard UI)
        'profile_change_requests': profile_change_requests,

        'current_lang': lang,
        'current_quarter': current_quarter,
        'current_year': current_year,
        'pending_approvals': pending_approvals,
    }

    response = render(request, 'qpr/hod_dashboard.html', context)

    # ===============================
    # NO CACHE
    # ===============================
    response['Cache-Control'] = 'no-cache, no-store, must-revalidate'
    response['Pragma'] = 'no-cache'
    response['Expires'] = '0'
    print("HOD name:", hod_name)
    print("HOD empcode:", hod_profile.employee_code)
    print("Pending count:", pending_approvals.count())


    return response
@login_required
def manager_dashboard(request):
    """Manager Dashboard - Manage system access and employee records"""
    if not (user_has_role(request.user, ['manager', 'admin']) or request.user.is_superuser):
        return redirect('/')
    
    manager_office = getattr(request.user.profile, 'office_code', None)

    users = CustomUser.objects.select_related('profile').filter(profile__office_code=manager_office).order_by('-date_joined')

    # Get employee codes from users in this office (only numeric codes)
    office_employee_codes_qs = CustomUser.objects.filter(profile__office_code=manager_office).values_list('profile__employee_code', flat=True)
    office_employee_codes = []
    for code in office_employee_codes_qs:
        if code is None:
            continue
        # Accept integers or strings that represent integers; skip non-numeric values
        try:
            office_employee_codes.append(int(str(code).strip()))
        except (ValueError, TypeError):
            continue

    # Fetch employee records directly by numeric empcode
    if office_employee_codes:
        raw_employees = Employee.objects.filter(empcode__in=office_employee_codes).order_by('-lastupdate')
    else:
        raw_employees = Employee.objects.none()
    
    employee_data = []
    
    for emp in raw_employees:
        # Get the user by employee_code (which is the empcode)
        user = CustomUser.objects.filter(profile__employee_code=emp.empcode).first()

        # --- QPR DATA ---
        qpr_status_text = "Not Started"
        qpr_is_submitted = False
        latest_qpr_id = None
        qpr_last_updated = None
        
        # Get ID safely
        linked_user_id = getattr(user, 'id', None) if user else None
        
        if user:
            latest_qpr = QPRRecord.objects.filter(user=user).order_by('-updated_at').first()
            if latest_qpr:
                qpr_is_submitted = latest_qpr.is_submitted
                qpr_status_text = "Submitted" if qpr_is_submitted else "Draft"
                latest_qpr_id = getattr(latest_qpr, 'id', None)
                qpr_last_updated = latest_qpr.updated_at

        employee_data.append({
            'empcode': emp.empcode,
            'name': emp.ename,
            'designation': emp.designation,
            'hname': emp.hname,
            'user_id': linked_user_id,
            'status': emp.status,
            'lastupdate': emp.lastupdate,
            'qpr_status': qpr_status_text,
            'qpr_is_submitted': qpr_is_submitted,
            'qpr_id': latest_qpr_id,
            'qpr_last_updated': qpr_last_updated
        })

    # Pending profile edit requests targeted to this manager
    pending_profile_requests = ManagerRequest.objects.filter(hod=request.user, request_type='profile', status='pending')
    
    # QPR edit requests (pending and approved) from employees in this manager's office
    manager_office = getattr(request.user.profile, 'office_code', None)
    pending_qpr_edits = []
    edit_requests_by_user = {}  # Map user_id -> EditRequest for quick lookup (pending or approved)
    
    if manager_office:
        # Get both pending and approved edit requests
        edit_requests = EditRequest.objects.filter(
            request_type='qpr',
            status__in=['pending', 'approved']  # Include both pending and approved
        ).select_related('user').filter(
            user__profile__office_code=manager_office
        )
        
        # Filter pending ones for the quick actions section
        pending_qpr_edits = [req for req in edit_requests if req.status == 'pending']
        
        # Build dictionary for template lookup (stores latest edit request per user)
        for req in edit_requests:
            # Only store if newer or first one
            if req.user_id not in edit_requests_by_user or req.created_at > edit_requests_by_user[req.user_id].created_at:
                edit_requests_by_user[req.user_id] = req
    
    # Enrich employee_data with edit request info
    for emp in employee_data:
        emp['pending_edit_request'] = None
        emp['approved_edit_request'] = None
        
        if emp['user_id'] in edit_requests_by_user:
            req = edit_requests_by_user[emp['user_id']]
            if req.status == 'pending':
                emp['pending_edit_request'] = req
            elif req.status == 'approved':
                emp['approved_edit_request'] = req
    
    context = {
        'users': users,
        'employees': employee_data,
        'pending_profile_requests': pending_profile_requests,
        'pending_qpr_edits': pending_qpr_edits,
    }
    return render(request, 'manager_dashboard.html', context)
@login_required
def admin_dashboard(request):
    if user_role(request.user) != 'admin': return redirect('/')
    admin_state = request.user.profile.office_state
    if not admin_state:
        messages.warning(request, "Mandatory: You must set your Office State in your profile before accessing the Admin Dashboard.")
        return redirect('profile')
    
    users = CustomUser.objects.filter(is_active=True, is_archived=False, profile__office_state=admin_state).order_by('-date_joined')
    archived_users = ArchivedUser.objects.all().order_by('-archived_at')

    current_quarter = get_current_quarter()
    current_year = get_current_year_label()
    
    hod_stats = []
    hods = UserProfile.objects.filter(roles__name='hod', office_state=admin_state).order_by('name')
    for hod_profile in hods:
        hod_key = hod_profile.hod_name or hod_profile.name or hod_profile.employee_code
        hod_display = hod_profile.name or hod_key or 'UNKNOWN'
        # Count only approved users for admin HOD statistics
        users_under_hod = UserProfile.objects.filter(roles__name='user', hod_name__iexact=hod_key, approval_status__iexact='approved',office_state=admin_state)
        total_users = users_under_hod.count()
        profile_complete = sum(1 for p in users_under_hod if p.profile_updated)
        qpr_complete = sum( 1 for p in users_under_hod if QPRRecord.objects.filter( user=p.user, quarter=current_quarter, year=current_year, is_submitted=True ).exists())
        completion_pct = int((qpr_complete / total_users) * 100) if total_users > 0 else 0
        hod_stats.append({
            'hod_name': str(hod_display).upper(),
            'total_employees': total_users,
            'profile_completed': profile_complete,
            'qpr_completed': qpr_complete,
            'completion_percentage': completion_pct,
        })
    # Consider only approved users when deriving unique HOD names
    unique_hod_names = set(UserProfile.objects.filter(roles__name='user', approval_status__iexact='approved', office_state=admin_state).exclude(hod_name__isnull=True).values_list('hod_name', flat=True))
    actual_hod_names = set(UserProfile.objects.filter(roles__name='hod', office_state=admin_state).values_list('hod_name', flat=True))
    uncovered = unique_hod_names - actual_hod_names
    for hod_name in sorted(uncovered):
        users_under_hod = UserProfile.objects.filter(roles__name='user', hod_name__iexact=hod_name, approval_status__iexact='approved', office_state=admin_state)
        total_users = users_under_hod.count()
        qpr_complete = sum(1 for p in users_under_hod if QPRRecord.objects.filter(user=p.user, status='Submitted').exists())
        completion_pct = int((qpr_complete / total_users) * 100) if total_users > 0 else 0
        hod_stats.append({
            'hod_name': str(hod_name).upper(),
            'total_employees': total_users,
            'profile_completed': sum(1 for p in users_under_hod if p.profile_updated),
            'qpr_completed': qpr_complete,
            'completion_percentage': completion_pct,
        })
    # 3. Pending Requests
    pending_requests = ManagerRequest.objects.filter(status='pending', hod__roles__name='user', hod__profile__office_state=admin_state)
    context = {
        'role': 'admin',  # Explicitly set role for template to avoid showing other roles' content
        'hod_stats': hod_stats, 
        'manager_requests': pending_requests,
        'users': users,
        'archived_users': archived_users
    }
    response = render(request, 'dashboard.html', context) # Renders UNIFIED DASHBOARD
    response['Cache-Control'] = 'no-cache, no-store, must-revalidate'
    response['Pragma'] = 'no-cache'
    response['Expires'] = '0'
    return response

# ==================== ADMIN/MANAGER ACTIONS (RESTORED) ====================

@login_required
def admin_create_hod(request):
    if not user_has_role(request.user, 'admin'): return redirect('/')
    admin_state = request.user.profile.office_state
    found_name = ''
    if request.method == 'POST':
        emp_code = request.POST.get('emp_code', '').strip()
        if not emp_code:
            messages.error(request, 'Employee code is required')
        else:
            # Check if employee code exists in registered users
            try:
                profile = UserProfile.objects.get(employee_code=emp_code, office_state=admin_state)
                display_name = profile.name or profile.user.get_full_name() or profile.user.username
                found_name = display_name
                if profile.roles.filter(name='hod').exists():
                    messages.error(request, 'This user is already assigned a HOD role')
                else:
                    # Assign HOD role (sync to both profile and user)
                    hod_role = Role.objects.get(name='hod')
                    user_role_obj = Role.objects.get(name='user')
                    profile.roles.add(hod_role, user_role_obj)
                    profile.approval_status = 'approved'
                    # Ensure CustomUser.roles is in sync
                    try:
                        profile.user.roles.add(hod_role, user_role_obj)
                        profile.user.save()
                    except Exception:
                        pass
                    profile.hod_name = emp_code
                    profile.profile_updated = True
                    profile.save()
                    messages.success(request, f'HOD {display_name} created!')
                    return redirect('qpr_admin_dashboard')
            except UserProfile.DoesNotExist:
                messages.error(request, 'User has not registered or entered employee code is incorrect')
    return render(request, 'qpr/admin_create_hod.html', {'found_name': found_name})

@login_required
def admin_create_manager(request):
    if not user_has_role(request.user, 'admin'): return redirect('/')
    admin_state = request.user.profile.office_state
    found_name = ''
    if request.method == 'POST':
        emp_code = request.POST.get('emp_code', '').strip()
        if not emp_code:
            messages.error(request, 'Employee code is required')
        else:
            try:
                profile = UserProfile.objects.get(employee_code=emp_code, office_state=admin_state)
                display_name = profile.name or profile.user.get_full_name() or profile.user.username
                found_name = display_name
                if profile.roles.filter(name='manager').exists():
                    messages.error(request, 'This user is already assigned a Manager role')
                else:
                    manager_role = Role.objects.get(name='manager')
                    user_role_obj = Role.objects.get(name='user')
                    profile.roles.add(manager_role, user_role_obj)
                    profile.approval_status = 'approved'
                    try:
                        profile.user.roles.add(manager_role, user_role_obj)
                        profile.user.save()
                    except Exception:
                        pass
                    profile.profile_updated = True
                    profile.save()
                    messages.success(request, f'Manager {display_name} created!')
                    return redirect('qpr_admin_dashboard')
            except UserProfile.DoesNotExist:
                messages.error(request, 'User has not registered or entered employee code is incorrect')
    return render(request, 'qpr/admin_create_manager.html', {'found_name': found_name})


def admin_api_get_employee_details(request):
    """API endpoint to fetch employee details by employee code"""
    admin_state = request.user.profile.office_state
    emp_code = request.GET.get('emp_code', '').strip()
    
    if not emp_code:
        return JsonResponse({'error': 'Employee code is required'}, status=400)
    
    try:
        profile = UserProfile.objects.get(employee_code=emp_code, office_state=admin_state)
        roles = list(profile.roles.values_list('name', flat=True))
        display_name = profile.name or profile.user.get_full_name() or profile.user.username
        return JsonResponse({
            'success': True,
            'name': display_name,
            'employee_code': profile.employee_code,
            'roles': roles or ['user']
        })
    except UserProfile.DoesNotExist:
        return JsonResponse({
            'error': 'User has not registered or entered employee code is incorrect'
        }, status=404)

@login_required
def api_create_office(request):
    if request.method != 'POST':
        return redirect('qpr_admin_dashboard')
    if not user_has_role(request.user, 'admin'):
        messages.error(request, 'Permission denied')
        return redirect('qpr_admin_dashboard')
    admin_state = getattr(request.user.profile, 'office_state', '').strip()
    if not admin_state:
        messages.error(request, "Your profile is missing a state. Please update your profile first.")
        return redirect('profile')
    code = request.POST.get('office_code', '').strip()
    name = request.POST.get('office_name', '').strip()
    if not code or not name:
        messages.error(request, 'Office code and name are required')
        return redirect('qpr_admin_dashboard')

    from .models import Office
    office, created = Office.objects.get_or_create(
        code=code, 
        defaults={'name': name,'state': admin_state}
    )
    if not created:
        messages.error(request, 'Office code already exists')
        return redirect('qpr_admin_dashboard')

    messages.success(request, f'Office {office.code} - {office.name} created for {admin_state}')
    return redirect('qpr_admin_dashboard')


def api_list_offices(request):
    """Return list of offices for dropdowns"""
    from .models import Office
    offices = list(Office.objects.all().values('code', 'name'))
    return JsonResponse({'offices': offices})

@login_required
def admin_approve_request(request, request_id):
    if not user_has_role(request.user, 'admin'): return redirect('/')
    try:
        req = ManagerRequest.objects.get(id=request_id)
        if request.method == 'POST':
            action = request.POST.get('action')
            if action == 'approve':
                req.status = 'approved'
                req.save()
                # ManagerRequest approved — do not set global is_edit_allowed here.
                messages.success(request, 'Approved!')
            elif action == 'reject':
                req.status = 'rejected'
                req.save()
                messages.success(request, 'Rejected!')
        return redirect('qpr_admin_dashboard')
    except ManagerRequest.DoesNotExist:
        return redirect('qpr_admin_dashboard')

@login_required
def admin_employee_list(request):
    if not user_has_role(request.user, 'admin'): return redirect('/')
    admin_state = request.user.profile.office_state
    employee_code_filter = request.GET.get('employee_code', '').strip()
    name_filter = request.GET.get('name', '').strip()
    quarter_filter = request.GET.get('quarter', '').strip()
    year_filter = request.GET.get('year', '').strip()
    if not quarter_filter:
        quarter_filter = get_current_quarter()
    if not year_filter:
        year_filter = get_current_year_label()

    hods = UserProfile.objects.filter(roles__name='hod', office_state=admin_state).order_by('name')
    hod_groups = []
    
    # Collect all unique quarters and years for filter dropdowns
    all_qpr_records = QPRRecord.objects.all()
    all_quarters = sorted(set(all_qpr_records.values_list('quarter', flat=True).filter(quarter__isnull=False)))
    all_years = sorted(set(all_qpr_records.values_list('year', flat=True).filter(year__isnull=False)), reverse=True)

    current_year = get_current_year_label()

    if current_year not in all_years:
        all_years.insert(0, current_year)
    
    from .models import Employee
    for hod_profile in hods:
        users_under_hod = UserProfile.objects.filter(roles__name='user', hod_name=hod_profile.hod_name).order_by('name')
        user_details = []
        for user_profile in users_under_hod:
            # employee_code may be None; normalize to empty string
            emp_code_val = (user_profile.employee_code or '').strip()
            if employee_code_filter and employee_code_filter.lower() not in emp_code_val.lower():
                continue

            user_name = (user_profile.name or user_profile.user.get_full_name() or user_profile.user.username) or ''

            # Try to fill missing name from Employee table when profile name is blank or 'None'
            emp_record = None
            if (not user_name) or user_name.strip().lower() in ['', 'none']:
                # Attempt integer conversion for codes with leading zeros (e.g., '003')
                if emp_code_val:
                    try:
                        emp_int = int(emp_code_val)
                        emp_record = Employee.objects.filter(empcode=emp_int).first()
                    except Exception:
                        emp_record = Employee.objects.filter(empcode=emp_code_val).first()
                # fallback: try to match by username
                if not emp_record:
                    emp_record = Employee.objects.filter(empcode=user_profile.user.username).first()
                if emp_record and emp_record.ename:
                    user_name = emp_record.ename

            if name_filter and name_filter.lower() not in (user_name or '').lower():
                continue

            qpr_record = QPRRecord.objects.filter( user=user_profile.user, quarter=quarter_filter, year=year_filter ).first()

            # Fill office info: prefer profile, then latest QPR, else Employee.hname
            office_name_val = user_profile.office_name or (qpr_record.officeName if qpr_record else '')
            office_code_val = user_profile.office_code or (qpr_record.officeCode if qpr_record else '')

            if (not office_name_val or office_name_val.strip() == '') and emp_record:
                office_name_val = getattr(emp_record, 'hname', '') or office_name_val

            user_details.append({
                'emp_code': user_profile.employee_code,
                'name': user_name,
                'email': user_profile.user.email,
                'office_name': office_name_val or 'Not Set',
                'office_code': office_code_val or 'Not Set',
                'quarter': quarter_filter,
                'year': year_filter,
                'qpr_status': qpr_record.status if qpr_record else 'Not Submitted',
            })
        if user_details:
            hod_groups.append({
                'hod_name': hod_profile.hod_name, 
                'hod_email': hod_profile.user.email,
                'hod_emp_code': hod_profile.employee_code,
                'user_count': len(user_details), 
                'users': user_details
            })
    
    context = {
        'hod_groups': hod_groups,
        'all_quarters': all_quarters,
        'all_years': all_years,
        'quarter_filter': quarter_filter,
        'year_filter': year_filter,
        'employee_code_filter': employee_code_filter,
        'name_filter': name_filter,
    }
    return render(request, 'qpr/admin_employee_list.html', context)

@user_passes_test(lambda u: user_has_role(u, ['hod', 'admin']))
def update_designation(request, user_id):
    if request.method == "POST":
        target_user = get_object_or_404(CustomUser, id=user_id)
        new_desig = request.POST.get('designation')
        emp = Employee.objects.filter(empcode=target_user.username).first()
        if emp:
            emp.designation = new_desig
            emp.save()
            messages.success(request, "Designation updated.")
        else:
            messages.error(request, "Employee record not found.")
    return redirect('manager_dashboard')

@user_passes_test(lambda u: u.is_authenticated and (user_has_role(u, ['hod', 'admin']) or u.is_superuser))
def manage_user_action(request, user_id, action):
    if action == 'unlock_qpr':
        if not (user_has_role(request.user, ['manager', 'admin']) or request.user.is_superuser):
            messages.error(request, translate_text("Unauthorized", request.session.get('lang', 'en')))
            return redirect('manager_dashboard')
        try:
            qpr = QPRRecord.objects.get(id=user_id)
            qpr.is_submitted = False
            qpr.status = 'Draft'
            qpr.save()
            messages.success(request, translate_text("QPR Form unlocked successfully.", request.session.get('lang', 'en')))
        except QPRRecord.DoesNotExist:
            messages.error(request, translate_text("QPR Record not found.", request.session.get('lang', 'en')))
        return redirect('manager_dashboard')

    try:
    # First try treating it as user id
        target_user = CustomUser.objects.get(id=user_id)
    except CustomUser.DoesNotExist:
    # If not found, treat it as employee_code
        profile = get_object_or_404(UserProfile, employee_code=user_id)
        target_user = profile.user

    lang = request.session.get('lang', 'en')
    
    if action in ['archive', 'unarchive']:
        if not (user_has_role(request.user, ['admin']) or request.user.is_superuser):
            messages.error(request, translate_text("Only Admins can perform this action.", lang))
            return redirect('manager_dashboard')
        if not request.user.is_superuser:
            admin_state = request.user.profile.office_state
            target_state = getattr(target_user.profile, 'office_state', None)
            if admin_state != target_state:
                messages.error(request, translate_text("You can only manage users within your own state.", lang))
                return redirect('dashboard')
        
        if action == 'archive':
            target_user.is_active = False
            target_user.is_archived = True
            target_user.save()
            messages.success(request, translate_text("User archived.", lang))
        elif action == 'unarchive':
            target_user.is_active = True
            target_user.is_archived = False
            target_user.save()
            messages.success(request, translate_text("User restored.", lang))

    # 2. Manager Actions
    elif action == 'unlock_record':
        emp = Employee.objects.filter( empcode=int(target_user.profile.employee_code)).first()
        if emp:
            emp.status = 'draft'
            emp.save()
            target_user.is_edit_allowed = True
            target_user.save()
            messages.success(request, "Record unlocked.")
    return redirect('manager_dashboard')

# ==================== QPR REPORTING & HOD ====================

@login_required
def qpr_form(request):
    # Prefill QPR form with user's profile values when present (read-only in form)
    profile = getattr(request.user, 'profile', None)
    if not profile or profile.approval_status != 'approved':
        messages.error(request, "Access Denied: Your account must be approved by your HOD before you can submit a QPR.")
        return redirect('dashboard')
    
    # Auto-create current financial year if it doesn't exist
    ensure_current_financial_year()
    
    profile_office_name = profile.office_name if profile and profile.office_name else ''
    profile_office_code = profile.office_code if profile and profile.office_code else ''
    profile_phone = profile.phone if profile and profile.phone else ''
    profile_email = profile.email if profile and profile.email else (request.user.get_email() if hasattr(request.user, 'get_email') else '')

    # Build a list of already used quarters for this user (quarter, year, record_id)
    used = []
    for r in QPRRecord.objects.filter(user=request.user):
        used.append({'quarter': r.quarter, 'year': r.year or '', 'record_id': r.pk})

    

    context = {
        'profile_office_name': profile_office_name,
        'profile_office_code': profile_office_code,
        'profile_phone': profile_phone,
        'profile_email': profile_email,
        'profile_office_name_filled': bool(profile_office_name),
        'profile_office_code_filled': bool(profile_office_code),
        'profile_phone_filled': bool(profile_phone),
        'profile_email_filled': bool(profile_email),
        'used_quarters_json': json.dumps(used),
    }

    # Determine current quarter for preselection using server local date (respects TIME_ZONE)
    today = timezone.localdate()
    month = today.month
    
    # Quarter mapping (Indian FY Apr-Mar): Apr-Jun -> Jun 30, Jul-Sep -> Sep 30, Oct-Dec -> Dec 31, Jan-Mar -> Mar 31
    if 4 <= month <= 6:
        current_quarter = '30 जून / Jun 30'
    elif 7 <= month <= 9:
        current_quarter = '30 सितंबर / Sep 30'
    elif 10 <= month <= 12:
        current_quarter = '31 दिसंबर / Dec 31'
    else:
        current_quarter = '31 मार्च / Mar 31'

    # Compute current financial year string (e.g. "2025-2026") where fiscal year runs Apr-Mar
    fiscal_year_start = today.year - 1 if month < 4 else today.year
    current_financial_year = f"{fiscal_year_start}-{fiscal_year_start + 1}"

    # Build financial_years list from FinancialYear table's earliest recorded start
    # up to the current fiscal year. If none exist, start from current fiscal year.
    from .models import FinancialYear
    fy_qs = FinancialYear.objects.filter(is_active=True)
    min_start = fy_qs.aggregate(Min('start_year'))['start_year__min']
    if min_start is None:
        min_start = fiscal_year_start
    financial_years = []
    for s in range(min_start, fiscal_year_start + 1):
        financial_years.append({'start_year': s, 'end_year': s + 1})

    context.update({
        'current_quarter': current_quarter,
        'current_year': current_financial_year,
        'financial_years': financial_years,
        'user_role': getattr(request.user, 'role', None),
        'active_role': request.session.get('active_role', getattr(request.user, 'role', None)),
        'server_month': today.month,
        'server_year': today.year,
        'profile_language_region': profile.language_region if profile else '',
    })
    # Preload user's existing QPR records so client-side JS can read them without calling the API
    try:
        records_qs = QPRRecord.objects.filter(user=request.user).order_by('-id')
        records = []
        for r in records_qs:
            d = serialize_qpr_record(r)
            # For form preload, only owner can edit; compute approval flags
            edit_approved = False
            if getattr(r, 'is_submitted', False):
                edit_approved = EditRequest.objects.filter(
                    user=request.user,
                    request_type='qpr',
                    qpr_record_id=r.pk,
                    status='approved'
                ).exists()
            d['edit_approved'] = edit_approved
            d['can_edit'] = (not getattr(r, 'is_submitted', False)) or edit_approved
            d['has_pending_edit_request'] = EditRequest.objects.filter(
                user=request.user, request_type='qpr', qpr_record_id=r.pk, status='pending'
            ).exists()
            records.append(d)
    except Exception:
        records = []
    import json as _json
    context['records_json'] = _json.dumps(records, default=str)

    # Precompute availability for the selected/default date so client doesn't need to call the API
    try:
        selected_date = timezone.localdate()
        # For the user-facing QPR form we must restrict selectable dates to today only
        availability = _allowed_frequencies_for_date(request.user, selected_date, allow_future_days=False)
        context['availability_json'] = _json.dumps(availability, default=str)
        context['selected_date'] = selected_date.isoformat()
    except Exception:
        context['availability_json'] = None
        context['selected_date'] = timezone.localdate().isoformat()

    return render(request, 'qpr/qpr_form.html', context)

@login_required
def report_list(request):
    # Allow HODs to view another employee's reports via ?emp_code=<employee_code>
    emp_code = (request.GET.get('emp_code') or '').strip()
    target_user = request.user
    is_hod_view = False
    if emp_code:
        from django.shortcuts import get_object_or_404
        try:
            profile = UserProfile.objects.select_related('user').get(employee_code=emp_code)
        except UserProfile.DoesNotExist:
            messages.error(request, "Employee not found.")
            return redirect('qpr_hod_dashboard' if user_has_role(request.user, 'hod') else 'qpr_user_dashboard')

        target_user = profile.user
        is_hod_view = (getattr(target_user, 'id', None) != getattr(request.user, 'id', None))

        # Authorization: allow admin/superuser, HODs (for their employees), and Managers for same office
        if is_hod_view:
            allowed = False
            # Admins and superusers always allowed
            if user_has_role(request.user, 'admin') or request.user.is_superuser:
                allowed = True

            # HODs: allowed but enforce same hod_name scope
            if not allowed and user_has_role(request.user, 'hod'):
                allowed = True
                requester_hod = (getattr(request.user.profile, 'hod_name', None) or getattr(request.user.profile, 'name', None))
                target_hod = (getattr(target_user, 'profile', None) and (getattr(target_user.profile, 'hod_name', None) or getattr(target_user.profile, 'name', None)))
                if requester_hod and target_hod and str(requester_hod).strip().lower() != str(target_hod).strip().lower():
                    messages.error(request, "Unauthorized to view reports for this employee.")
                    return redirect('qpr_hod_dashboard')

            # Managers: allow if same office_code
            if not allowed and user_has_role(request.user, 'manager'):
                mgr_profile = getattr(request.user, 'userprofile', None) or getattr(request.user, 'profile', None)
                mgr_office = getattr(mgr_profile, 'office_code', None)
                tgt_profile = getattr(target_user, 'profile', None)
                tgt_office = getattr(tgt_profile, 'office_code', None)
                if mgr_office and tgt_office and str(mgr_office).strip() == str(tgt_office).strip():
                    allowed = True

            if not allowed:
                messages.error(request, "Unauthorized to view other user's reports.")
                return redirect('home')

    context = {
        'target_user_id': getattr(target_user, 'id', ''),
        'is_hod_view': is_hod_view,
    }
    # Preload records for client-side rendering without calling API
    try:
        records_qs = QPRRecord.objects.filter(user=target_user).order_by('-id')
        records = []
        for r in records_qs:
            d = serialize_qpr_record(r)
            # If viewing another user's records (HOD/admin), don't allow edit via list
            if getattr(target_user, 'id', None) != getattr(request.user, 'id', None):
                d['can_edit'] = False
                d['edit_approved'] = False
                d['has_pending_edit_request'] = EditRequest.objects.filter(
                    user=target_user, request_type='qpr', qpr_record_id=r.pk, status='pending'
                ).exists()
            else:
                edit_approved = False
                if getattr(r, 'is_submitted', False):
                    edit_approved = EditRequest.objects.filter(
                        user=request.user,
                        request_type='qpr',
                        qpr_record_id=r.pk,
                        status='approved'
                    ).exists()
                d['edit_approved'] = edit_approved
                d['can_edit'] = (not getattr(r, 'is_submitted', False)) or edit_approved
                d['has_pending_edit_request'] = EditRequest.objects.filter(
                    user=request.user, request_type='qpr', qpr_record_id=r.pk, status='pending'
                ).exists()
                # Debug: log EditRequest rows for troublesome example (user_id=4, record_id=1)
                try:
                    if getattr(target_user, 'id', None) == 4 and getattr(r, 'pk', None) == 1:
                        ers = list(EditRequest.objects.filter(user=target_user, qpr_record_id=r.pk).values_list('id', 'status'))
                        print(f"[DEBUG] report_list - EditRequests for user=4 record=1: {ers}")
                except Exception:
                    pass
            records.append(d)
    except Exception:
        records = []
    import json as _json
    context['records_json'] = _json.dumps(records, default=str)
    # Compute period summary for the current quarter/year so client can render daily/weekly/monthly/quarterly lists
    try:
        quarter = get_current_quarter()
        year = get_current_year_label()
        try:
            q_start, q_end = _quarter_label_to_daterange(quarter, year)
        except Exception:
            q_start = None
            q_end = None

        # Default region from profile
        default_region = ''
        try:
            profile = getattr(target_user, 'profile', None)
            if profile and getattr(profile, 'language_region', None):
                default_region = profile.language_region
        except Exception:
            default_region = ''

        # DAILY: include every working day (Mon-Sat) with submitted flag and coverage info
        daily = []
        if q_start and q_end:
            cur = q_start
            while cur <= q_end:
                if cur.weekday() <= 5:  # Mon-Sat
                    # totals from daily records
                    totals = _aggregate_records_with_fallback(target_user, cur, cur, preferred='daily')
                    rec_daily = QPRRecord.objects.filter(user=target_user, is_submitted=True, frequency__iexact='daily', period_start=cur).first()
                    exists_daily = bool(rec_daily)
                    # determine coverage by higher-level records
                    covered_by = None
                    region = getattr(rec_daily, 'region', '') if rec_daily else ''
                    if not exists_daily:
                        # check weekly, monthly, quarterly coverage
                        if QPRRecord.objects.filter(user=target_user, is_submitted=True, frequency__iexact='weekly', period_start__lte=cur, period_end__gte=cur).exists():
                            covered_by = 'weekly'
                        elif QPRRecord.objects.filter(user=target_user, is_submitted=True, frequency__iexact='monthly', period_start__lte=cur, period_end__gte=cur).exists():
                            covered_by = 'monthly'
                        elif QPRRecord.objects.filter(user=target_user, is_submitted=True, frequency__iexact='quarterly', period_start__lte=cur, period_end__gte=cur).exists():
                            covered_by = 'quarterly'
                        # derive region from candidates if not set
                        if not region:
                            cand = QPRRecord.objects.filter(user=target_user, is_submitted=True, period_start__lte=cur, period_end__gte=cur).first()
                            region = getattr(cand, 'region', '') if cand else ''

                    daily.append({
                        'period_start': cur.isoformat(),
                        'period_end': cur.isoformat(),
                        'totals': totals,
                        'has_daily': exists_daily,
                        'covered_by': covered_by,
                        'region': region or default_region or ''
                    })
                cur = cur + timedelta(days=1)

        # WEEKLY: iterate Mon-Sat weeks overlapping quarter
        weekly = []
        if q_start and q_end:
            w_start = q_start - timedelta(days=q_start.weekday())
            while w_start <= q_end:
                w_end = w_start + timedelta(days=5)
                display_start = max(w_start, q_start)
                display_end = min(w_end, q_end)
                totals = _aggregate_records_with_fallback(target_user, display_start, display_end, preferred='weekly')
                daily_count = QPRRecord.objects.filter(user=target_user, is_submitted=True, frequency__iexact='daily', period_start__range=(display_start, display_end)).count()
                expected_days = 0
                for d in range((display_end - display_start).days + 1):
                    dt = display_start + timedelta(days=d)
                    if dt.weekday() <= 5 and q_start <= dt <= q_end:
                        expected_days += 1
                missing_days = max(0, expected_days - daily_count)
                weekly_submitted = QPRRecord.objects.filter(user=target_user, is_submitted=True, frequency__iexact='weekly', period_start__lte=display_start, period_end__gte=display_end).exists()
                weekly_rec = QPRRecord.objects.filter(user=target_user, is_submitted=True, frequency__iexact='weekly', period_start__lte=display_start, period_end__gte=display_end).first()
                region_week = getattr(weekly_rec, 'region', '') if weekly_rec else ''
                if not region_week:
                    cand = QPRRecord.objects.filter(user=target_user, is_submitted=True, frequency__iexact='daily', period_start__range=(display_start, display_end)).first()
                    region_week = getattr(cand, 'region', '') if cand else ''
                weekly.append({
                    'period_start': display_start.isoformat(),
                    'period_end': display_end.isoformat(),
                    'totals': totals,
                    'daily_count': daily_count,
                    'expected_days': expected_days,
                    'missing_days': missing_days,
                    'weekly_submitted': weekly_submitted,
                    'region': region_week or default_region or ''
                })
                w_start = w_start + timedelta(days=7)

        # MONTHLY
        monthly = []
        if q_start and q_end:
            m = q_start
            while m <= q_end:
                month_start = date(m.year, m.month, 1)
                if m.month == 12:
                    month_end = date(m.year, 12, 31)
                else:
                    month_end = date(m.year, m.month + 1, 1) - timedelta(days=1)
                if month_end > q_end:
                    month_end = q_end
                if month_start < q_start:
                    month_start = q_start
                totals = _aggregate_records_with_fallback(target_user, month_start, month_end, preferred='monthly')
                daily_count = QPRRecord.objects.filter(user=target_user, is_submitted=True, frequency__iexact='daily', period_start__range=(month_start, month_end)).count()
                monthly_submitted = QPRRecord.objects.filter(user=target_user, is_submitted=True, frequency__iexact='monthly', period_start__lte=month_start, period_end__gte=month_end).exists()
                monthly_rec = QPRRecord.objects.filter(user=target_user, is_submitted=True, frequency__iexact='monthly', period_start__lte=month_start, period_end__gte=month_end).first()
                region_month = getattr(monthly_rec, 'region', '') if monthly_rec else ''
                if not region_month:
                    cand = QPRRecord.objects.filter(user=target_user, is_submitted=True, frequency__iexact='daily', period_start__range=(month_start, month_end)).first()
                    region_month = getattr(cand, 'region', '') if cand else ''
                monthly.append({
                    'period_start': month_start.isoformat(),
                    'period_end': month_end.isoformat(),
                    'totals': totals,
                    'daily_count': daily_count,
                    'monthly_submitted': monthly_submitted,
                    'region': region_month or default_region or ''
                })
                # next month
                if m.month == 12:
                    m = date(m.year + 1, 1, 1)
                else:
                    m = date(m.year, m.month + 1, 1)

        # QUARTERLY
        if q_start and q_end:
            quarterly_totals = _aggregate_records_with_fallback(target_user, q_start, q_end, preferred='quarterly')
            quarterly_submitted = QPRRecord.objects.filter(user=target_user, is_submitted=True, frequency__iexact='quarterly', period_start=q_start, period_end=q_end).exists()
            quarterly = {'period_start': q_start.isoformat(), 'period_end': q_end.isoformat(), 'totals': quarterly_totals, 'submitted': quarterly_submitted}
        else:
            quarterly = None

        summary = {
            'quarter_label': quarter,
            'year_label': year,
            'quarter_start': q_start.isoformat() if q_start else None,
            'quarter_end': q_end.isoformat() if q_end else None,
            'daily': daily,
            'weekly': weekly,
            'monthly': monthly,
            'quarterly': quarterly
        }
        context['summary_json'] = _json.dumps(summary, default=str)
    except Exception:
        context['summary_json'] = 'null'
    return render(request, 'qpr/report_list.html', context)


@login_required
def finalize_qpr(request):
    """Mark user as finalized for current quarter (via POST request)"""
    if request.method != 'POST':
        return JsonResponse({'error': 'Invalid method'}, status=400)
    
    try:
        from django.http import JsonResponse
        from website.models import QPRFinalization
        
        quarter = get_current_quarter()
        year = get_current_year_label()
        
        # Create or get finalization record
        finalization, created = QPRFinalization.objects.get_or_create(
            user=request.user,
            quarter=quarter,
            year=year
        )
        
        return JsonResponse({
            'success': True,
            'message': f'QPR finalized for {quarter} {year}'
        })
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({'error': str(e)}, status=500)


@login_required
def report_detail(request, record_id):
    # Support HOD view via ?user_id= (existing) and division mode via ?division=1
    user_id = request.GET.get('user_id') or None
    is_hod_view = False
    target_user_id = ''

    # division mode: render aggregated quarterly report for HOD
    division_flag = request.GET.get('division')
    if division_flag == '1':
        # Allow HODs, admins/superusers, and Managers (subject to office match) to access division aggregation
        if not (user_has_role(request.user, 'hod') or user_has_role(request.user, 'admin') or request.user.is_superuser or user_has_role(request.user, 'manager')):
            messages.error(request, 'Unauthorized')
            return redirect('home')

        # If a snapshot record id was passed (report_detail/<id>/?division=1), prefer loading that snapshot
        hod_profile = None
        try:
            # record_id comes from the URL path; try fetching a frozen quarterly snapshot
            if record_id:
                rec = QPRRecord.objects.filter(id=record_id, frequency__iexact='quarterly', is_quarterly_frozen=True).select_related('user').first()
            else:
                rec = None
        except Exception:
            rec = None

        if rec:
            # Authorization: HODs/admins/superusers allowed; Managers allowed only if office matches
            if user_has_role(request.user, 'manager') and not (user_has_role(request.user, 'hod') or user_has_role(request.user, 'admin') or request.user.is_superuser):
                mgr_profile = getattr(request.user, 'userprofile', None) or getattr(request.user, 'profile', None)
                mgr_office = getattr(mgr_profile, 'office_code', None)
                # prefer officeCode stored on the snapshot if present
                rec_office = getattr(rec, 'officeCode', None) or (getattr(rec.user, 'profile', None) and getattr(rec.user.profile, 'office_code', None))
                if not (mgr_office and rec_office and str(mgr_office).strip() == str(rec_office).strip()):
                    messages.error(request, 'Unauthorized')
                    return redirect('home')

            # Serialize snapshot and validate: if the stored snapshot does not match
            # recomputed aggregated totals for the HOD's division, ignore it and
            # render recomputed aggregation instead. This prevents personal
            # frozen QPRs from being shown as division snapshots.
            try:
                snap = serialize_qpr_record(rec)
            except Exception:
                snap = None

            try:
                # Determine HOD identity from the record owner
                owner_profile = getattr(rec.user, 'profile', None)
                owner_hod_name = (owner_profile.hod_name or owner_profile.name) if owner_profile else None
                owner_hod_name = owner_hod_name.strip() if owner_hod_name else None

                # Compute expected aggregated totals for that HOD
                current_quarter = get_current_quarter()
                current_year = get_current_year_label()
                try:
                    q_start, q_end = _quarter_label_to_daterange(current_quarter, current_year)
                except Exception:
                    q_start = None
                    q_end = None

                expected = {k: 0 for k in NUMERIC_KEYS}
                if owner_hod_name and q_start and q_end:
                    users_under = UserProfile.objects.filter(roles__name='user', hod_name__iexact=owner_hod_name).select_related('user')
                    for p in users_under:
                        try:
                            u = getattr(p, 'user', None)
                            if not u:
                                continue
                            ut = _aggregate_records_with_fallback(u, q_start, q_end, preferred='quarterly') or {k: 0 for k in NUMERIC_KEYS}
                            for k in NUMERIC_KEYS:
                                try:
                                    expected[k] += int(ut.get(k, 0) or 0)
                                except Exception:
                                    continue
                        except Exception:
                            continue

                # If snap matches expected aggregated totals for numeric keys,
                # treat it as a valid division snapshot; otherwise fall back to
                # recomputing the aggregation and render that.
                valid_snapshot = False
                if snap is not None:
                    try:
                        all_match = True
                        for k in NUMERIC_KEYS:
                            s_val = int(snap.get(k, 0) or 0)
                            e_val = int(expected.get(k, 0) or 0)
                            if s_val != e_val:
                                all_match = False
                                break
                        if all_match:
                            valid_snapshot = True
                    except Exception:
                        valid_snapshot = False
            except Exception:
                valid_snapshot = False

            if valid_snapshot and snap is not None:
                try:
                    import json as _json
                    initial_qpr_json = _json.dumps(snap, default=str)
                except Exception:
                    initial_qpr_json = '{}'
                return render(request, 'qpr/report_detail.html', {'qpr': snap, 'initial_qpr_json': initial_qpr_json, 'is_division': True})
            # else: fall through to recompute aggregation for this HOD and render below

        # Identify HOD name: prefer profile.hod_name (set by admin), fallback to profile.name
        hod_profile = getattr(request.user, 'profile', None)
        hod_name_val = None
        if hod_profile:
            hod_name_val = (hod_profile.hod_name or hod_profile.name)
        hod_name_val = hod_name_val.strip() if hod_name_val else None
        if not hod_name_val:
            messages.error(request, 'HOD identity not found')
            return redirect('qpr_hod_dashboard')

        # If requester is a Manager (not a HOD/admin), enforce same office_code as the HOD
        if user_has_role(request.user, 'manager') and not user_has_role(request.user, 'hod') and not (user_has_role(request.user, 'admin') or request.user.is_superuser):
            mgr_profile = getattr(request.user, 'userprofile', None) or getattr(request.user, 'profile', None)
            mgr_office = getattr(mgr_profile, 'office_code', None)
            hod_office = getattr(hod_profile, 'office_code', None)
            if not (mgr_office and hod_office and str(mgr_office).strip() == str(hod_office).strip()):
                messages.error(request, 'Unauthorized')
                return redirect('home')

        # Get users under this HOD by matching UserProfile.hod_name == hod_profile.name
        users_under = UserProfile.objects.filter(roles__name='user', hod_name__iexact=hod_name_val).select_related('user')

        current_quarter = get_current_quarter()
        current_year = get_current_year_label()

        # Determine quarter date range
        try:
            q_start, q_end = _quarter_label_to_daterange(current_quarter, current_year)
        except Exception:
            q_start = None
            q_end = None

        # Initialize aggregated totals
        aggregated = {k: 0 for k in NUMERIC_KEYS}
        aggregated['quarter'] = current_quarter
        aggregated['year'] = current_year
        aggregated['frequency'] = 'quarterly'
        aggregated['officeName'] = (hod_name_val or '') + " (Division)"
        aggregated['officeCode'] = ''
        aggregated['region'] = ''
        aggregated['phone'] = ''
        aggregated['email'] = ''
        aggregated['s9_date'] = ''
        aggregated['s10_date'] = ''
        aggregated['s12_1'] = ''
        aggregated['s12_2'] = ''
        aggregated['s12_3'] = ''

        # DEBUG logs
        try:
            print("DIVISION MODE ACTIVE")
            print("HOD:", hod_name_val)
            print("USERS UNDER HOD:", [p.user.id for p in users_under if getattr(p, 'user', None)])
        except Exception:
            pass

        # Aggregate per-user using the same robust fallback aggregation used
        # for individual records (daily->weekly->monthly->quarterly). This
        # avoids skipping records that lack explicit period_start/period_end
        # and ensures division totals reflect stored data.
        try:
            processed_user_ids = set()
            for profile in users_under:
                try:
                    user_obj = getattr(profile, 'user', None)
                    if not user_obj:
                        continue
                    # track processed user ids to avoid double-counting HOD later
                    try:
                        processed_user_ids.add(int(user_obj.id))
                    except Exception:
                        pass
                    print("USER:", user_obj.id)
                    user_totals = _aggregate_records_with_fallback(user_obj, q_start, q_end, preferred='quarterly') or {k: 0 for k in NUMERIC_KEYS}
                    try:
                        print("USER_TOTALS:", user_totals)
                    except Exception:
                        pass
                    any_nonzero = False
                    for k in NUMERIC_KEYS:
                        try:
                            v = int(user_totals.get(k, 0) or 0)
                            aggregated[k] += v
                            if v != 0:
                                any_nonzero = True
                        except Exception:
                            continue
                    if any_nonzero:
                        record_count += 1
                except Exception:
                    continue

            # Include HOD's own totals using the same fallback aggregation
            try:
                hod_user_id = getattr(request.user, 'id', None)
                # If the HOD was already included in users_under aggregation, skip adding again
                if hod_user_id and hod_user_id in processed_user_ids:
                    try:
                        print("HOD already included in user list; skipping additional aggregation for HOD:", hod_user_id)
                    except Exception:
                        pass
                else:
                    hod_totals = _aggregate_records_with_fallback(request.user, q_start, q_end, preferred='quarterly') or {k: 0 for k in NUMERIC_KEYS}
                    try:
                        print("HOD TOTALS:", hod_totals)
                    except Exception:
                        pass
                    for k in NUMERIC_KEYS:
                        try:
                            aggregated[k] += int(hod_totals.get(k, 0) or 0)
                        except Exception:
                            continue
            except Exception:
                pass

            try:
                print("FINAL AGGREGATED:", {k: aggregated.get(k, 0) for k in NUMERIC_KEYS})
            except Exception:
                pass
        except Exception:
            pass

        # Build aggregated_qpr structure expected by template JS
        aggregated_qpr = {
            'quarter': aggregated.get('quarter'),
            'year': aggregated.get('year'),
            'frequency': aggregated.get('frequency'),
            'officeName': aggregated.get('officeName'),
            'officeCode': aggregated.get('officeCode'),
            'region': aggregated.get('region'),
            'phone': aggregated.get('phone'),
            'email': aggregated.get('email'),
        }
        for k in NUMERIC_KEYS:
            aggregated_qpr[k] = aggregated.get(k, 0)
        aggregated_qpr['s9_date'] = aggregated.get('s9_date', '')
        aggregated_qpr['s10_date'] = aggregated.get('s10_date', '')
        aggregated_qpr['s12_1'] = aggregated.get('s12_1', '')
        aggregated_qpr['s12_2'] = aggregated.get('s12_2', '')
        aggregated_qpr['s12_3'] = aggregated.get('s12_3', '')

        # Pass JSON string to template for immediate rendering
        try:
            import json
            # Include cumulative quarterly totals so the frontend `valFor` helper
            # prefers cumulative values for non-daily views (weekly/monthly/quarterly).
            aggregated_qpr['cumulative'] = {'quarterly': {k: aggregated.get(k, 0) for k in NUMERIC_KEYS}}
            initial_qpr_json = json.dumps(aggregated_qpr)
        except Exception:
            initial_qpr_json = '{}'

        return render(request, 'qpr/report_detail.html', {'qpr': aggregated_qpr, 'initial_qpr_json': initial_qpr_json, 'is_division': True})

    # Fallback: existing per-record behavior
    if user_id:
        try:
            uid = int(user_id)
            target = CustomUser.objects.filter(id=uid).first()
        except Exception:
            target = None
        if not target:
            messages.error(request, 'User not found')
            return redirect('qpr_report_list')
        # Authorization: match logic in `report_list` so Managers in the same office can view
        if target.id != request.user.id:
            allowed = False
            # Admins and superusers always allowed
            if user_has_role(request.user, 'admin') or request.user.is_superuser:
                allowed = True

            # HODs: allowed but enforce same hod_name scope
            if not allowed and user_has_role(request.user, 'hod'):
                allowed = True
                requester_hod = (getattr(request.user.profile, 'hod_name', None) or getattr(request.user.profile, 'name', None))
                target_hod = (getattr(target, 'profile', None) and (getattr(target.profile, 'hod_name', None) or getattr(target.profile, 'name', None)))
                if requester_hod and target_hod and str(requester_hod).strip().lower() != str(target_hod).strip().lower():
                    messages.error(request, 'Unauthorized')
                    return redirect('qpr_hod_dashboard')

            # Managers: allow if same office_code
            if not allowed and user_has_role(request.user, 'manager'):
                mgr_profile = getattr(request.user, 'userprofile', None) or getattr(request.user, 'profile', None)
                mgr_office = getattr(mgr_profile, 'office_code', None)
                tgt_profile = getattr(target, 'profile', None)
                tgt_office = getattr(tgt_profile, 'office_code', None)
                if mgr_office and tgt_office and str(mgr_office).strip() == str(tgt_office).strip():
                    allowed = True

            if not allowed:
                messages.error(request, 'Unauthorized')
                return redirect('home')
        is_hod_view = (target.id != request.user.id)
        target_user_id = target.id

    # Preload the record JSON for client-side rendering to avoid API calls
    record_json = '{}'
    try:
        rec = QPRRecord.objects.filter(pk=record_id, user=target if 'target' in locals() and target is not None else request.user).first()
        if rec:
            try:
                import json as _json
                record_data = serialize_qpr_record(rec)
                # If client requested a particular view (weekly/monthly/quarterly),
                # compute aggregated totals for that period so the UI can show
                # per-period aggregates instead of the single-record values.
                view_as = (request.GET.get('view_as') or '').lower()
                try:
                    ps = getattr(rec, 'period_start', None)
                    pe = getattr(rec, 'period_end', None) or ps
                    # ensure ps/pe are date objects
                    if ps:
                        from datetime import timedelta, date
                        # prepare container
                        record_data.setdefault('cumulative', {})
                        # WEEKLY
                        if view_as == 'weekly' or view_as == 'weekly' or True:
                            try:
                                # compute week containing ps (Monday start)
                                wk_start = ps - timedelta(days=(ps.weekday() or 0))
                                wk_end = wk_start + timedelta(days=6)
                                wk_tot = _aggregate_records_with_fallback(rec.user, wk_start, wk_end, preferred='weekly')
                                record_data['cumulative']['weekly'] = wk_tot
                                # Accumulate Section 11 text for weekly view
                                try:
                                    s12_1_txt = _aggregate_section11_text_for_range(rec.user, wk_start, wk_end, 'innovative_work', source_frequency='daily')
                                    s12_2_txt = _aggregate_section11_text_for_range(rec.user, wk_start, wk_end, 'special_events', source_frequency='daily')
                                    s12_3_txt = _aggregate_section11_text_for_range(rec.user, wk_start, wk_end, 'hindi_medium_works', source_frequency='daily')
                                    record_data.setdefault('cumulative_text', {})
                                    record_data['cumulative_text']['weekly'] = {'s12_1': s12_1_txt, 's12_2': s12_2_txt, 's12_3': s12_3_txt}
                                except Exception:
                                    pass
                            except Exception:
                                pass
                        # MONTHLY
                        try:
                            m_start = date(ps.year, ps.month, 1)
                            if ps.month == 12:
                                m_end = date(ps.year, 12, 31)
                            else:
                                m_end = date(ps.year, ps.month + 1, 1) - timedelta(days=1)
                            m_tot = _aggregate_records_with_fallback(rec.user, m_start, m_end, preferred='monthly')
                            record_data['cumulative']['monthly'] = m_tot
                            # Accumulate Section 11 text for monthly view
                            try:
                                s12_1_txt = _aggregate_section11_text_for_range(rec.user, m_start, m_end, 'innovative_work', source_frequency='daily')
                                s12_2_txt = _aggregate_section11_text_for_range(rec.user, m_start, m_end, 'special_events', source_frequency='daily')
                                s12_3_txt = _aggregate_section11_text_for_range(rec.user, m_start, m_end, 'hindi_medium_works', source_frequency='daily')
                                record_data.setdefault('cumulative_text', {})
                                record_data['cumulative_text']['monthly'] = {'s12_1': s12_1_txt, 's12_2': s12_2_txt, 's12_3': s12_3_txt}
                            except Exception:
                                pass
                        except Exception:
                            pass
                        # QUARTERLY (fiscal Apr-Mar quarters)
                        try:
                            mon = ps.month
                            yr = ps.year
                            if 4 <= mon <= 6:
                                q_start = date(yr, 4, 1); q_end = date(yr, 6, 30)
                            elif 7 <= mon <= 9:
                                q_start = date(yr, 7, 1); q_end = date(yr, 9, 30)
                            elif 10 <= mon <= 12:
                                q_start = date(yr, 10, 1); q_end = date(yr, 12, 31)
                            else:
                                # Jan-Mar
                                q_start = date(yr, 1, 1); q_end = date(yr, 3, 31)
                            q_tot = _aggregate_records_with_fallback(rec.user, q_start, q_end, preferred='quarterly')
                            record_data['cumulative']['quarterly'] = q_tot
                            # Accumulate Section 11 text for quarterly view
                            try:
                                s12_1_txt = _aggregate_section11_text_for_range(rec.user, q_start, q_end, 'innovative_work', source_frequency='daily')
                                s12_2_txt = _aggregate_section11_text_for_range(rec.user, q_start, q_end, 'special_events', source_frequency='daily')
                                s12_3_txt = _aggregate_section11_text_for_range(rec.user, q_start, q_end, 'hindi_medium_works', source_frequency='daily')
                                record_data.setdefault('cumulative_text', {})
                                record_data['cumulative_text']['quarterly'] = {'s12_1': s12_1_txt, 's12_2': s12_2_txt, 's12_3': s12_3_txt}
                            except Exception:
                                pass
                        except Exception:
                            pass
                except Exception:
                    pass
                record_json = _json.dumps(record_data, default=str)
                # Compute edit approval flags for the detail preload (same logic as API)
                try:
                    edit_approved = False
                    has_pending_edit_request = False
                    if rec.is_submitted:
                        edit_approved = EditRequest.objects.filter(
                            user=rec.user,
                            request_type='qpr',
                            qpr_record_id=rec.pk,
                            status='approved'
                        ).exists()
                        has_pending_edit_request = EditRequest.objects.filter(
                            user=rec.user,
                            request_type='qpr',
                            qpr_record_id=rec.pk,
                            status='pending'
                        ).exists()
                    record_data['edit_approved'] = edit_approved
                    record_data['can_edit'] = (rec.user == request.user and not rec.is_submitted) or edit_approved
                    record_data['has_pending_edit_request'] = has_pending_edit_request
                    record_json = _json.dumps(record_data, default=str)
                except Exception:
                    pass
            except Exception:
                record_json = '{}'
    except Exception:
        record_json = '{}'

    return render(request, 'qpr/report_detail.html', {'record_id': record_id, 'is_hod_view': is_hod_view, 'target_user_id': target_user_id, 'record_json': record_json})

@login_required
def typing_usage_report_form(request, record_id):
    """Display form for typing usage report"""
    qpr_record = get_object_or_404(QPRRecord, id=record_id, user=request.user)
    
    if request.method == 'POST':
        form = TypingUsageReportForm(request.POST)
        if form.is_valid():
            total_words = form.cleaned_data['total_words']
            hindi_words = form.cleaned_data['hindi_words']
            
            # Create or update the typing usage report
            report, created = TypingUsageReport.objects.update_or_create(
                qpr_record=qpr_record,
                defaults={'total_words': total_words, 'hindi_words': hindi_words}
            )
            
            return redirect('typing_usage_report_view', record_id=record_id)
    else:
        # Pre-fill form if report exists
        try:
            report = TypingUsageReport.objects.get(qpr_record=qpr_record)
            form = TypingUsageReportForm(initial={
                'total_words': report.total_words,
                'hindi_words': report.hindi_words
            })
        except TypingUsageReport.DoesNotExist:
            form = TypingUsageReportForm()
    
    context = {
        'form': form,
        'record_id': record_id,
        'office_name': qpr_record.officeName
    }
    return render(request, 'qpr/typing_usage_report.html', context)

@login_required
def typing_usage_report_view(request, record_id):
    """Display typing usage report with details"""
    qpr_record = get_object_or_404(QPRRecord, id=record_id, user=request.user)
    
    # Get employee profile
    user_profile = request.user.profile
    employee_name = user_profile.name or request.user.username
    designation = user_profile.office_name  # Or you can get it from Employee model
    
    # Try to get designation from Employee model if available
    try:
        employee = Employee.objects.get(empcode=user_profile.employee_code)
        designation = employee.designation or designation
    except Employee.DoesNotExist:
        pass
    
    # Get section7 data (notings data) using safe attribute access
    section7 = getattr(qpr_record, 'section7', None)
    if section7:
        total_notes = getattr(section7, 'total_pages', 0) or 0
        hindi_notes = getattr(section7, 'hindi_pages', 0) or 0
    else:
        total_notes = 0
        hindi_notes = 0
    
    # Get typing usage report data
    try:
        typing_report = TypingUsageReport.objects.get(qpr_record=qpr_record)
        total_words = typing_report.total_words or 0
        hindi_words = typing_report.hindi_words or 0
    except TypingUsageReport.DoesNotExist:
        total_words = 0
        hindi_words = 0
    
    # Calculate percentages
    notes_hindi_percentage = (hindi_notes / total_notes * 100) if total_notes > 0 else 0
    words_hindi_percentage = (hindi_words / total_words * 100) if total_words > 0 else 0
    
    context = {
        'record_id': record_id,
        'office_name': qpr_record.officeName,
        'employee_name': employee_name,
        'designation': designation,
        'total_notes': total_notes,
        'hindi_notes': hindi_notes,
        'notes_hindi_percentage': round(notes_hindi_percentage, 2),
        'total_words': total_words,
        'hindi_words': hindi_words,
        'words_hindi_percentage': round(words_hindi_percentage, 2),
    }
    
    return render(request, 'qpr/typing_usage_report_view.html', context)

@login_required
def hod_detail_list(request):
    """HOD Detail List with profile change request approvals"""
    if not user_has_role(request.user, 'hod'): 
        return redirect('/')
    
    # Determine hod_name from profile (fallback to profile.name)
    hod_profile = getattr(request.user, 'profile', None)
    hod_name = (hod_profile.hod_name or hod_profile.name) if hod_profile else None
    hod_name = hod_name.strip() if hod_name else None

    # Robustly include users who have the 'user' role on either UserProfile or CustomUser
    if hod_name:
        user_role_q = Q(roles__name='user') | Q(user__roles__name='user')
        # Only include approved profiles under this HOD
        users_under_hod = UserProfile.objects.filter(
            user_role_q & Q(hod_name__iexact=hod_name) & Q(approval_status__iexact='approved')
        ).select_related('user').distinct()
    else:
        users_under_hod = UserProfile.objects.filter(user=request.user, approval_status__iexact='approved').select_related('user')
    
    users_data = []
    current_quarter = get_current_quarter()
    current_year = get_current_year_label()
    today = timezone.localdate()
    for user_profile in users_under_hod:
        user = user_profile.user
        qpr_records = user.qpr_records.all()
        office_code = ''
        office_name = ''
        if qpr_records.exists():
            first_qpr = qpr_records.first()
            office_code = first_qpr.officeCode
            office_name = first_qpr.officeName

        # Normalize employee code and try to fill missing profile fields from Employee
        emp_code_val = (user_profile.employee_code or '').strip()
        emp_record = None
        if emp_code_val:
            try:
                emp_int = int(emp_code_val)
                from .models import Employee
                emp_record = Employee.objects.filter(empcode=emp_int).first()
            except Exception:
                from .models import Employee
                emp_record = Employee.objects.filter(empcode=emp_code_val).first()

        # Build display name: prefer profile.name, then full name, then Employee.ename, then username
        display_name = user_profile.name or user.get_full_name() or ''
        if not display_name or display_name.strip().lower() in ['', 'none']:
            if emp_record and emp_record.ename:
                display_name = emp_record.ename
            else:
                display_name = user.username

        # Fill office info: prefer profile, then QPR, then Employee.hname
        office_name_val = user_profile.office_name or office_name or ''
        office_code_val = user_profile.office_code or office_code or ''
        if (not office_name_val or office_name_val.strip() == '') and emp_record:
            office_name_val = getattr(emp_record, 'hname', '') or office_name_val

        has_pending = ManagerRequest.objects.filter(hod=user, request_type='qpr', status='pending').exists()
        current_qpr = qpr_records.filter( quarter=current_quarter, year=current_year ).first()
        # do not include per-user quarterly action fields here (removed from template)
        qpr_complete_flag = current_qpr.is_submitted if current_qpr else False
        # Has the user submitted a daily QPR for today?
        try:
            submitted_today = user.qpr_records.filter(frequency__iexact='daily', period_start=today, is_submitted=True).exists()
        except Exception:
            submitted_today = False
        users_data.append({
            'profile': user_profile, 
            'user': user, 
            'employee_code': user_profile.employee_code,
            'name': display_name, 
            'office_code': office_code_val or 'Not Set', 
            'office_name': office_name_val or 'Not Set',
            'profile_complete': user_profile.profile_updated, 
            'qpr_complete': current_qpr.is_submitted if current_qpr else False,
            'qpr_record_id': current_qpr.id if current_qpr else None,
            'has_pending_edit_request': has_pending,
            'submitted_today': submitted_today,
        })
    
    # NEW: Fetch pending profile change requests for this HOD
    from .models import ProfileChangeRequest
    profile_change_requests = ProfileChangeRequest.objects.filter(
        hod=request.user,
        status__in=['pending', 'approved', 'rejected']
    ).select_related('profile__user').order_by('-requested_at')

    # Calculate finalization counts
    all_users_ids = list(set([up.user.id for up in users_under_hod if getattr(up, 'user', None)]))
    # Include HOD themselves if they have 'user' role and are not already included
    if user_has_role(request.user, 'user') and request.user.id not in all_users_ids:
        all_users_ids.append(request.user.id)
    
    total_users = len(all_users_ids)
    finalized_count = 0
    if total_users > 0:
        from .models import QPRFinalization
        finalized_count = QPRFinalization.objects.filter(
            user_id__in=all_users_ids,
            quarter=current_quarter,
            year=current_year
        ).count()
    
    all_finalized = (total_users > 0 and finalized_count == total_users)

    context = {
        'users_data': users_data, 
        'hod_name': hod_name, 
        'current_quarter': current_quarter, 
        'current_year': current_year,
        'profile_change_requests': profile_change_requests,
        'finalized_count': finalized_count,
        'total_users': total_users,
        'all_finalized': all_finalized,
    }
    
    response = render(request, 'qpr/hod_detail_list.html', context)
    response['Cache-Control'] = 'no-cache, no-store, must-revalidate'
    response['Pragma'] = 'no-cache'
    response['Expires'] = '0'
    return response

@login_required
def toggle_freeze_qpr(request, qpr_record_id):
    """HOD can freeze/unfreeze quarterly QPR reports"""
    if not user_has_role(request.user, 'hod'):
        return redirect('/')
    
    if request.method != 'POST':
        return redirect('qpr_hod_detail_list')
    
    try:
        qpr_record = QPRRecord.objects.get(id=qpr_record_id, frequency='quarterly')
    except QPRRecord.DoesNotExist:
        return JsonResponse({'error': 'Quarterly record not found'}, status=404)
    
    # Check if HOD is authorized (record belongs to user in their department)
    hod_profile = getattr(request.user, 'profile', None)
    hod_name = (hod_profile.hod_name or hod_profile.name) if hod_profile else None
    
    if hod_name:
        user_profile = getattr(qpr_record.user, 'profile', None)
        record_hod_name = (user_profile.hod_name or user_profile.name) if user_profile else None
        if record_hod_name != hod_name.strip():
            return JsonResponse({'error': 'Unauthorized'}, status=403)
    
    # Check if we're at quarter end (can only freeze at quarter end)
    today = date.today()
    quarter_end_dates = get_quarter_end_dates()  # You'll need to create this helper
    
    is_at_quarter_end = today >= quarter_end_dates['current']
    
    # Toggle freeze (can always unfreeze, but can only freeze at quarter-end)
    if qpr_record.is_quarterly_frozen:
        qpr_record.is_quarterly_frozen = False
        message = 'Quarterly report unfrozen'
    else:
        if is_at_quarter_end:
            qpr_record.is_quarterly_frozen = True
            message = 'Quarterly report frozen'
        else:
            days_until_end = (quarter_end_dates['current'] - today).days
            message = f'Can only freeze at quarter end (in {days_until_end} days)'
            return JsonResponse({'error': message}, status=400)
    
    qpr_record.save()
    
    # Redirect back to HOD detail list
    return redirect('qpr_hod_detail_list')

@login_required
def freeze_division_snapshot(request):
    """Compute division aggregation for HOD and save as a quarterly frozen QPRRecord snapshot.

    This endpoint expects POST and will create a new QPRRecord for the HOD user
    with frequency='quarterly', is_quarterly_frozen=True and is_submitted=True.
    Duplicate freezes for same quarter/year are rejected.
    """
    if not user_has_role(request.user, 'hod'):
        messages.error(request, 'Unauthorized access')
        return redirect('qpr_hod_detail_list')

    if request.method != 'POST':
        messages.error(request, 'Invalid request method')
        return redirect('qpr_hod_detail_list')

    hod_profile = getattr(request.user, 'profile', None)
    hod_name = (hod_profile.hod_name or hod_profile.name) if hod_profile else None
    current_quarter = get_current_quarter()
    current_year = get_current_year_label()

    # Prevent duplicate freeze for same HOD + quarter
    existing = QPRRecord.objects.filter(user=request.user, frequency__iexact='quarterly', quarter=current_quarter, year=current_year, is_quarterly_frozen=True)
    if existing.exists():
        messages.error(request, 'You have already frozen for this quarter')
        return redirect('qpr_hod_detail_list')

    # Find users under this HOD
    if hod_name:
        user_role_q = Q(roles__name='user') | Q(user__roles__name='user')
        users_under_hod = UserProfile.objects.filter(user_role_q & Q(hod_name__iexact=hod_name)).select_related('user').distinct()
    else:
        users_under_hod = UserProfile.objects.filter(user=request.user).select_related('user')

    user_ids = list(users_under_hod.values_list('user__id', flat=True))

    # Compute aggregated totals for the quarter by summing per-user cumulative totals
    totals = {k: 0 for k in NUMERIC_KEYS}
    record_count = 0
    try:
        # Determine quarter date range
        try:
            q_start, q_end = _quarter_label_to_daterange(current_quarter, current_year)
        except Exception:
            q_start = None
            q_end = None

        if user_ids and q_start and q_end:
            for uid in user_ids:
                try:
                    u = CustomUser.objects.filter(id=uid).first()
                    if not u:
                        continue
                    # Use fallback aggregation to compute this user's totals for the quarter
                    user_totals = _aggregate_records_with_fallback(u, q_start, q_end, preferred='quarterly') or {k: 0 for k in NUMERIC_KEYS}
                    any_nonzero = False
                    for k in NUMERIC_KEYS:
                        try:
                            v = int(user_totals.get(k, 0) or 0)
                            totals[k] += v
                            if v != 0:
                                any_nonzero = True
                        except Exception:
                            continue
                    if any_nonzero:
                        record_count += 1
                except Exception:
                    continue
    except Exception:
        totals = {k: 0 for k in NUMERIC_KEYS}

    officeName = getattr(hod_profile, 'office_name', '') or ''
    officeCode = getattr(hod_profile, 'office_code', '') or ''

    # DEBUG: log totals before creating snapshot
    try:
        print("[DEBUG] freeze_division_snapshot - TOTALS BEFORE SAVE:", totals)
    except Exception:
        pass

    # Create snapshot record (header only) and persist section data
    qpr_fields = {
        'user': request.user,
        'frequency': 'quarterly',
        'quarter': current_quarter,
        'year': current_year,
        'is_quarterly_frozen': True,
        'is_submitted': True,
        'officeName': officeName,
        'officeCode': officeCode,
    }

    new_rec = QPRRecord.objects.create(**qpr_fields)

    # DEBUG: log snapshot id after create
    try:
        print(f"[DEBUG] freeze_division_snapshot - Created snapshot id={getattr(new_rec, 'id', None)}")
    except Exception:
        pass

    # Save section-level aggregated totals into related Section models
    try:
        _save_section_data(new_rec, totals)
    except Exception:
        try:
            new_rec.delete()
        except Exception:
            pass
        messages.error(request, 'Failed to save aggregated section data')
        return redirect('qpr_hod_detail_list')

    # DEBUG: verify saved snapshot sections
    try:
        # reload record and related sections
        nr = QPRRecord.objects.filter(id=new_rec.id).first()
        from .models import Section1FilesData, Section2MeetingsData
        s1 = Section1FilesData.objects.filter(qpr_record=nr).first()
        s2 = Section2MeetingsData.objects.filter(qpr_record=nr).first()
        print("[DEBUG] freeze_division_snapshot - SNAPSHOT SAVED id=", getattr(nr, 'id', None))
        if s1:
            print("[DEBUG] SNAPSHOT S1:", getattr(s1, 'total_files', None), getattr(s1, 'hindi_files', None))
        if s2:
            print("[DEBUG] SNAPSHOT S2:", getattr(s2, 'meetings_count', None), getattr(s2, 'hindi_minutes', None))
    except Exception:
        pass

    messages.success(request, 'Division quarter frozen successfully. Any further changes in QPR will not be shown in the state aggregation.')
    return redirect('qpr_hod_detail_list')

# ==================== APIs ====================


@login_required
def qpr_records_view(request):
    records = QPRRecord.objects.filter(user=request.user).order_by('-id')
    return render(request, 'qpr_records.html', {
        'records': records
    })


@login_required
def qpr_user_report_list(request):
    """List the current user's quarterly QPRs with Edit/View actions."""
    records = QPRRecord.objects.filter(user=request.user, frequency__iexact='quarterly').order_by('-period_start')
    return render(request, 'qpr/user_report_list.html', {
        'records': records
    })


@login_required
def manager_qpr_list(request):
    """List ManagerQPR records created by the current user."""
    from .models import ManagerQPR
    records = ManagerQPR.objects.filter(user=request.user).order_by('-created_at')
    return render(request, 'qpr/manager_qpr_list.html', {'records': records})


@login_required
def admin_qpr_list(request):
    """List AdminQPR records created by the current user."""
    from .models import AdminQPR
    records = AdminQPR.objects.filter(user=request.user).order_by('-created_at')
    return render(request, 'qpr/admin_qpr_list.html', {'records': records})


@login_required
def qpr_save_record(request):
    if request.method != 'POST':
        return redirect('qpr_records')

    data = request.POST

    # Ignore accidental posts from role-specific QPR forms (manager/admin)
    role_form = (data.get('role_form') or '').strip().lower()
    if role_form:
        try:
            if role_form == 'manager':
                return redirect('manager_qpr_form')
            if role_form == 'admin':
                return redirect('admin_qpr_form')
        except Exception:
            return redirect('qpr_records')

    try:
        year = (data.get('year') or '').strip()

        today = timezone.localdate()
        current_start = today.year if today.month >= 4 else today.year - 1

        if year:
            try:
                selected_start = int(year.split('-')[0])
            except:
                messages.error(request, "Invalid year format")
                return redirect('qpr_records')

            if selected_start > current_start:
                messages.error(request, "Future financial year not allowed")
                return redirect('qpr_records')

        record_id = data.get('id')
        details = json.loads(data.get('details', '{}'))

        # ================= UPDATE =================
        if record_id:
            record = get_object_or_404(QPRRecord, pk=record_id, user=request.user)

            record.officeName = data.get('officeName', '')
            record.officeCode = (data.get('officeCode', '') or '').replace('*', '')
            record.region = data.get('region', '')
            record.quarter = data.get('quarter', '')
            record.year = data.get('year', '')
            record.status = data.get('status', 'Draft')
            record.phone = data.get('phone', '')
            record.email = data.get('email', '')
            record.frequency = data.get('frequency', 'quarterly')
            record.is_submitted = (record.status == 'Submitted')

            allowed_quarters = get_allowed_quarters(record.year)
            if record.quarter and record.quarter not in allowed_quarters:
                messages.error(request, "Invalid quarter selection")
                return redirect('qpr_records')

            ps, pe = record.period_start, record.period_end
            if not ps or not pe:
                try:
                    ps, pe = _quarter_label_to_daterange(record.quarter, record.year)
                except:
                    ps, pe = None, None

            if ps and pe and is_period_overlapping(request.user, ps, pe, exclude_id=record.pk, new_frequency=record.frequency):
                messages.error(request, "This update overlaps with an existing report.")
                return redirect('qpr_records')

            record.period_start = ps
            record.period_end = pe
            record.save()

            # If manager temporarily allowed edits, revoke that flag
            if getattr(request.user, 'is_edit_allowed', False):
                request.user.is_edit_allowed = False
                request.user.save(update_fields=['is_edit_allowed'])

            # Whenever a record is submitted, ensure per-record EditRequest rows are consumed
            if record.is_submitted:
                ManagerRequest.objects.filter(hod=request.user, request_type='qpr', status='approved').delete()

                # Debug: show counts before updating
                try:
                    before = list(EditRequest.objects.filter(user=request.user, qpr_record_id=record.pk).values_list('id','status'))
                    print(f"[DEBUG] before update EditRequests for user={request.user.id} record={record.pk}: {before}")
                except Exception:
                    pass

                # Mark any approved EditRequest(s) for this record as temp use
                EditRequest.objects.filter(
                    user=request.user,
                    request_type='qpr',
                    qpr_record_id=record.pk,
                    status='approved'
                ).update(status='temp use')
                # Reject any pending requests for this same record
                EditRequest.objects.filter(
                    user=request.user,
                    request_type='qpr',
                    qpr_record_id=record.pk,
                    status='pending'
                ).update(status='rejected')

                try:
                    after = list(EditRequest.objects.filter(user=request.user, qpr_record_id=record.pk).values_list('id','status'))
                    print(f"[DEBUG] after update EditRequests for user={request.user.id} record={record.pk}: {after}")
                except Exception:
                    pass

            _save_section_data(record, details)

        # ================= CREATE =================
        else:
            is_submitted = (data.get('status', 'Draft') == 'Submitted')

            frequency = (data.get('frequency') or '').strip()
            selected_date_str = (data.get('selected_date') or '').strip()

            if not frequency:
                messages.error(request, "Frequency is required")
                return redirect('qpr_records')

            if frequency in ['daily', 'weekly', 'monthly'] and not selected_date_str:
                messages.error(request, "Date is required")
                return redirect('qpr_records')

            # Parse and validate selected_date only for non-quarterly frequencies
            selected_date = None
            if frequency != 'quarterly':
                try:
                    selected_date = datetime.strptime(selected_date_str, '%Y-%m-%d').date() if selected_date_str else None
                except:
                    messages.error(request, "Invalid date")
                    return redirect('qpr_records')

                if selected_date:
                    today = timezone.localdate()

                    if selected_date.weekday() == 6:
                        messages.error(request, "Sunday not allowed")
                        return redirect('qpr_records')

                    if selected_date > today:
                        messages.error(request, "Too far in future")
                        return redirect('qpr_records')

                    try:
                        cur_q_start, _ = _get_quarter_range_for_date(today)
                        sel_q_start, _ = _get_quarter_range_for_date(selected_date)
                        if sel_q_start > cur_q_start:
                            messages.error(request, "Future quarter not allowed")
                            return redirect('qpr_records')
                    except:
                        pass

                availability = _allowed_frequencies_for_date(request.user, selected_date)
                if frequency not in availability.get('allowed', []):
                    messages.error(request, f"Allowed: {availability.get('allowed', [])}")
                    return redirect('qpr_records')
            else:
                # Quarterly frequency: no selected_date checks; availability implicitly allowed
                selected_date = None

            ps, pe = compute_period(
                frequency,
                selected_date=selected_date,
                quarter=data.get('quarter'),
                year=data.get('year')
            )

            if ps and pe and is_period_overlapping(request.user, ps, pe, new_frequency=frequency):
                messages.error(request, "Overlapping period")
                return redirect('qpr_records')

            quarter = data.get('quarter', '').strip()
            year = data.get('year', '').strip() or None

            allowed_quarters = get_allowed_quarters(year)
            if quarter and quarter not in allowed_quarters:
                messages.error(request, f"Invalid quarter: {allowed_quarters}")
                return redirect('qpr_records')

            exists = QPRRecord.objects.filter(
                user=request.user,
                frequency__iexact=frequency,
                period_start=ps,
                is_submitted=True
            )

            if quarter:
                exists = exists.filter(quarter=quarter)
                if year:
                    exists = exists.filter(year=year)

            if exists.exists():
                messages.error(request, "Report already exists")
                return redirect('qpr_records')

            record = QPRRecord.objects.create(
                user=request.user,
                officeName=data.get('officeName', ''),
                officeCode=(data.get('officeCode', '') or '').replace('*', ''),
                region=data.get('region', ''),
                quarter=quarter,
                year=year,
                status=data.get('status', 'Draft'),
                frequency=frequency,
                period_start=ps,
                period_end=pe,
                phone=data.get('phone', ''),
                email=data.get('email', ''),
                is_submitted=is_submitted
            )

            _save_section_data(record, details)

        messages.success(request, "Saved successfully")
        # Redirect based on which form was actually submitted (form_type field)
        # This is more reliable than checking active_role for multi-role users
        form_type = (data.get('form_type') or '').strip().lower()
        if form_type == 'manager':
            return redirect('manager_qpr_list')
        elif form_type == 'admin':
            return redirect('admin_qpr_list')
        else:
            # Default to user/HOD report list for user or HOD forms
            return redirect('qpr_report_list')

    except Exception as e:
        messages.error(request, str(e))
        return redirect('qpr_records')


@login_required
def qpr_delete_record(request, id):
    if request.method == "POST":
        QPRRecord.objects.filter(pk=id, user=request.user).delete()
        messages.success(request, "Deleted successfully")
    return redirect('qpr_records')

@login_required
def print_qpr_report(request, record_id):
    """Render a server-side printable version of the QPR record (matches view)."""
    try:
        record = QPRRecord.objects.get(pk=record_id)
    except QPRRecord.DoesNotExist:
        return redirect('qpr_report_list')
    
    if not (
        record.user == request.user or
        user_has_role(request.user, ['manager', 'admin']) or
        request.user.is_superuser
    ):
        return redirect('dashboard')

    data = serialize_qpr_record(record)
    # Render server-side template with the same fields used by report_detail
    return render(request, 'qpr/print_report.html', {'r': data})

@login_required
@csrf_exempt
def request_edit_api(request):
    """API endpoint for requesting QPR edits"""
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            request_type = data.get('request_type')
            record_id = data.get('record_id')
            reason = data.get('reason', '')
            
            if not request_type or not record_id:
                return JsonResponse({'error': 'Missing required fields'}, status=400)
            
            # For QPR requests, create an EditRequest
            if request_type == 'qpr':
                try:
                    record = QPRRecord.objects.get(pk=record_id, user=request.user)
                except QPRRecord.DoesNotExist:
                    return JsonResponse({'error': 'Record not found'}, status=404)
                
                # Check if there's already a pending request
                existing = EditRequest.objects.filter(
                    user=request.user,
                    request_type='qpr',
                    qpr_record_id=record_id,
                    status__in=['pending', 'approved']
                ).first()
                
                if existing:
                    return JsonResponse({'error': 'Request already exists with status: ' + existing.status}, status=400)
                
                # Create the EditRequest
                edit_req = EditRequest.objects.create(
                    user=request.user,
                    request_type='qpr',
                    qpr_record_id=record_id,
                    reason=reason,
                    status='pending'
                )
                
                # Send notification to manager(s)
                from .utils import send_system_email
                manager_office = record.officeCode
                managers = UserProfile.objects.filter(
                    office_code=manager_office,
                    roles__name='manager'
                ).select_related('user')
                
                for profile in managers:
                    try:
                        msg = f"Employee {request.user.get_full_name() or request.user.username} has requested permission to edit their QPR submission.\n\nReason: {reason}"
                        send_system_email(
                            profile.user,
                            request,
                            'manager_alert',
                            extra_context={'body_text': msg, 'subject': 'QPR Edit Request'}
                        )
                    except Exception:
                        # on email failure, continue without raising
                        pass
                
                return JsonResponse({'success': True, 'message': 'Edit request submitted to manager'})
            
            # For profile requests, create the old way with ManagerRequest
            admin_users = User.objects.filter(profile__roles__name='admin')
            for admin_user in admin_users:
                ManagerRequest.objects.create(hod=request.user, user=admin_user, request_type=request_type, reason=f"Edit request: {reason}")
            return JsonResponse({'success': True, 'message': 'Request sent'})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)}, status=500)
    return JsonResponse({'error': 'Invalid method'}, status=400)


@login_required
def employee_form(request):
    if request.session.get('active_role') != 'user': return redirect('dashboard')
    profile = getattr(request.user, 'profile', None)
    from .models import Employee

    # If an Employee record already exists for this user's empcode, let them edit it.
    emp_record = None
    if profile and profile.employee_code:
        emp_record = Employee.objects.filter(empcode=profile.employee_code).first()

    if request.method == 'POST':
        if emp_record:
            form = EmployeeForm(request.POST, instance=emp_record)
        else:
            form = EmployeeForm(request.POST)

        if form.is_valid():
            obj = form.save(commit=False)
            # Ensure empcode is set from user's profile and do not auto-submit; user saves manually
            if profile and profile.employee_code:
                obj.empcode = profile.employee_code
            obj.lastupdate = timezone.now()
            obj.save()
            messages.success(request, 'Employee record saved successfully.')
            return redirect('dashboard')
    else:
        if emp_record:
            form = EmployeeForm(instance=emp_record)
        else:
            initial = {}
            if profile and profile.employee_code:
                initial['empcode'] = profile.employee_code
            initial['ename'] = request.user.first_name or request.user.username
            form = EmployeeForm(initial=initial)

    return render(request, "employeeform.html", {"form": form})


# ==================== EDIT REQUEST WORKFLOW ====================

@login_required
def request_profile_edit(request):
    """User requests to edit their profile"""
    lang = request.session.get('lang', 'en')
    
    if request.method == 'POST':
        try:
            # Check if already pending
            pending_request = EditRequest.objects.filter(
                user=request.user,
                request_type='profile',
                status='pending'
            ).first()
            
            if pending_request:
                messages.warning(request, translate_text("You already have a pending profile edit request. Please wait for approval.", lang))
                return redirect('qpr_user_profile')
            
            reason = request.POST.get('reason', '')
            profile_data = {
                'username': request.POST.get('username', ''),
                'email': request.POST.get('email', ''),
                'name': request.POST.get('name', ''),
                'office_name': request.POST.get('office_name', ''),
                'office_code': request.POST.get('office_code', ''),
            }
            
            EditRequest.objects.create(
                user=request.user,
                request_type='profile',
                requested_data=profile_data,
                reason=reason,
                status='pending'
            )
            messages.success(request, translate_text("Profile edit request submitted to admin for approval. You will not be able to submit again until approved or rejected.", lang))
            
            # Send notification to admins
            admins = CustomUser.objects.filter(roles__name='admin', is_active=True)
            for admin in admins:
                msg = f"User {request.user.username} ({request.user.profile.employee_code}) has requested to edit their profile."
                send_system_email(admin, request, 'manager_alert', extra_context={'body_text': msg})
            
            return redirect('qpr_user_profile')
        except Exception as e:
            messages.error(request, translate_text("Error submitting request.", lang))
            return redirect('qpr_user_profile')
    
    context = {'profile': request.user.profile, 'current_lang': lang}
    return render(request, 'qpr/request_profile_edit.html', context)


@login_required
def request_qpr_edit(request, record_id):
    """User requests to edit their submitted QPR"""
    lang = request.session.get('lang', 'en')
    
    try:
        qpr_record = QPRRecord.objects.get(id=record_id, user=request.user)
    except QPRRecord.DoesNotExist:
        messages.error(request, translate_text("QPR record not found.", lang))
        return redirect('qpr_report_list')
    
    # Check if quarterly is frozen
    if qpr_record.frequency == 'quarterly' and qpr_record.is_quarterly_frozen:
        messages.error(request, translate_text("This quarterly report is frozen and cannot be edited.", lang))
        return redirect('qpr_report_detail', record_id=record_id)
    
    if request.method == 'POST':
        try:
            reason = request.POST.get('reason', '')
            
            # Check if already pending
            pending_request = EditRequest.objects.filter(
                user=request.user,
                request_type='qpr',
                qpr_record_id=record_id,
                status='pending'
            ).exists()
            
            if pending_request:
                messages.warning(request, translate_text("You already have a pending QPR edit request for this record.", lang))
            else:
                qpr_data = {
                    'qpr_id': record_id,
                    'office_name': qpr_record.officeName,
                    'quarter': qpr_record.quarter,
                    'year': qpr_record.year,
                }
                
                EditRequest.objects.create(
                    user=request.user,
                    request_type='qpr',
                    qpr_record_id=record_id,
                    requested_data=qpr_data,
                    reason=reason,
                    status='pending'
                )
                messages.success(request, translate_text("QPR edit request submitted to admin for approval.", lang))
                
                # Send notification to admins
                admins = CustomUser.objects.filter(roles__name='admin', is_active=True)
                for admin in admins:
                    msg = f"User {request.user.username} ({request.user.profile.employee_code}) has requested to edit QPR for {qpr_record.quarter}."
                    send_system_email(admin, request, 'manager_alert', extra_context={'body_text': msg})
            
            return redirect('qpr_report_detail', record_id=record_id)
        except Exception as e:
            messages.error(request, translate_text("Error submitting request.", lang))
            return redirect('qpr_report_detail', record_id=record_id)
    
    context = {'qpr_record': qpr_record, 'current_lang': lang}
    return render(request, 'qpr/request_qpr_edit.html', context)


@login_required
def admin_edit_requests(request):
    """Admin view all pending edit requests"""
    if not user_has_role(request.user, ['admin']):
        return redirect('/')
    lang = request.session.get('lang', 'en')
    admin_state = request.user.profile.office_state
    
    status_filter = request.GET.get('status', 'pending')
    request_type_filter = request.GET.get('type', '')
    
    edit_requests = EditRequest.objects.filter(
        user__profile__office_state=admin_state
    ).select_related('user', 'approved_by')
    
    if status_filter:
        edit_requests = edit_requests.filter(status=status_filter)
    
    if request_type_filter:
        edit_requests = edit_requests.filter(request_type=request_type_filter)
    
    context = {
        'edit_requests': edit_requests,
        'status_filter': status_filter,
        'request_type_filter': request_type_filter,
        'statuses': [('pending', 'Pending'), ('approved', 'Approved'), ('rejected', 'Rejected'), ('temp use', 'Temp Use')],
        'types': [('profile', 'Profile'), ('qpr', 'QPR')],
        'current_lang': lang,
    }
    return render(request, 'qpr/admin_edit_requests.html', context)


@login_required
def approve_edit_request(request, request_id):
    """Manager approves a QPR edit request"""
    if not user_has_role(request.user, 'manager'):
        return redirect('/')
    lang = request.session.get('lang', 'en')
    
    try:
        edit_request = EditRequest.objects.get(id=request_id)
    except EditRequest.DoesNotExist:
        messages.error(request, translate_text("Request not found.", lang))
        return redirect('manager_dashboard')
    
    if request.method == 'POST':
        try:
            admin_notes = request.POST.get('admin_notes', '')
            
            edit_request.status = 'approved'
            edit_request.approved_by = request.user
            edit_request.approved_at = now()
            edit_request.admin_notes = admin_notes
            edit_request.save()
            # Approved EditRequest — user is notified; do not set global is_edit_allowed here.
            
            # Send notification to user
            msg = f"Your {edit_request.get_request_type_display().lower()} edit request has been approved."
            if admin_notes:
                msg += f"\n\nAdmin Notes: {admin_notes}"
            send_system_email(
                edit_request.user,
                request,
                'manager_alert',
                extra_context={'body_text': msg, 'subject': 'Edit Request Approved'}
            )
            
            messages.success(request, translate_text("Edit request approved.", lang))
            return redirect('manager_dashboard')
        except Exception as e:
            messages.error(request, translate_text(f"Error approving request: {str(e)}", lang))
            return redirect('manager_dashboard')
    
    context = {'edit_request': edit_request, 'current_lang': lang}
    return render(request, 'qpr/approve_edit_request.html', context)


@login_required
def reject_edit_request(request, request_id):
    """Manager rejects a QPR edit request"""
    if not user_has_role(request.user, 'manager'):
        return redirect('/')
    lang = request.session.get('lang', 'en')
    
    try:
        edit_request = EditRequest.objects.get(id=request_id)
    except EditRequest.DoesNotExist:
        messages.error(request, translate_text("Request not found.", lang))
        return redirect('manager_dashboard')
    
    if request.method == 'POST':
        try:
            admin_notes = request.POST.get('admin_notes', '')
            
            if not admin_notes:
                messages.error(request, translate_text("Please provide a reason for rejection.", lang))
                return render(request, 'qpr/reject_edit_request.html', {'edit_request': edit_request})
            
            edit_request.status = 'rejected'
            edit_request.approved_by = request.user
            edit_request.approved_at = now()
            edit_request.admin_notes = admin_notes
            edit_request.save()
            
            # Send notification to user
            msg = f"Your {edit_request.get_request_type_display().lower()} edit request has been rejected.\n\nReason: {admin_notes}"
            send_system_email(
                edit_request.user,
                request,
                'manager_alert',
                extra_context={'body_text': msg, 'subject': 'Edit Request Rejected'}
            )
            
            messages.success(request, translate_text("Edit request rejected.", lang))
            return redirect('manager_dashboard')
        except Exception as e:
            messages.error(request, translate_text(f"Error rejecting request: {str(e)}", lang))
            return redirect('manager_dashboard')
    
    context = {'edit_request': edit_request, 'current_lang': lang}
    return render(request, 'qpr/reject_edit_request.html', context)


def typing_data_report(request):
    lang = request.session.get('lang', 'en')
    if user_has_role(request.user, 'admin'):
        admin_state = request.user.profile.office_state
        typing_reports = TypingUsageReport.objects.filter(
            qpr_record__user__profile__office_state=admin_state
        ).select_related('qpr_record__user__profile', 'qpr_record__section7')
    else:
        typing_reports = TypingUsageReport.objects.select_related(
            'qpr_record__user__profile', 'qpr_record__section7'
        ).all()
    typing_reports = TypingUsageReport.objects.select_related(
        'qpr_record__user__profile',
        'qpr_record__section7'
    ).all()
    data = []
    for report in typing_reports:
        qpr_record = report.qpr_record
        user_profile = qpr_record.user.profile if qpr_record.user else None
        employee_name = (user_profile.name if user_profile else None) or (qpr_record.user.username if qpr_record.user else 'Unknown')
        designation = 'N/A'
        office_code = (user_profile.office_code if user_profile else None) or 'N/A'

        try:
            if user_profile and user_profile.employee_code:
                employee = Employee.objects.get(empcode=user_profile.employee_code)
                designation = employee.designation or 'N/A'
                office_code = (user_profile.office_code if user_profile else None) or 'N/A'
        except Employee.DoesNotExist:
            pass
        
        # Get section7 data using safe attribute access
        section7 = getattr(qpr_record, 'section7', None)
        if section7:
            total_notes = getattr(section7, 'total_pages', 0) or 0
            hindi_notes = getattr(section7, 'hindi_pages', 0) or 0
        else:
            total_notes = 0
            hindi_notes = 0
        
        notes_hindi_percentage = (hindi_notes / total_notes * 100) if total_notes > 0 else 0
        words_hindi_percentage = ((report.hindi_words or 0) / (report.total_words or 1) * 100) if (report.total_words and report.total_words > 0) else 0
        
        data.append({
            'serial_no': len(data) + 1,
            'employee_name': employee_name,
            'designation': designation,
            'office_code': office_code,
            'total_notes': total_notes,
            'hindi_notes': hindi_notes,
            'notes_hindi_percentage': round(notes_hindi_percentage, 2),
            'total_words': report.total_words or 0,
            'hindi_words': report.hindi_words or 0,
            'words_hindi_percentage': round(words_hindi_percentage, 2),
            'year': qpr_record.year,
            'quarter': qpr_record.quarter,
        })
    
    context = {
        'typing_data': data,
        'years': sorted(set(r['year'] for r in data if r['year']), reverse=True),
        'quarters': sorted(set(r['quarter'] for r in data if r['quarter'])),
        'current_lang': lang,
    }
    return render(request, 'qpr/typing_data_report.html', context)


# ==================== USER HOD SELECTION ====================

@login_required
def api_user_change_hod(request):
    """API endpoint for users to change their assigned HOD"""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'POST method required'}, status=405)

    try:
        # Support both JSON (AJAX) and standard form POST submissions.
        is_ajax = False
        if request.content_type and 'application/json' in request.content_type:
            data = json.loads(request.body)
            new_hod_name = data.get('hod_name', '').strip()
            is_ajax = True
        else:
            new_hod_name = request.POST.get('hod_name', '').strip()

        if not new_hod_name:
            if not is_ajax:
                messages.error(request, 'HOD name is required')
                return redirect('dashboard')
            return JsonResponse({'success': False, 'error': 'HOD name is required'}, status=400)

        # Check if user is HOD or Manager - they shouldn't be able to change HOD
        if user_has_role(request.user, ['hod', 'manager', 'admin']):
            if not is_ajax:
                messages.error(request, 'Only users can change their HOD')
                return redirect('dashboard')
            return JsonResponse({'success': False, 'error': 'Only users can change their HOD'}, status=403)

        # Get user's profile
        try:
            profile = UserProfile.objects.get(user=request.user)
        except UserProfile.DoesNotExist:
            if not is_ajax:
                messages.error(request, 'User profile not found')
                return redirect('dashboard')
            return JsonResponse({'success': False, 'error': 'User profile not found'}, status=404)

        # Verify the selected HOD exists (check both profile.roles and the user's roles)
        hod_exists = UserProfile.objects.filter(
            Q(roles__name='hod') | Q(user__roles__name='hod'),
            hod_name__iexact=new_hod_name
        ).exists()
        if not hod_exists:
            if not is_ajax:
                messages.error(request, 'Selected HOD does not exist')
                return redirect('dashboard')
            return JsonResponse({'success': False, 'error': 'Selected HOD does not exist'}, status=400)

        # Update the HOD
        old_hod = profile.hod_name
        profile.hod_name = new_hod_name
        profile.save()

        # If this was a standard form submit, use messages and redirect back to dashboard
        if not is_ajax:
            messages.success(request, f'HOD changed successfully from {old_hod or "None"} to {new_hod_name}')
            return redirect('dashboard')

        return JsonResponse({
            'success': True,
            'message': f'HOD changed successfully from {old_hod or "None"} to {new_hod_name}',
            'new_hod': new_hod_name
        })

    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'error': 'Invalid JSON'}, status=400)
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


# ==================== HOD MANAGEMENT (ADMIN ONLY) ====================

@csrf_exempt
@login_required
def api_update_hod(request):
    """API endpoint to update HOD name and employee code (Admin only)"""
    if not user_has_role(request.user, ['admin']):
        return JsonResponse({'success': False, 'error': 'Access denied. Admin only.'}, status=403)
    
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            old_hod_name = data.get('old_hod_name')  # Current HOD name
            new_hod_name = data.get('new_hod_name')  # New HOD name
            old_employee_code = data.get('old_employee_code')  # Current HOD employee code
            new_employee_code = data.get('new_employee_code')  # New HOD employee code
            
            if not old_hod_name or not new_hod_name or not old_employee_code or not new_employee_code:
                return JsonResponse({
                    'success': False,
                    'error': 'All fields required: old_hod_name, new_hod_name, old_employee_code, new_employee_code'
                }, status=400)
            
            # Find the HOD user profile
            try:
                hod_profile = UserProfile.objects.get(employee_code=old_employee_code, roles__name='hod')
            except UserProfile.DoesNotExist:
                return JsonResponse({
                    'success': False,
                    'error': f'HOD with employee code {old_employee_code} not found'
                }, status=404)
            
            # Check if new_employee_code is already taken
            if new_employee_code != old_employee_code:
                if UserProfile.objects.filter(employee_code=new_employee_code).exists():
                    return JsonResponse({
                        'success': False,
                        'error': f'Employee code {new_employee_code} is already in use'
                    }, status=400)
            
            # Update HOD profile
            hod_profile.name = new_hod_name
            hod_profile.hod_name = new_hod_name
            hod_profile.employee_code = new_employee_code
            hod_profile.user.username = new_employee_code  # Update Django User username
            hod_profile.user.save()
            hod_profile.save()
            
            # Update all users under this HOD (update their hod_name reference)
            UserProfile.objects.filter(
                role='user',
                hod_name__iexact=old_hod_name
            ).update(hod_name=new_hod_name)
            
            return JsonResponse({
                'success': True,
                'message': f'HOD updated successfully! {old_hod_name} → {new_hod_name}, {old_employee_code} → {new_employee_code}',
                'new_hod_name': new_hod_name
            })
        
        except Exception as e:
            return JsonResponse({
                'success': False,
                'error': f'Server error: {str(e)}'
            }, status=500)
    
    return JsonResponse({'error': 'Invalid method'}, status=400)
def send_reminder_email(request, user_id):
    user_profile = getattr(request.user, 'profile', None)
    if not user_profile or not user_profile.roles.filter(name='hod').exists(): 
        return redirect('/')
        
    if request.method == 'POST':
        target_user = get_object_or_404(CustomUser, id=user_id)
        lang = request.session.get('lang', 'en')
        target_profile = getattr(target_user, 'profile', None)
        if target_profile and target_profile.hod_name == user_profile.hod_name:
            send_system_email(target_user, request, 'reminder')
            messages.success(request, translate_text(f"Reminder email sent successfully to {target_user.username}.", lang))
        else:
            messages.error(request, translate_text("Unauthorized action.", lang))
            
    return redirect('qpr_hod_detail_list')
@login_required
def export_employee_pdf(request):
    if request.session.get('active_role') != 'user':
        return redirect('dashboard')

    try:
        profile = getattr(request.user, 'profile', None)

        if profile and profile.employee_code:
            user_empcode = int(profile.employee_code)
        else:
            user_empcode = int(request.user.username)

    except (ValueError, TypeError):
        messages.error(request, "Invalid employee code.")
        return redirect('dashboard')

    employees = Employee.objects.filter(empcode=user_empcode, status='submitted')
    lang = request.GET.get('lang', 'en')

    # Translation dictionary (same as your JS dictionary)
    hindi_dict = {
        "Passed": "उत्तीर्ण",
        "Did not Appear": "उपस्थित नहीं हुए",
        "Failed": "अनुत्तीर्ण",
        "Good": "अच्छा",
        "Average": "औसत",
        "Basic": "बुनियादी",
        "Hindi": "हिंदी",
        "English": "अंग्रेजी",
        "Both": "दोनों",
        "Gazetted": "राजपत्रित",
        "Non-Gazetted": "अराजपत्रित",
        "Scientist-F": "वैज्ञानिक-एफ",
        "Scientist-G": "वैज्ञानिक-जी",
        "Scientist-E": "वैज्ञानिक-ई",
        "Scientist-D": "वैज्ञानिक-डी",
        "Scientist-C": "वैज्ञानिक-सी",
        "Scientist-B": "वैज्ञानिक-बी",
        "Section Officer": "अनुभाग अधिकारी",
        "Senior Secretariate Assistant": "वरिष्ठ सचिवालय सहायक",
        "Scientific/Technical Assistant-A": "वैज्ञानिक/तकनीकी सहायक-ए",
        "Scientific/Technical Assistant-B": "वैज्ञानिक/तकनीकी सहायक-बी",
        "Scientific Officer/Engineer-SB": "वैज्ञानिक अधिकारी/इंजीनियर-एसबी",
        "Pending": "लंबित",
    }

    def t(value):
        """Translate value if lang is Hindi"""
        if not value or value == '-':
            return '-'
        if lang == 'hi':
            return hindi_dict.get(str(value), str(value))
        return str(value)

    buffer = io.BytesIO()

    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib import colors
    from reportlab.lib.units import mm

    page = landscape(A4)
    margin = 15 * mm

    doc = SimpleDocTemplate(
        buffer, pagesize=page,
        rightMargin=margin, leftMargin=margin,
        topMargin=margin, bottomMargin=margin
    )

    header_style = ParagraphStyle('Header', fontName='HindiFont', fontSize=8,
        leading=11, textColor=colors.white, alignment=1)
    cell_style = ParagraphStyle('Cell', fontName='HindiFont', fontSize=8,
        leading=11, alignment=1)
    title_style = ParagraphStyle('Title', fontName='HindiFont', fontSize=14,
        leading=18, spaceAfter=6)
    subtitle_style = ParagraphStyle('Subtitle', fontName='HindiFont', fontSize=9,
        leading=12, spaceAfter=10, textColor=colors.HexColor('#555555'))

    col_widths = [18*mm, 28*mm, 28*mm, 38*mm, 18*mm, 22*mm, 22*mm, 20*mm, 20*mm, 18*mm, 20*mm, 17*mm]

    # Headers — translated if Hindi
    if lang == 'hi':
        header_texts = [
            "एम्पकोड", "अंग्रेजी में नाम", "हिंदी में नाम", "पद का नाम",
            "टाइपिंग", "हिंदी<br/>प्रवीणता", "राजपत्र", "प्रबोध",
            "प्रवीण", "प्रज्ञा", "पारंगत", "सेवानिवृत्ति<br/>तिथि"
        ]
        title_text = "सबमिट किए गए कर्मचारी रिकॉर्ड"
    else:
        header_texts = [
            "Emp<br/>Code", "Name in<br/>English", "Name in<br/>Hindi", "Designation",
            "Typing", "Hindi<br/>Proficiency", "Gazet", "Prabodh",
            "Praveen", "Pragya", "Parangat", "Superann.<br/>Date"
        ]
        title_text = "Submitted Employee Records"

    headers = [Paragraph(h, header_style) for h in header_texts]
    table_data = [headers]

    for emp in employees:
        raw_date = emp.get_super_annuation_date()
        raw_date.strftime('%Y-%m-%d')

        row = [
            Paragraph(str(emp.empcode or '-'), cell_style),
            Paragraph(str(emp.ename or '-'), cell_style),
            Paragraph(str(emp.hname or '-'), cell_style),
            Paragraph(t(emp.designation), cell_style),
            Paragraph(t(emp.typing), cell_style),
            Paragraph(t(emp.hindiproficiency), cell_style),
            Paragraph(t(emp.gazet), cell_style),
            Paragraph(t(emp.prabodh), cell_style),
            Paragraph(t(emp.praveen), cell_style),
            Paragraph(t(emp.pragya), cell_style),
            Paragraph(t(emp.parangat), cell_style),
            Paragraph(masked_date, cell_style),
        ]
        table_data.append(row)

    table = Table(table_data, colWidths=col_widths, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1a6496')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, -1), 'HindiFont'),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#cccccc')),
        ('BOX', (0, 0), (-1, -1), 1, colors.HexColor('#1a6496')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f0f6fb')]),
        ('TOPPADDING', (0, 0), (-1, -1), 5),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
        ('LEFTPADDING', (0, 0), (-1, -1), 3),
        ('RIGHTPADDING', (0, 0), (-1, -1), 3),
    ]))

    from datetime import date
    elements = [
        Paragraph(title_text, title_style),
        Paragraph(f"Generated on: {date.today().strftime('%d %B %Y')}", subtitle_style),
        Spacer(1, 4*mm),
        table,
    ]

    doc.build(elements)
    buffer.seek(0)
    return FileResponse(buffer, as_attachment=True, filename='employee_records.pdf')


@login_required
def manager_report(request):
    if not (user_has_role(request.user, ['manager', 'admin']) or request.user.is_superuser):
        return redirect('/')

    manager_profile = getattr(request.user, 'userprofile', None) or getattr(request.user, 'profile', None)
    office_code = getattr(manager_profile, 'office_code', None)

    manager_data = []
    hod_profiles = UserProfile.objects.none()
    if office_code:
        hod_profiles = UserProfile.objects.filter(
            office_code=office_code,
            roles__name__iexact='hod',
            approval_status='approved'
        ).select_related('user').distinct()

        for hod_profile in hod_profiles:
            try:
                # Use robust display name for HOD: prefer profile.name, then user full name, then username
                user_obj = getattr(hod_profile, 'user', None)
                hod_display_name = (hod_profile.name or '').strip()
                # If hod_profile.name looks like an empcode (numeric) or is missing,
                # prefer profile-linked employee name or user's full name
                def _looks_like_empcode(s):
                    try:
                        return str(s).strip().isdigit()
                    except Exception:
                        return False

                if (not hod_display_name or _looks_like_empcode(hod_display_name)) and user_obj:
                    # try Employee record linked to profile
                    emp_rec = None
                    try:
                        emp_rec = getattr(hod_profile, 'employee', None)
                    except Exception:
                        emp_rec = None
                    if emp_rec and getattr(emp_rec, 'ename', None):
                        hod_display_name = emp_rec.ename.strip()
                    else:
                        try:
                            hod_display_name = (user_obj.get_full_name() or user_obj.username or '').strip()
                        except Exception:
                            hod_display_name = (getattr(user_obj, 'username', '') or '').strip()

                # Find employees that report to this HOD. Try multiple matching keys because
                # different records may store HOD identifier as name, empcode, or username.
                employees_qs = UserProfile.objects.filter(
                    approval_status='approved',
                ).select_related('user')

                match_keys = []
                if hod_profile.name:
                    match_keys.append(hod_profile.name)
                if getattr(hod_profile, 'employee_code', None):
                    match_keys.append(str(hod_profile.employee_code))
                if user_obj and getattr(user_obj, 'username', None):
                    match_keys.append(user_obj.username)

                # Build OR query across possible hod_name variants
                if match_keys:
                    q = Q()
                    for key in set(match_keys):
                        if key and str(key).strip():
                            q |= Q(hod_name__iexact=str(key).strip())
                    if q:
                        employees_qs = employees_qs.filter(q)

                # Build employee dicts. We'll ensure the HOD appears as the first entry
                emp_dicts = []
                for p in employees_qs:
                    try:
                        u = getattr(p, 'user', None)
                        uid = getattr(u, 'id', None)
                        if uid is None:
                            continue
                        name = (p.name or '') or (getattr(u, 'get_full_name', lambda: None)() or getattr(u, 'username', ''))
                        # try to get Hindi name from linked Employee record if available
                        try:
                            emp_obj = getattr(p, 'employee', None)
                            hname_val = getattr(emp_obj, 'hname', '') if emp_obj is not None else ''
                        except Exception:
                            hname_val = ''
                        empcode = getattr(p, 'employee_code', '') or getattr(u, 'username', '') or ''
                        empcode = str(empcode)
                        ipnum = getattr(p, 'ip_number', '') or ''
                        emp_dicts.append({'id': uid, 'name': name, 'hname': hname_val or name, 'empcode': empcode, 'ip': ipnum})
                    except Exception:
                        continue

                # Ensure HOD is present at the top of the employees list
                try:
                    hod_user_id = getattr(user_obj, 'id', None)
                    hod_empcode = getattr(hod_profile, 'employee_code', '') or (getattr(user_obj, 'username', '') or '')
                    hod_ip = getattr(hod_profile, 'ip_number', '') or ''
                    # also store Hindi name for HOD (if linked employee exists)
                    try:
                        hod_emp_obj = getattr(hod_profile, 'employee', None)
                        hod_hname = getattr(hod_emp_obj, 'hname', '') if hod_emp_obj is not None else ''
                    except Exception:
                        hod_hname = ''
                    hod_entry = {'id': hod_user_id, 'name': hod_display_name, 'hname': hod_hname or hod_display_name, 'empcode': str(hod_empcode), 'ip': hod_ip}
                    # remove existing occurrence of HOD if present
                    emp_dicts = [e for e in emp_dicts if e.get('id') != hod_user_id]
                    emp_dicts.insert(0, hod_entry)
                except Exception:
                    pass

                manager_data.append({
                    'hod_name': hod_display_name,
                    'hod_hname': (getattr(hod_profile, 'employee', None) and getattr(hod_profile.employee, 'hname', '')) or hod_display_name,
                    'hod_id': getattr(user_obj, 'id', None),
                    'employees': emp_dicts,
                    'division_frozen': False,
                    'division_qpr_id': None,
                })
            except Exception:
                continue

    # Detect frozen division QPR per HOD and compute state totals (current quarter/year)
    try:
        # Build mapping from hod user id -> index in manager_data for reliable lookup
        hod_user_map = {}
        hod_user_ids = []
        for idx, entry in enumerate(manager_data):
            hid = entry.get('hod_id')
            if hid is not None:
                hod_user_map[hid] = idx
                hod_user_ids.append(hid)

        current_quarter = get_current_quarter()
        current_year = get_current_year_label()
        total_hods = len(hod_user_ids)

        state_totals = {k: 0 for k in NUMERIC_KEYS}
        frozen_count = 0
        # --- NEW REPLACEMENT LOGIC ---
        # --- BULLETPROOF REPLACEMENT LOGIC ---
       # --- DEBUG REPLACEMENT LOGIC ---
        frozen_count = 0
        for hid in hod_user_ids:
            hod_rec = QPRRecord.objects.filter(
                user_id=hid, 
                is_quarterly_frozen=True, 
                quarter=current_quarter, 
                year=current_year
            ).first()

            if hod_rec:
                frozen_count += 1
                
                # Capture the office code
                hod_office = hod_rec.user.profile.office_code
                
                # Fetch team members explicitly
                team_ids = list(UserProfile.objects.filter(office_code=hod_office).values_list('user_id', flat=True))
                
                # --- DIAGNOSTIC PRINT 1 ---
                print(f"DEBUG: HOD {hid} found. Office: {hod_office}. Team IDs: {team_ids}")

                team_records = QPRRecord.objects.filter(
                    user_id__in=team_ids,
                    quarter=current_quarter,
                    year=current_year
                )

                # --- DIAGNOSTIC PRINT 2 ---
                print(f"DEBUG: Found {team_records.count()} records for this team.")

                for rec in team_records:
                    d = serialize_qpr_record(rec)
                    for k in NUMERIC_KEYS:
                        v = d.get(k)
                        if v:
                            val = int(v)
                            state_totals[k] += val
                            # --- DIAGNOSTIC PRINT 3 ---
                            if k == 's2_meetings': # Check one specific key
                                print(f"DEBUG: Adding {val} from User {rec.user_id}. Current Subtotal: {state_totals[k]}")

                if hid in hod_user_map:
                    idx = hod_user_map[hid]
                    manager_data[idx]['division_frozen'] = True
                    manager_data[idx]['division_qpr_id'] = hod_rec.id
        
        # Also include ManagerQPR/AdminQPR records to compute state totals
        try:
            from .models import ManagerQPR, AdminQPR
            mgr_qprs = ManagerQPR.objects.filter(quarter=current_quarter, financial_year=current_year, user__profile__office_code=office_code)
            for mq in mgr_qprs:
                try:
                    mvals = _serialize_managerqpr(mq)
                    for k in NUMERIC_KEYS:
                        try:
                            v = int(mvals.get(k, 0) or 0)
                            state_totals[k] += v
                        except Exception:
                            continue
                except Exception:
                    continue

            adm_qprs = AdminQPR.objects.filter(quarter=current_quarter, financial_year=current_year, user__profile__office_code=office_code)
            for aq in adm_qprs:
                try:
                    avals = _serialize_adminqpr(aq)
                    for k in NUMERIC_KEYS:
                        try:
                            v = int(avals.get(k, 0) or 0)
                            state_totals[k] += v
                        except Exception:
                            continue
                except Exception:
                    continue
        except Exception:
            pass


    except Exception:
        state_totals = {k: 0 for k in NUMERIC_KEYS}
        frozen_count = 0
        total_hods = hod_profiles.count() if 'hod_profiles' in locals() else 0

    # Build single aggregated QPR-like dict for template and items list
    state_qpr = None
    state_qpr_items = []
    try:
        state_qpr = {'quarter': current_quarter, 'year': current_year, 'frequency': 'quarterly', 'officeName': 'State Aggregated', 'officeCode': ''}
        for k in NUMERIC_KEYS:
            state_qpr[k] = state_totals.get(k, 0)
            state_qpr_items.append((k, state_qpr[k]))
    except Exception:
        state_qpr = None

    return render(request, 'qpr/manager_report.html', {
        'manager_data': manager_data,
        'state_totals': state_totals,
        'state_qpr': state_qpr,
        'state_qpr_items': state_qpr_items,
        'frozen_count': frozen_count,
        'total_hods': total_hods,
        'current_lang': request.session.get('lang', 'en'),
    })


@login_required
def manager_state_qpr(request):
    """Render aggregated State QPR (from frozen HOD snapshots) as a single QPR view.

    Uses the same `report_detail.html` but passes `initial_qpr_json` containing
    the aggregated QPR so the client-side loader will render it directly.
    """
    if not (user_has_role(request.user, ['manager', 'admin']) or request.user.is_superuser):
        return redirect('/')

    manager_profile = getattr(request.user, 'profile', None)
    office_code = getattr(manager_profile, 'office_code', None)

    if not office_code:
        messages.error(request, 'Manager office not found')
        return redirect('manager_report')

    hod_profiles = UserProfile.objects.filter(office_code=office_code, roles__name__iexact='hod').select_related('user')
    hod_user_ids = [getattr(h.user, 'id', None) for h in hod_profiles if getattr(h.user, 'id', None) is not None]

    current_quarter = get_current_quarter()
    current_year = get_current_year_label()

    # Aggregate from frozen HOD snapshots for current quarter/year
    state_totals = {k: 0 for k in NUMERIC_KEYS}
    # Extras for sections 9-11 (non-numeric fields)
    s9_sub_committees_total = 0
    s9_meetings_total = 0
    s9_agenda_any = False
    s9_dates = []
    s10_dates = []
    s12_1_texts = []
    s12_2_texts = []
    s12_3_texts = []
    if hod_user_ids:
        hod_qprs = QPRRecord.objects.filter(user_id__in=hod_user_ids, frequency__iexact='quarterly', is_quarterly_frozen=True, quarter=current_quarter, year=current_year)
        for rec in hod_qprs:
            try:
                d = serialize_qpr_record(rec)
            except Exception:
                continue
            for k in NUMERIC_KEYS:
                try:
                    v = d.get(k)
                    if v is None or v == '':
                        continue
                    state_totals[k] += int(v)
                except Exception:
                    continue
            # collect section 9/10/11 fields when present
            try:
                sc = d.get('s9_sub_committees')
                if sc not in (None, ''):
                    try:
                        s9_sub_committees_total += int(sc)
                    except Exception:
                        pass
                sm = d.get('s9_meetings_count')
                if sm not in (None, ''):
                    try:
                        s9_meetings_total += int(sm)
                    except Exception:
                        pass
                agenda = d.get('s9_agenda_hindi')
                if agenda not in (None, ''):
                    if str(agenda).strip().lower() in ('yes','y','true','1','हाँ','हां'):
                        s9_agenda_any = True
                if d.get('s9_date'):
                    s9_dates.append(str(d.get('s9_date')))
                if d.get('s10_date'):
                    s10_dates.append(str(d.get('s10_date')))
                if d.get('s12_1'):
                    s12_1_texts.append(str(d.get('s12_1')).strip())
                if d.get('s12_2'):
                    s12_2_texts.append(str(d.get('s12_2')).strip())
                if d.get('s12_3'):
                    s12_3_texts.append(str(d.get('s12_3')).strip())
            except Exception:
                pass

    # Include ManagerQPR and AdminQPR records for this financial year/quarter
    try:
        from .models import ManagerQPR, AdminQPR
        mgr_qprs = ManagerQPR.objects.filter(quarter=current_quarter, financial_year=current_year, user__profile__office_code=office_code)
        for mq in mgr_qprs:
            try:
                mvals = _serialize_managerqpr(mq)
                for k in NUMERIC_KEYS:
                    try:
                        state_totals[k] += int(mvals.get(k, 0) or 0)
                    except Exception:
                        continue
            except Exception:
                continue
            # manager qpr may have section9/10/11-like attrs; include if present
            try:
                sc = getattr(mq, 's9_sub_committees', None) or getattr(mq, 's9_sub_committees_count', None)
                if sc not in (None, ''):
                    try:
                        s9_sub_committees_total += int(sc)
                    except Exception:
                        pass
                sm = getattr(mq, 's9_meetings_count', None) or getattr(mq, 's9_meetings_organized', None)
                if sm not in (None, ''):
                    try:
                        s9_meetings_total += int(sm)
                    except Exception:
                        pass
                agenda = getattr(mq, 's9_agenda_hindi', None)
                if agenda not in (None, ''):
                    if str(agenda).strip().lower() in ('yes','y','true','1','हाँ','हां'):
                        s9_agenda_any = True
                if getattr(mq, 's9_meeting_date', None):
                    s9_dates.append(str(getattr(mq, 's9_meeting_date')))
                if getattr(mq, 's10_meeting_date', None):
                    s10_dates.append(str(getattr(mq, 's10_meeting_date')))
                if getattr(mq, 's11_innovative_work', None):
                    s12_1_texts.append(str(getattr(mq, 's11_innovative_work')).strip())
                if getattr(mq, 's11_special_events', None):
                    s12_2_texts.append(str(getattr(mq, 's11_special_events')).strip())
                if getattr(mq, 's11_hindi_medium_works', None):
                    s12_3_texts.append(str(getattr(mq, 's11_hindi_medium_works')).strip())
            except Exception:
                pass

        adm_qprs = AdminQPR.objects.filter(quarter=current_quarter, financial_year=current_year, user__profile__office_code=office_code)
        for aq in adm_qprs:
            try:
                avals = _serialize_adminqpr(aq)
                for k in NUMERIC_KEYS:
                    try:
                        state_totals[k] += int(avals.get(k, 0) or 0)
                    except Exception:
                        continue
            except Exception:
                continue
            # admin qpr fields
            try:
                sc = getattr(aq, 'a_s9_sub_committees', None) or getattr(aq, 's9_sub_committees', None)
                if sc not in (None, ''):
                    try:
                        s9_sub_committees_total += int(sc)
                    except Exception:
                        pass
                sm = getattr(aq, 'a_s9_meetings_count', None) or getattr(aq, 's9_meetings_count', None) or getattr(aq, 'a_s9_meetings_organized', None)
                if sm not in (None, ''):
                    try:
                        s9_meetings_total += int(sm)
                    except Exception:
                        pass
                agenda = getattr(aq, 'a_s9_agenda_hindi', None) or getattr(aq, 's9_agenda_hindi', None)
                if agenda not in (None, ''):
                    if str(agenda).strip().lower() in ('yes','y','true','1','हाँ','हां'):
                        s9_agenda_any = True
                if getattr(aq, 'a_s9_date', None) or getattr(aq, 's9_date', None) or getattr(aq, 'a_s9_meeting_date', None):
                    s9_dates.append(str(getattr(aq, 'a_s9_date', None) or getattr(aq, 's9_date', None) or getattr(aq, 'a_s9_meeting_date', None)))
                if getattr(aq, 'a_s10_date', None) or getattr(aq, 's10_date', None) or getattr(aq, 'a_s10_meeting_date', None):
                    s10_dates.append(str(getattr(aq, 'a_s10_date', None) or getattr(aq, 's10_date', None) or getattr(aq, 'a_s10_meeting_date', None)))
                if getattr(aq, 'a_s11_innovative_work', None) or getattr(aq, 's11_innovative_work', None):
                    s12_1_texts.append(str(getattr(aq, 'a_s11_innovative_work', None) or getattr(aq, 's11_innovative_work', None)).strip())
                if getattr(aq, 'a_s11_special_events', None) or getattr(aq, 's11_special_events', None):
                    s12_2_texts.append(str(getattr(aq, 'a_s11_special_events', None) or getattr(aq, 's11_special_events', None)).strip())
                if getattr(aq, 'a_s11_hindi_medium_works', None) or getattr(aq, 's11_hindi_medium_works', None):
                    s12_3_texts.append(str(getattr(aq, 'a_s11_hindi_medium_works', None) or getattr(aq, 's11_hindi_medium_works', None)).strip())
            except Exception:
                pass
    except Exception:
        pass

    state_qpr = {
        'quarter': current_quarter,
        'year': current_year,
        'frequency': 'quarterly',
        'officeName': 'State Aggregated',
        'officeCode': '',
    }
    for k in NUMERIC_KEYS:
        state_qpr[k] = state_totals.get(k, 0)

    # Attach aggregated Section 9-11 fields
    state_qpr['s9_sub_committees'] = s9_sub_committees_total or 0
    state_qpr['s9_meetings_count'] = s9_meetings_total or 0
    state_qpr['s9_agenda_hindi'] = 'Yes' if s9_agenda_any else 'No'
    state_qpr['s9_date'] = s9_dates[0] if s9_dates else ''
    state_qpr['s10_date'] = s10_dates[0] if s10_dates else ''
    state_qpr['s12_1'] = ' | '.join([t for t in s12_1_texts if t]) if s12_1_texts else ''
    state_qpr['s12_2'] = ' | '.join([t for t in s12_2_texts if t]) if s12_2_texts else ''
    state_qpr['s12_3'] = ' | '.join([t for t in s12_3_texts if t]) if s12_3_texts else ''

    try:
        initial_qpr_json = json.dumps(state_qpr)
    except Exception:
        initial_qpr_json = None

    return render(request, 'qpr/report_detail.html', {
        'initial_qpr_json': initial_qpr_json,
    })


@login_required
def manager_report_detail(request, year, quarter):
    """Show list of users who submitted for given quarter and year, grouped by HOD."""
    if not (user_has_role(request.user, ['manager', 'admin']) or request.user.is_superuser):
        return redirect('/')

    manager_office = getattr(request.user.profile, 'office_code', None)
    if not manager_office:
        first = QPRRecord.objects.filter(user=request.user).first()
        manager_office = first.officeCode if first else None

    if not manager_office:
        return redirect('manager_report')

    users_qs = UserProfile.objects.filter(office_code=manager_office).filter(
        Q(roles__name='user') | Q(user__roles__name='user')
    ).select_related('user').distinct().order_by('hod_name', 'name')
    total_users = users_qs.count()

    # Normalize year: if year contains '2025' treat as empty string (DB stores year='')
    normalized_year = year

    # Find submitted QPRs matching the office, normalized year, and quarter
    submitted = QPRRecord.objects.filter(officeCode=manager_office, year=normalized_year, quarter=quarter, is_submitted=True)
    submitted_users_count = submitted.values('user').distinct().count()
    total_users = users_qs.count()

    # Flag indicating whether all users have submitted (used to control view behavior)
    all_submitted = submitted_users_count >= total_users

    # Build grouping by HOD and include every employee; mark submitted status
    grouped = {}

    # Map submitted records by user id
    submitted_map = {r.user.id: r for r in submitted if r.user is not None}

    for up in users_qs:
        user = up.user
        rec = submitted_map.get(getattr(user, 'id', None))
        hod = up.hod_name or 'Unassigned'
        if hod not in grouped:
            grouped[hod] = []
        # Provide both English and Hindi names so templates can choose by language
        try:
            emp_hname = ''
            emp_obj = getattr(up, 'employee', None)
            if emp_obj is not None:
                emp_hname = getattr(emp_obj, 'hname', '') or ''
        except Exception:
            emp_hname = ''

        grouped[hod].append({
            'name': up.name or getattr(user, 'username', ''),
            'hname': emp_hname or (up.name or getattr(user, 'username', '')),
            'empcode': up.employee_code,
            'email': up.email or '',
            'submitted': bool(rec),
            'submitted_at': rec.updated_at if rec else None,
            'qpr_record_id': rec.id if rec else None,
        })

    context = {
        'year': year if year else '2025-2026', # <--- Add the fallback here too!
        'quarter': quarter,
        'office_code': manager_office,
        'grouped': grouped,
        'all_submitted': all_submitted,
        'submitted_users_count': submitted_users_count,
        'total_users': total_users,
        'current_lang': request.session.get('lang', 'en'),
    }
    
    return render(request, 'qpr/manager_report_detail.html', context)


@login_required
def qpr_certificate(request, record_id):
    """Render a printable certificate for a QPR record (open in new tab and print)."""
    try:
        rec = QPRRecord.objects.get(pk=record_id)
    except QPRRecord.DoesNotExist:
        return redirect('manager_report')

    # Permission: only manager/admin or same office manager can view
    if not (user_has_role(request.user, ['manager', 'admin']) or request.user.is_superuser):
        return redirect('/')

    # If user is manager, ensure office matches their profile where possible
    mgr_office = getattr(request.user.profile, 'office_code', None)
    if user_role(request.user) == 'manager' and mgr_office and mgr_office != rec.officeCode:
        return redirect('manager_report')

    context = {
        'record': rec,
    }
    return render(request, 'qpr/certificate.html', context)


@login_required
def manager_report_detail_by_record(request, record_id):
    try:
        rec = QPRRecord.objects.get(pk=record_id)
    except QPRRecord.DoesNotExist:
        return redirect('manager_report')
    
    safe_year = rec.year if rec.year else '2025-2026'
    # Delegate to manager_report_detail using the record's year and quarter
    return manager_report_detail(request, rec.year, rec.quarter)


@login_required
def certificate_form_view(request, record_id):
    """Auto-populate certificate data and redirect to display"""
    try:
        record = QPRRecord.objects.get(pk=record_id)
    except QPRRecord.DoesNotExist:
        return redirect('manager_report')

    # Only manager/admin can view certificate for submitted records
    if not (user_has_role(request.user, ['manager', 'admin']) or request.user.is_superuser):
        return redirect('/')

    # Get or create certificate data with auto-populated values from record
    cert_data, created = CertificateData.objects.get_or_create(
        qpr_record=record,
        defaults={
            'financial_year': record.year if record.year else '2025-2026',
            'quarter_ending': record.quarter if record.quarter else '',
        }
    )
    
    # Redirect to Part II form view (manager fills Part II here)
    return redirect('certificate_part2_list')





@login_required
def certificate_display_view(request, record_id):
    """Display the certificate in Enclosure format"""
    try:
        record = QPRRecord.objects.get(pk=record_id)
    except QPRRecord.DoesNotExist:
        return redirect('manager_report')

    # Only manager/admin can view certificate
    if not (user_has_role(request.user, ['manager', 'admin']) or request.user.is_superuser):
        return redirect('/')

    # Get certificate data
    cert_data = CertificateData.objects.filter(qpr_record=record).first()
    if not cert_data:
        # Redirect to form if certificate data doesn't exist
        return redirect('certificate_form', record_id=record.id)

    context = {
        'record': record,
        'cert_data': cert_data,
    }
    return render(request, 'qpr/certificate_display.html', context)


@login_required
def certificate_part2_view(request):
    """Standalone Certificate Part II view.
    Accessible only to users with the 'manager' role. Not linked to any QPR records.
    GET: render the `certificate_part2.html` form.
    Other methods: return 405.
    """
    # Enforce manager-only access
    if not user_has_role(request.user, 'manager'):
        return HttpResponseForbidden('Forbidden')

    if request.method != 'GET':
        return HttpResponseNotAllowed(['GET'])

    # Render the standalone form (no QPR/record context)
    return render(request, 'qpr/certificate_part2.html', {})
    if 'computer_training_trained' in payload:
        part2.computer_training_trained = int(payload.get('computer_training_trained') or 0)
    if 'computer_training_working' in payload:
        part2.computer_training_working = int(payload.get('computer_training_working') or 0)
    if 'total_computers' in payload:
        part2.total_computers = int(payload.get('total_computers') or 0)
    if 'hindi_enabled_computers' in payload:
        part2.hindi_enabled_computers = int(payload.get('hindi_enabled_computers') or 0)
    if 'officials_issued_rule_8_4_orders' in payload:
        part2.officials_issued_rule_8_4_orders = int(payload.get('officials_issued_rule_8_4_orders') or 0)
    if 'training_total_duration_hours' in payload:
        part2.training_total_duration_hours = int(payload.get('training_total_duration_hours') or 0)
    if 'training_imparted_hindi' in payload:
        part2.training_imparted_hindi = int(payload.get('training_imparted_hindi') or 0)
    if 'training_imparted_english' in payload:
        part2.training_imparted_english = int(payload.get('training_imparted_english') or 0)
    if 'training_imparted_mixed' in payload:
        part2.training_imparted_mixed = int(payload.get('training_imparted_mixed') or 0)
    if 'sec8_total_sections' in payload:
        part2.sec8_total_sections = int(payload.get('sec8_total_sections') or 0)
    if 'sec8_inspected_sections' in payload:
        part2.sec8_inspected_sections = int(payload.get('sec8_inspected_sections') or 0)
    if 'sec8_total_sub_offices' in payload:
        part2.sec8_total_sub_offices = int(payload.get('sec8_total_sub_offices') or 0)
    if 'sec8_inspected_sub_offices' in payload:
        part2.sec8_inspected_sub_offices = int(payload.get('sec8_inspected_sub_offices') or 0)
    if 'magazines_total' in payload:
        part2.magazines_total = int(payload.get('magazines_total') or 0)
    if 'magazines_hindi' in payload:
        part2.magazines_hindi = int(payload.get('magazines_hindi') or 0)
    if 'magazines_english' in payload:
        part2.magazines_english = int(payload.get('magazines_english') or 0)
    if 'expenditure_total_books' in payload:
        try:
            part2.expenditure_total_books = Decimal(str(payload.get('expenditure_total_books') or 0))
        except Exception:
            part2.expenditure_total_books = Decimal('0.00')
    if 'expenditure_hindi_books' in payload:
        try:
            part2.expenditure_hindi_books = Decimal(str(payload.get('expenditure_hindi_books') or 0))
        except Exception:
            part2.expenditure_hindi_books = Decimal('0.00')
    # Dates and text
    if 'hindi_event_start_date' in payload:
        part2.hindi_event_start_date = payload.get('hindi_event_start_date') or None
    if 'hindi_event_end_date' in payload:
        part2.hindi_event_end_date = payload.get('hindi_event_end_date') or None
    if 'seminar_date' in payload:
        part2.seminar_date = payload.get('seminar_date') or None
    if 'seminar_subject' in payload:
        part2.seminar_subject = payload.get('seminar_subject') or ''
    if 'other_activities_date' in payload:
        part2.other_activities_date = payload.get('other_activities_date') or None
    if 'other_activities_subject' in payload:
        part2.other_activities_subject = payload.get('other_activities_subject') or ''

    # Other scalar fields can be mapped similarly if included in payload
    part2.save()

    # If action is 'submit' mark submitted and lock editing
    if action == 'submit':
        from django.utils import timezone
        part2.is_submitted = True
        part2.submitted_at = timezone.now()
        part2.submitted_by = request.user
        part2.save()
        # Lock editing until manager unlocks via manager table
        record.is_editing_allowed = False
        record.save(update_fields=['is_editing_allowed'])

    # Replace staff_knowledge rows
    if 'staff_knowledge' in payload:
        part2.staff_knowledge.all().delete()
        for item in payload.get('staff_knowledge', []):
            StaffHindiKnowledge.objects.create(
                report=part2,
                category=item.get('category', ''),
                officers_count=int(item.get('officers_count') or 0),
                employees_count=int(item.get('employees_count') or 0),
                total_count=int(item.get('total_count') or 0)
            )

    # Typing/Stenography rows
    if 'typing_knowledge' in payload:
        part2.typing_knowledge.all().delete()
        for item in payload.get('typing_knowledge', []):
            from website.models import TypingStenographyKnowledge
            TypingStenographyKnowledge.objects.create(
                report=part2,
                category=item.get('category', ''),
                total_no=int(item.get('total_no') or 0),
                trained_in_hindi=int(item.get('trained_in_hindi') or 0),
                work_in_hindi=int(item.get('work_in_hindi') or 0),
                yet_to_be_trained=int(item.get('yet_to_be_trained') or 0)
            )

    # Translation knowledge
    if 'translation_knowledge' in payload:
        part2.translation_knowledge.all().delete()
        for item in payload.get('translation_knowledge', []):
            from website.models import TranslationKnowledge
            TranslationKnowledge.objects.create(
                report=part2,
                category=item.get('category', ''),
                officers_count=int(item.get('officers_count') or 0),
                employees_count=int(item.get('employees_count') or 0),
                total_count=int(item.get('total_count') or 0)
            )

    # Code/manuals
    if 'code_manuals' in payload:
        part2.codes_manuals.all().delete()
        for item in payload.get('code_manuals', []):
            from website.models import CodeManualStandardForms
            CodeManualStandardForms.objects.create(
                report=part2,
                category=item.get('category', ''),
                total_no=int(item.get('total_no') or 0),
                bilingual_no=int(item.get('bilingual_no') or 0)
            )

    # Officers work (sections 11 & 12)
    if 'officers_work' in payload:
        part2.officers_work.all().delete()
        for item in payload.get('officers_work', []):
            from website.models import OfficersWorkInHindi
            OfficersWorkInHindi.objects.create(
                report=part2,
                level=item.get('level', ''),
                total_officers=int(item.get('total_officers') or 0),
                knowledge_of_hindi=int(item.get('knowledge_of_hindi') or 0),
                not_doing=int(item.get('not_doing') or 0),
                doing_upto_25=int(item.get('doing_upto_25') or 0),
                doing_26_to_50=int(item.get('doing_26_to_50') or 0),
                doing_51_to_75=int(item.get('doing_51_to_75') or 0),
                doing_more_76=int(item.get('doing_more_76') or 0),
                doing_cent_percent=int(item.get('doing_cent_percent') or 0)
            )

    # Websites
    if 'websites' in payload:
        part2.websites.all().delete()
        for w in payload.get('websites', []):
            from website.models import WebsiteDetail
            if not w.get('url'): continue
            WebsiteDetail.objects.create(
                report=part2,
                url=w.get('url'),
                status=w.get('status') or ''
            )

    # Chairperson / contact
    if 'chairperson' in payload:
        ch = payload.get('chairperson') or {}
        part2.chairperson_name = ch.get('name') or ''
        part2.chairperson_designation = ch.get('designation') or ''
        part2.chairperson_phone = ch.get('phone') or ''
        part2.chairperson_fax = ch.get('fax') or ''
        part2.chairperson_email = ch.get('email') or ''
        part2.save()

    # Replace hindi_posts rows
    if 'hindi_posts' in payload:
        part2.hindi_posts.all().delete()
        for p in payload.get('hindi_posts', []):
            if not p.get('designation'):
                continue
            HindiPost.objects.create(
                report=part2,
                designation=p.get('designation'),
                sanctioned=int(p.get('sanctioned') or 0),
                vacant=int(p.get('vacant') or 0)
            )

    return JsonResponse({'success': True, 'message': 'Part II saved', 'edit_count': record.cert_edit_count})


# ==================== NEW CERTIFICATE PART 2 VIEWS (NO API CALLS) ====================

@login_required
def certificate_part2_list(request):
    """List all submitted certificates for the logged-in manager"""
    lang = request.session.get('lang', 'en')
    
    # Only managers can access
    if not user_has_role(request.user, 'manager'):
        return HttpResponseForbidden('Only managers can access certificates.')
    
    # Get all certificates for this user
    certificates = QPRPartTwo.objects.filter(user=request.user).order_by('-created_at')
    
    context = {
        'certificates': certificates,
        'current_lang': lang
    }
    return render(request, 'qpr/certificate_part2_list.html', context)


@login_required
def certificate_part2_new(request):
    """Create a new certificate - redirects to form with quarter/year selection"""
    lang = request.session.get('lang', 'en')
    
    # Only managers can access
    if not user_has_role(request.user, 'manager'):
        return HttpResponseForbidden('Only managers can access certificates.')
    
    if request.method == 'POST':
        quarter = request.POST.get('quarter', '').strip()
        year = request.POST.get('year', '').strip()
        
        if not quarter or not year:
            messages.error(request, 'Quarter and Year are required.')
            return redirect('certificate_part2_list')
        
        # Check if certificate already exists for this quarter/year
        existing = QPRPartTwo.objects.filter(
            user=request.user,
            quarter=quarter,
            year=year
        ).first()
        
        if existing:
            # Redirect to edit existing certificate
            return redirect('certificate_part2_edit', pk=existing.id)
        
        # Create new certificate
        certificate = QPRPartTwo.objects.create(
            user=request.user,
            quarter=quarter,
            year=year,
            financial_year=year
        )
        
        # Create default rows for sections 11 & 12 (Officers Work)
        from website.models import OfficersWorkInHindi
        OfficersWorkInHindi.objects.create(
            report=certificate,
            level='ds_and_above',
            total_officers=0,
            knowledge_of_hindi=0,
            not_doing=0,
            doing_upto_25=0,
            doing_26_to_50=0,
            doing_51_to_75=0,
            doing_more_76=0,
            doing_cent_percent=0
        )
        OfficersWorkInHindi.objects.create(
            report=certificate,
            level='below_ds',
            total_officers=0,
            knowledge_of_hindi=0,
            not_doing=0,
            doing_upto_25=0,
            doing_26_to_50=0,
            doing_51_to_75=0,
            doing_more_76=0,
            doing_cent_percent=0
        )
        
        return redirect('certificate_part2_form', pk=certificate.id)
    
    # Show quarter/year selection form
    quarters = ['Q1', 'Q2', 'Q3', 'Q4']
    years = ['2024-25', '2025-26']  # Adjust as needed
    
    context = {
        'quarters': quarters,
        'years': years,
        'current_lang': lang
    }
    return render(request, 'qpr/certificate_part2_select_quarter.html', context)


@login_required
def certificate_part2_form(request, pk):
    """Form to edit/view certificate"""
    lang = request.session.get('lang', 'en')
    
    # Get the certificate
    try:
        certificate = QPRPartTwo.objects.get(pk=pk, user=request.user)
    except QPRPartTwo.DoesNotExist:
        messages.error(request, 'Certificate not found.')
        return redirect('certificate_part2_list')
    
    # Only manager can edit their own certificates
    if certificate.user != request.user:
        return HttpResponseForbidden('You cannot edit this certificate.')
    
    # Ensure default rows exist for sections 11 & 12 (Officers Work)
    from website.models import OfficersWorkInHindi
    OfficersWorkInHindi.objects.get_or_create(
        report=certificate,
        level='ds_and_above',
        defaults={
            'total_officers': 0,
            'knowledge_of_hindi': 0,
            'not_doing': 0,
            'doing_upto_25': 0,
            'doing_26_to_50': 0,
            'doing_51_to_75': 0,
            'doing_more_76': 0,
            'doing_cent_percent': 0
        }
    )
    OfficersWorkInHindi.objects.get_or_create(
        report=certificate,
        level='below_ds',
        defaults={
            'total_officers': 0,
            'knowledge_of_hindi': 0,
            'not_doing': 0,
            'doing_upto_25': 0,
            'doing_26_to_50': 0,
            'doing_51_to_75': 0,
            'doing_more_76': 0,
            'doing_cent_percent': 0
        }
    )
    
    if request.method == 'POST':
        action = request.POST.get('action', 'save')
        
        # Update certificate fields from POST
        certificate.is_notified_rule_10_4 = request.POST.get('is_notified_rule_10_4') == 'true'
        certificate.total_sub_offices = int(request.POST.get('total_sub_offices', 0) or 0)
        certificate.notified_sub_offices = int(request.POST.get('notified_sub_offices', 0) or 0)
        certificate.computer_training_total_staff = int(request.POST.get('computer_training_total_staff', 0) or 0)
        certificate.computer_training_trained = int(request.POST.get('computer_training_trained', 0) or 0)
        certificate.computer_training_working = int(request.POST.get('computer_training_working', 0) or 0)
        certificate.total_computers = int(request.POST.get('total_computers', 0) or 0)
        certificate.hindi_enabled_computers = int(request.POST.get('hindi_enabled_computers', 0) or 0)
        certificate.hindi_work_percentage = float(request.POST.get('hindi_work_percentage', 0) or 0)
        certificate.officials_issued_rule_8_4_orders = int(request.POST.get('officials_issued_rule_8_4_orders', 0) or 0)
        certificate.training_total_duration_hours = int(request.POST.get('training_total_duration_hours', 0) or 0)
        certificate.training_imparted_hindi = int(request.POST.get('training_imparted_hindi', 0) or 0)
        certificate.training_imparted_english = int(request.POST.get('training_imparted_english', 0) or 0)
        certificate.training_imparted_mixed = int(request.POST.get('training_imparted_mixed', 0) or 0)
        certificate.sec8_total_sections = int(request.POST.get('sec8_total_sections', 0) or 0)
        certificate.sec8_inspected_sections = int(request.POST.get('sec8_inspected_sections', 0) or 0)
        certificate.sec8_total_sub_offices = int(request.POST.get('sec8_total_sub_offices', 0) or 0)
        certificate.sec8_inspected_sub_offices = int(request.POST.get('sec8_inspected_sub_offices', 0) or 0)
        certificate.magazines_total = int(request.POST.get('magazines_total', 0) or 0)
        certificate.magazines_hindi = int(request.POST.get('magazines_hindi', 0) or 0)
        certificate.magazines_english = int(request.POST.get('magazines_english', 0) or 0)
        certificate.expenditure_total_books = float(request.POST.get('expenditure_total_books', 0) or 0)
        certificate.expenditure_hindi_books = float(request.POST.get('expenditure_hindi_books', 0) or 0)
        
        # Date fields
        hindi_event_start = request.POST.get('hindi_event_start_date')
        if hindi_event_start:
            certificate.hindi_event_start_date = hindi_event_start
        hindi_event_end = request.POST.get('hindi_event_end_date')
        if hindi_event_end:
            certificate.hindi_event_end_date = hindi_event_end
        seminar_date = request.POST.get('seminar_date')
        if seminar_date:
            certificate.seminar_date = seminar_date
        
        certificate.seminar_subject = request.POST.get('seminar_subject', '')
        
        other_act_date = request.POST.get('other_activities_date')
        if other_act_date:
            certificate.other_activities_date = other_act_date
        certificate.other_activities_subject = request.POST.get('other_activities_subject', '')
        
        # Chairperson info
        certificate.chairperson_name = request.POST.get('chairperson_name', '')
        certificate.chairperson_designation = request.POST.get('chairperson_designation', '')
        certificate.chairperson_phone = request.POST.get('chairperson_phone', '')
        certificate.chairperson_fax = request.POST.get('chairperson_fax', '')
        certificate.chairperson_email = request.POST.get('chairperson_email', '')

        # Section 2(ii) - Typing/Stenography

        categories = ['stenographer', 'typist_clerk', 'tax_postal']

        for cat in categories:
            TypingStenographyKnowledge.objects.update_or_create(
        report=certificate,
        category=cat,
        defaults={
            'total_no': int(request.POST.get(f'typing_total_{cat}', 0) or 0),
            'trained_in_hindi': int(request.POST.get(f'typing_trained_{cat}', 0) or 0),
            'work_in_hindi': int(request.POST.get(f'typing_working_{cat}', 0) or 0),
            'yet_to_be_trained': int(request.POST.get(f'typing_yet_{cat}', 0) or 0),
            }
        )
        
        # Section 2(iii) - Translation

        TranslationKnowledge.objects.update_or_create(
            report=certificate,
            category='engaged',
            defaults={
                'officers_count': int(request.POST.get('translation_officers_engaged', 0) or 0),
                'employees_count': int(request.POST.get('translation_employees_engaged', 0) or 0),
                'total_count': int(request.POST.get('translation_total_engaged', 0) or 0),
                }
            )

        TranslationKnowledge.objects.update_or_create(
            report=certificate,
            category='trained',
            defaults={
                'officers_count': int(request.POST.get('translation_officers_trained', 0) or 0),
                'employees_count': int(request.POST.get('translation_employees_trained', 0) or 0),
                'total_count': int(request.POST.get('translation_total_trained', 0) or 0),
                }
            )

        TranslationKnowledge.objects.update_or_create(
            report=certificate,
            category='yet_to_be_trained',
            defaults={
            'officers_count': int(request.POST.get('translation_officers_yet', 0) or 0),
            'employees_count': int(request.POST.get('translation_employees_yet', 0) or 0),
            'total_count': int(request.POST.get('translation_total_yet', 0) or 0),
                }
            )
        
        # Section 2(i) - Staff Data

        StaffHindiKnowledge.objects.update_or_create(
            report=certificate,
            category='total',
            defaults={
            'officers_total': int(request.POST.get('staff_total_officers', 0) or 0),
            'employees_total': int(request.POST.get('staff_total_employees', 0) or 0),
            'total_count': int(request.POST.get('staff_total_total', 0) or 0),
            }
        )

        StaffHindiKnowledge.objects.update_or_create(
            report=certificate,
            category='secretarial',
            defaults={
                'officers_total': int(request.POST.get('staff_secretarial_officers', 0) or 0),
                'employees_total': int(request.POST.get('staff_secretarial_employees', 0) or 0),
                'total_count': int(request.POST.get('staff_secretarial_total', 0) or 0),
                }
        )

        StaffHindiKnowledge.objects.update_or_create(
            report=certificate,
            category='knowledge',
            defaults={
                'officers_working': int(request.POST.get('staff_knowledge_officers_working', 0) or 0),
                'officers_proficient': int(request.POST.get('staff_knowledge_officers_proficient', 0) or 0),
                'employees_working': int(request.POST.get('staff_knowledge_employees_working', 0) or 0),
                'employees_proficient': int(request.POST.get('staff_knowledge_employees_proficient', 0) or 0),
                'total_count': int(request.POST.get('staff_knowledge_total', 0) or 0),
            }
        )

        StaffHindiKnowledge.objects.update_or_create(
            report=certificate,
            category='being_trained',
            defaults={
                'officers_total': int(request.POST.get('staff_trained_officers', 0) or 0),
                'employees_total': int(request.POST.get('staff_trained_employees', 0) or 0),
                'total_count': int(request.POST.get('staff_trained_total', 0) or 0),
            }
        )

        StaffHindiKnowledge.objects.update_or_create(
            report=certificate,
            category='yet_to_be_trained',
            defaults={
                'officers_total': int(request.POST.get('staff_yet_officers', 0) or 0),
                'employees_total': int(request.POST.get('staff_yet_employees', 0) or 0),
                'total_count': int(request.POST.get('staff_yet_total', 0) or 0),
            }
        )

        # Section 5 - Codes/Manuals

        CodeManualStandardForms.objects.update_or_create(
            report=certificate,
            category='acts_rules',
            defaults={
            'total_no': int(request.POST.get('codes_total_acts_rules', 0) or 0),
            'bilingual_no': int(request.POST.get('codes_bilingual_acts_rules', 0) or 0),
            }
        )

        CodeManualStandardForms.objects.update_or_create(
            report=certificate,
            category='standard_forms',
            defaults={
            'total_no': int(request.POST.get('codes_total_standard_forms', 0) or 0),
            'bilingual_no': int(request.POST.get('codes_bilingual_standard_forms', 0) or 0),
            }
        )

        # Section 13 - Hindi Posts (MULTI ROW)

        post_names = request.POST.getlist('section13_post_name')
        hq_sanctioned_list = request.POST.getlist('section13_hq_sanctioned')
        hq_vacant_list = request.POST.getlist('section13_hq_vacant')
        sub_sanctioned_list = request.POST.getlist('section13_sub_sanctioned')
        sub_vacant_list = request.POST.getlist('section13_sub_vacant')

        # Delete old data
        certificate.hindi_posts.all().delete()

        for i in range(len(post_names)):
            if post_names[i]:  # skip empty rows
                HindiPost.objects.create(
                    report=certificate,
                    designation=post_names[i],
                    hq_sanctioned=int(hq_sanctioned_list[i] or 0),
                    hq_vacant=int(hq_vacant_list[i] or 0),
                    sub_sanctioned=int(sub_sanctioned_list[i] or 0),
                    sub_vacant=int(sub_vacant_list[i] or 0)
                )

# Section 14 - Websites

        urls = request.POST.getlist('section14_url')

        certificate.websites.all().delete()

        for i, url in enumerate(urls):
            if url:  # only create if URL is not empty
                status_key = f'section14_status_{i + 1}'
                status = request.POST.get(status_key, '')
                WebsiteDetail.objects.create(
                    report=certificate,
                    url=url,
                    status=status
                )
        
        # Section 11 - DS and Above

        OfficersWorkInHindi.objects.update_or_create(
            report=certificate,
            level='ds_and_above',
            defaults={
                'total_officers': int(request.POST.get('sec11_total', 0) or 0),
                'knowledge_of_hindi': int(request.POST.get('sec11_knowledge', 0) or 0),
                'not_doing': int(request.POST.get('sec11_not_doing', 0) or 0),
                'doing_upto_25': int(request.POST.get('sec11_0_25', 0) or 0),
                'doing_26_to_50': int(request.POST.get('sec11_26_50', 0) or 0),
                'doing_51_to_75': int(request.POST.get('sec11_51_75', 0) or 0),
                'doing_more_76': int(request.POST.get('sec11_76_99', 0) or 0),
                'doing_cent_percent': int(request.POST.get('sec11_100', 0) or 0),
            }
        )

# Section 12 - Below DS

        OfficersWorkInHindi.objects.update_or_create(
            report=certificate,
            level='below_ds',
            defaults={
                'total_officers': int(request.POST.get('sec12_total', 0) or 0),
                'knowledge_of_hindi': int(request.POST.get('sec12_knowledge', 0) or 0),
                'not_doing': int(request.POST.get('sec12_not_doing', 0) or 0),
                'doing_upto_25': int(request.POST.get('sec12_0_25', 0) or 0),
                'doing_26_to_50': int(request.POST.get('sec12_26_50', 0) or 0),
                'doing_51_to_75': int(request.POST.get('sec12_51_75', 0) or 0),
                'doing_more_76': int(request.POST.get('sec12_76_99', 0) or 0),
                'doing_cent_percent': int(request.POST.get('sec12_100', 0) or 0),
            }
        )

        if action == 'submit':
            # Mark as submitted
            if not certificate.quarter in ['Q1', 'Q2', 'Q3', 'Q4']:
                messages.error(request, 'Invalid quarter.')
                return render(request, 'qpr/certificate_part2_form.html', {
                    'part2': certificate,
                    'quarter': certificate.quarter,
                    'year': certificate.year,
                    'current_lang': lang
                })
            
            certificate.is_submitted = True
            certificate.submitted_by = request.user
            certificate.submitted_at = timezone.now()
            certificate.save()
            
            messages.success(request, translate_text('Certificate submitted successfully!', lang))
            return redirect('certificate_part2_list')
        else:
            # Save as draft
            certificate.save()
            messages.success(request, translate_text('Certificate saved as draft.', lang))
            return redirect('certificate_part2_form', pk=certificate.id)
    
    context = {
        'part2': certificate,
        'quarter': certificate.quarter,
        'year': certificate.year,
        'current_lang': lang
    }
    return render(request, 'qpr/certificate_part2_form.html', context)


@login_required
def certificate_part2_edit(request, pk):
    """Edit a certificate - same as form view"""
    return certificate_part2_form(request, pk)


@login_required
def certificate_part2_view(request, pk):
    """View a certificate (read-only)"""
    lang = request.session.get('lang', 'en')
    
    try:
        certificate = QPRPartTwo.objects.get(pk=pk, user=request.user)
    except QPRPartTwo.DoesNotExist:
        messages.error(request, 'Certificate not found.')
        return redirect('certificate_part2_list')
    
    context = {
        'certificate': certificate,
        'part2': certificate,
        'quarter': certificate.quarter,
        'year': certificate.year,
        'current_lang': lang,
        'readonly': True
    }
    return render(request, 'qpr/certificate_part2_form.html', context)


@login_required
def certificate_part2_print(request, pk):
    """Print a certificate - dedicated print page"""
    lang = request.session.get('lang', 'en')
    
    try:
        certificate = QPRPartTwo.objects.get(pk=pk, user=request.user)
    except QPRPartTwo.DoesNotExist:
        messages.error(request, 'Certificate not found.')
        return redirect('certificate_part2_list')
    
    # Get office name from user profile if available
    office_name = ""
    try:
        user_profile = request.user.profile
        office_name = user_profile.office_name if hasattr(user_profile, 'office_name') else ""
    except:
        pass
    
    context = {
        'certificate': certificate,
        'office_name': office_name,
        'part2': certificate,
        'current_lang': lang
    }
    return render(request, 'qpr/certificate_part2_comprehensive_print.html', context)


@login_required
def certificate_part2_delete(request, pk):
    """Delete a draft certificate"""
    lang = request.session.get('lang', 'en')
    
    try:
        certificate = QPRPartTwo.objects.get(pk=pk, user=request.user)
    except QPRPartTwo.DoesNotExist:
        messages.error(request, 'Certificate not found.')
        return redirect('certificate_part2_list')
    
    # Only draft certificates can be deleted
    if certificate.is_submitted:
        messages.error(request, 'Submitted certificates cannot be deleted.')
        return redirect('certificate_part2_list')
    
    if request.method == 'POST':
        certificate.delete()
        messages.success(request, translate_text('Certificate deleted.', lang))
        return redirect('certificate_part2_list')
    
    return redirect('certificate_part2_list')


@login_required
def manager_report_edit_view(request, record_id):
    """Unlock ONLY certificate Part-II for editing (max 2 times)"""

    try:

        if request.method != 'POST':
            return JsonResponse({'success': False, 'error': 'POST required'}, status=405)

        if not (user_has_role(request.user, ['manager', 'admin']) or request.user.is_superuser):
            return JsonResponse({'success': False, 'error': 'Permission denied'}, status=403)

        try:
            record = QPRRecord.objects.get(pk=record_id)
        except QPRRecord.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Record not found'}, status=404)

        # Verify manager office
        mgr_office = getattr(request.user.profile, 'office_code', None)
        if mgr_office != record.officeCode:
            return JsonResponse({'success': False, 'error': 'Unauthorized office'}, status=403)

        # LIMIT EDITS
        if record.cert_edit_count >= 2:
            return JsonResponse({
                'success': False,
                'error': 'Maximum edit attempts (2) reached'
            })

        # increase edit count
        record.cert_edit_count += 1
        record.is_editing_allowed = True
        record.save(update_fields=['cert_edit_count', 'is_editing_allowed'])

        # unlock ONLY the certificate
        part2 = getattr(record, 'part2', None)

        if part2:
            part2.is_submitted = False
            part2.save(update_fields=['is_submitted'])

        return JsonResponse({
            'success': True,
            'message': f'Certificate unlocked. Edit attempt {record.cert_edit_count}/2',
            'edit_count': record.cert_edit_count
        })

    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)
        
from gtts import gTTS
import os
from django.conf import settings

def generate_captcha_audio(text):
    # This creates the audio from the captcha text
    tts = gTTS(text=text, lang='en')
    filename = os.path.join(settings.MEDIA_ROOT, 'captcha_audio.mp3')
    tts.save(filename)
    return filename

@login_required
def print_all_qpr_reports(request, year, quarter):
    """Aggregates Part 1, Certificate, and Part 2 for all submitted employees for printing."""
    if not (user_has_role(request.user, ['manager', 'admin']) or request.user.is_superuser):
        return redirect('/')

    # Determine manager's office
    manager_office = getattr(request.user.profile, 'office_code', None)
    if not manager_office:
        first = QPRRecord.objects.filter(user=request.user).first()
        manager_office = first.officeCode if first else None

    if not manager_office:
        return redirect('manager_report')

    normalized_year = year
    
    # Fetch all submitted QPRs for this office, year, and quarter
    submitted_qprs = QPRRecord.objects.filter(
        officeCode=manager_office, 
        year=normalized_year, 
        quarter=quarter, 
        is_submitted=True
    ).select_related('user', 'part2', 'certificate_data').order_by('user__username')

    # Submitted QPRs count: suppressed debug output

    all_reports_data = []
    
    for record in submitted_qprs:
        # 1. Part 1 Data
        part1_data = serialize_qpr_record(record)
        
        # 2. Certificate Data
        cert_data = getattr(record, 'certificate_data', None)
        
        # 3. Part 2 Data
        part2 = getattr(record, 'part2', None)
        part2_data = {}
        if part2:
            part2_data = {
                'financial_year': part2.financial_year,
                'is_notified_rule_10_4': bool(part2.is_notified_rule_10_4),
                'total_sub_offices': part2.total_sub_offices,
                'notified_sub_offices': part2.notified_sub_offices,
                'staff_knowledge': list(part2.staff_knowledge.values('category', 'officers_count', 'employees_count', 'total_count')),
                'hindi_posts': list(part2.hindi_posts.values('designation', 'sanctioned', 'vacant')),
                'typing_knowledge': list(part2.typing_knowledge.values('category', 'total_no', 'trained_in_hindi', 'work_in_hindi', 'yet_to_be_trained')),
                'translation_knowledge': list(part2.translation_knowledge.values('category', 'officers_count', 'employees_count', 'total_count')),
                'code_manuals': list(part2.codes_manuals.values('category', 'total_no', 'bilingual_no')),
                'officers_work': list(part2.officers_work.values('level', 'total_officers', 'knowledge_of_hindi', 'not_doing', 'doing_upto_25', 'doing_26_to_50', 'doing_51_to_75', 'doing_more_76', 'doing_cent_percent')),
                'websites': list(part2.websites.values('url', 'status')),
                'chairperson': {
                    'name': part2.chairperson_name or '',
                    'designation': part2.chairperson_designation or '',
                    'phone': part2.chairperson_phone or '',
                    'fax': part2.chairperson_fax or '',
                    'email': part2.chairperson_email or ''
                }
            }

        all_reports_data.append({
            'record': record,
            'part1': part1_data,
            'cert': cert_data,
            'part2': part2,
            'part2_data': part2_data,
            'employee_name': record.user.get_full_name() or record.user.username,
            'emp_code': getattr(record.user.profile, 'employee_code', 'N/A') if hasattr(record.user, 'profile') else 'N/A'
        })

    context = {
        'year': year,
        'quarter': quarter,
        'reports': all_reports_data,
        'office_code': manager_office
    }
    writer = PdfWriter()
    temp_files = []

    # ---------- PART 1 : All Employee QPR Reports ----------
    for item in all_reports_data:

        html_string = render_to_string(
            "qpr/print_report.html",
            {"r": item['part1'], "record": item['record'], "current_lang": request.session.get("lang", "en")}
        )

        tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".pdf")
        HTML( string=html_string, base_url=request.build_absolute_uri("/")).write_pdf(tmp.name)
        reader = PdfReader(tmp.name)
        for page in reader.pages:
            writer.add_page(page)
        temp_files.append(tmp.name)


    # ---------- PART 2 : Manager Form ----------
    manager_item = all_reports_data[0] if all_reports_data else None

    if manager_item and manager_item['part2']:

        html_string = render_to_string(
            "qpr/print_certificate_part2.html",
            {
                "record": manager_item['record'],
                "part2": manager_item['part2'],
                "part2_data": manager_item['part2_data'],
                "current_lang": request.session.get("lang", "en")
            }
        )


        tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".pdf")
        HTML( string=html_string, base_url=request.build_absolute_uri("/")).write_pdf(tmp.name)
        reader = PdfReader(tmp.name)
        for page in reader.pages:
            writer.add_page(page)
        temp_files.append(tmp.name)


    # ---------- PART 3 : Final Certificate ----------
    if manager_item and manager_item['cert']:

        html_string = render_to_string(
            "qpr/certificate.html",
            {
                "record": manager_item['record'],
                "cert_data": manager_item['cert'],
                "current_lang": request.session.get("lang", "en")
            }
        )

 
        tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".pdf")
        HTML( string=html_string, base_url=request.build_absolute_uri("/")).write_pdf(tmp.name)
        reader = PdfReader(tmp.name)
        for page in reader.pages:
            writer.add_page(page)
        temp_files.append(tmp.name)


    # ---------- FINAL MERGED PDF ----------
    # ---------- FINAL PDF ----------
    output = tempfile.NamedTemporaryFile(delete=False, suffix=".pdf")

    with open(output.name, "wb") as f:
        writer.write(f)

    pdf_file = open(output.name, "rb")

    response = FileResponse(
        pdf_file,
        content_type="application/pdf"
    )

    response["Content-Disposition"] = f'inline; filename="QPR_{quarter}_{year}.pdf"'

    return response

# Debug helper: returns current user/session info (useful to verify AJAX session & roles)
def debug_whoami(request):
    try:
        user = request.user
        data = {
            'is_authenticated': user.is_authenticated,
            'username': user.username if user and user.is_authenticated else None,
            'roles': user_get_all_roles(user) if user and user.is_authenticated else [],
            'profile_office_code': getattr(getattr(user, 'profile', None), 'office_code', None),
            'session_key': request.session.session_key,
            'method': request.method,
        }
        return JsonResponse({'success': True, 'whoami': data})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})
    
# In views.py
def process_user_approval(request, profile_id, action):
    if not user_has_role(request.user, ['hod', 'admin']):
        messages.error(request, "Unauthorized action.")
        return redirect('dashboard')

    target_profile = get_object_or_404(UserProfile, id=profile_id)
    
    if action == 'approve':
        from .models import Employee
        # Fetch the master record to sync the correct English Name (ename)
        master_record = Employee.objects.filter(empcode=target_profile.employee_code).first()
        
        if master_record:
            # Sync the name to the profile so it shows on the dashboard
            target_profile.name = master_record.ename
            target_profile.employee = master_record
            
            # Sync to the Django User object so {{ user.first_name }} works
            target_user = target_profile.user
            target_user.first_name = master_record.ename
            target_user.save()
        
        target_profile.approval_status = 'approved'
        target_profile.save()
        
        send_system_email(target_profile.user, request, 'accepted_alert') 
        messages.success(request, f"User {target_profile.employee_code} approved.")
        
    elif action == 'reject':
        target_profile.approval_status = 'rejected'
        target_profile.save()
        send_system_email(target_profile.user, request, 'rejected_alert')
        messages.warning(request, f"User {target_profile.employee_code} rejected.")

    return redirect('qpr_hod_dashboard')
